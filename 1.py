import logging
import sqlite3
import re
import requests
import nest_asyncio
import random
import os
import traceback
from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import PlainTextResponse, JSONResponse
from pydantic import BaseModel
import aiohttp
import asyncio
import httpx
import sys
import uvicorn
from datetime import datetime, timedelta
import threading
import time
import pandas as pd
from openpyxl import load_workbook

print('=== [LOG] 1.py –∏–º–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞–Ω ===')
nest_asyncio.apply()

# --- –†–∞–±–æ—Ç–∞ —Å Excel –¥–∞–Ω–Ω—ã–º–∏ ---
GOOGLE_SHEETS_URL = "https://docs.google.com/spreadsheets/d/1rvb3QdanuukCyXnoQZZxz7HF6aJXm2de/export?format=xlsx&gid=1870986273"
excel_data = None

def load_excel_data():
    """–ó–∞–≥—Ä—É–∂–∞–µ—Ç –¥–∞–Ω–Ω—ã–µ –∏–∑ Google Sheets"""
    global excel_data
    try:
        logger.info("Loading data from Google Sheets...")
        
        import requests
        import io
        import ssl
        from openpyxl import load_workbook
        
        # –û—Ç–∫–ª—é—á–∞–µ–º –ø—Ä–æ–≤–µ—Ä–∫—É SSL –¥–ª—è Google Sheets
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        
        # –ó–∞–≥—Ä—É–∂–∞–µ–º –¥–∞–Ω–Ω—ã–µ —á–µ—Ä–µ–∑ requests
        session = requests.Session()
        session.verify = False  # –û—Ç–∫–ª—é—á–∞–µ–º SSL –ø—Ä–æ–≤–µ—Ä–∫—É
        
        response = session.get(GOOGLE_SHEETS_URL, timeout=30)
        response.raise_for_status()
        
        # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≤–æ –≤—Ä–µ–º–µ–Ω–Ω—ã–π —Ñ–∞–π–ª –¥–ª—è –∏–∑–≤–ª–µ—á–µ–Ω–∏—è –≥–∏–ø–µ—Ä—Å—Å—ã–ª–æ–∫
        import tempfile
        import os
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(response.content)
            tmp_file_path = tmp_file.name
        
        try:
            # –ó–∞–≥—Ä—É–∂–∞–µ–º —Å –ø–æ–º–æ—â—å—é openpyxl –¥–ª—è –∏–∑–≤–ª–µ—á–µ–Ω–∏—è –≥–∏–ø–µ—Ä—Å—Å—ã–ª–æ–∫
            wb = load_workbook(tmp_file_path)
            ws = wb.active
            
            # –ò–∑–≤–ª–µ–∫–∞–µ–º –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏ –∏–∑ —Å—Ç–æ–ª–±—Ü–∞ B
            hyperlinks = {}
            for row in range(4, ws.max_row + 1):  # –ù–∞—á–∏–Ω–∞–µ–º —Å 4-–π —Å—Ç—Ä–æ–∫–∏
                cell = ws.cell(row=row, column=2)  # –°—Ç–æ–ª–±–µ—Ü B
                if cell.hyperlink:
                    hyperlinks[row] = cell.hyperlink.target
                    logger.info(f"–ù–∞–π–¥–µ–Ω–∞ –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∞ –≤ —Å—Ç—Ä–æ–∫–µ {row}: {cell.hyperlink.target}")
            
            wb.close()
            
        finally:
            # –£–¥–∞–ª—è–µ–º –≤—Ä–µ–º–µ–Ω–Ω—ã–π —Ñ–∞–π–ª
            os.unlink(tmp_file_path)
        
        # –ß–∏—Ç–∞–µ–º Excel –∏–∑ –ø–∞–º—è—Ç–∏ —á–µ—Ä–µ–∑ pandas
        df = pd.read_excel(io.BytesIO(response.content), header=2, skiprows=[3])
        
        # –û—á–∏—â–∞–µ–º –¥–∞–Ω–Ω—ã–µ
        df = df.dropna(how='all')
        
        # –§–∏–ª—å—Ç—Ä—É–µ–º –ø–æ –Ω–∞–ª–∏—á–∏—é –±—Ä–µ–Ω–¥–∞ –∏ –∞—Ä–æ–º–∞—Ç–∞ (—Å—Ç–æ–ª–±—Ü—ã 5 –∏ 6)
        if len(df.columns) > 6:
            df = df[df.iloc[:, 5].notna() & df.iloc[:, 6].notna()]
        else:
            logger.warning("Not enough columns in Google Sheets data")
            raise Exception("Invalid data structure")
        
        # –ü–µ—Ä–µ–∏–º–µ–Ω–æ–≤—ã–≤–∞–µ–º —Å—Ç–æ–ª–±—Ü—ã –¥–ª—è —É–¥–æ–±—Å—Ç–≤–∞ (–Ω–æ–≤–∞—è —Å—Ç—Ä—É–∫—Ç—É—Ä–∞ —Ç–∞–±–ª–∏—Ü—ã)
        if len(df.columns) >= 13:
            column_mapping = {
                df.columns[1]: '–°—Å—ã–ª–∫–∞',     # –°—Ç–æ–ª–±–µ—Ü B (—Å—Å—ã–ª–∫–∞ –Ω–∞ –∞—Ä–æ–º–∞—Ç)
                df.columns[5]: '–ë—Ä–µ–Ω–¥',      # –°—Ç–æ–ª–±–µ—Ü 5
                df.columns[6]: '–ê—Ä–æ–º–∞—Ç',     # –°—Ç–æ–ª–±–µ—Ü 6
                df.columns[7]: '–ü–æ–ª',        # –°—Ç–æ–ª–±–µ—Ü 7
                df.columns[8]: '–§–∞–±—Ä–∏–∫–∞',    # –°—Ç–æ–ª–±–µ—Ü 8
                df.columns[9]: '–ö–∞—á–µ—Å—Ç–≤–æ',   # –°—Ç–æ–ª–±–µ—Ü 9 (—É–∂–µ –≤ —Ñ–æ—Ä–º–∞—Ç–µ TOP/Q1/Q2)
                df.columns[10]: '30 GR',     # –°—Ç–æ–ª–±–µ—Ü 10
                df.columns[11]: '50 GR',     # –°—Ç–æ–ª–±–µ—Ü 11
                df.columns[12]: '500 GR',    # –°—Ç–æ–ª–±–µ—Ü 12
                df.columns[13]: '1 KG',      # –°—Ç–æ–ª–±–µ—Ü 13
            }
            
            # –î–æ–±–∞–≤–ª—è–µ–º —Å—Ç–æ–ª–±—Ü—ã –¥–ª—è –Ω–æ—Ç –∏ —Å—Ç—Ä–∞–Ω—ã (–µ—Å–ª–∏ –æ–Ω–∏ –µ—Å—Ç—å)
            if len(df.columns) > 14:
                column_mapping[df.columns[14]] = '–°—Ç—Ä–∞–Ω–∞'
            if len(df.columns) > 15:
                column_mapping[df.columns[15]] = '–í–µ—Ä—Ö–Ω–∏–µ –Ω–æ—Ç—ã'
            if len(df.columns) > 16:
                column_mapping[df.columns[16]] = '–°—Ä–µ–¥–Ω–∏–µ –Ω–æ—Ç—ã'
            if len(df.columns) > 17:
                column_mapping[df.columns[17]] = '–ë–∞–∑–æ–≤—ã–µ –Ω–æ—Ç—ã'
            
            # –î–æ–±–∞–≤–ª—è–µ–º —Å—Ç–æ–ª–±—Ü—ã TOP LAST –∏ TOP ALL (—Å–¥–≤–∏–≥–∞–µ–º –∏–Ω–¥–µ–∫—Å—ã –∏–∑-–∑–∞ –¥–æ–±–∞–≤–ª–µ–Ω–Ω—ã—Ö —Å—Ç–æ–ª–±—Ü–æ–≤)
            if len(df.columns) > 18:
                column_mapping[df.columns[18]] = 'TOP LAST'
            if len(df.columns) > 19:
                column_mapping[df.columns[19]] = 'TOP ALL'
            
            df = df.rename(columns=column_mapping)
        else:
            logger.warning(f"Not enough columns: {len(df.columns)}")
            raise Exception("Invalid column structure")
        
        # –î–æ–±–∞–≤–ª—è–µ–º –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏ –∫ –¥–∞–Ω–Ω—ã–º
        if hyperlinks:
            # –°–æ–∑–¥–∞–µ–º —Å—Ç–æ–ª–±–µ—Ü —Å –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∞–º–∏
            df['–ì–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∞'] = ''
            for idx, row in df.iterrows():
                # –ò–Ω–¥–µ–∫—Å —Å—Ç—Ä–æ–∫–∏ –≤ Excel (–Ω–∞—á–∏–Ω–∞—è —Å 4-–π —Å—Ç—Ä–æ–∫–∏)
                excel_row = idx + 4
                if excel_row in hyperlinks:
                    df.at[idx, '–ì–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∞'] = hyperlinks[excel_row]
                    logger.info(f"–î–æ–±–∞–≤–ª–µ–Ω–∞ –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∞ –¥–ª—è —Å—Ç—Ä–æ–∫–∏ {idx}: {hyperlinks[excel_row]}")
        
        # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º —Ç–∏–ø—ã –¥–∞–Ω–Ω—ã—Ö
        price_columns = ['30 GR', '50 GR', '500 GR', '1 KG']
        for col in price_columns:
            if col in df.columns:
                # –£–±–∏—Ä–∞–µ–º —Å–∏–º–≤–æ–ª—ã –≤–∞–ª—é—Ç—ã –∏ –∫–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –≤ —á–∏—Å–ª–∞
                df[col] = df[col].astype(str).str.replace('‚ÇΩ', '').str.replace(' ', '').str.replace('nan', '')
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º —Å—Ç–æ–ª–±—Ü—ã –ø–æ–ø—É–ª—è—Ä–Ω–æ—Å—Ç–∏ (–¥–∞–Ω–Ω—ã–µ —É–∂–µ –≤ –ø—Ä–æ—Ü–µ–Ω—Ç–∞—Ö)
        if 'TOP LAST' in df.columns:
            df['TOP LAST'] = pd.to_numeric(df['TOP LAST'], errors='coerce')
        
        if 'TOP ALL' in df.columns:
            df['TOP ALL'] = pd.to_numeric(df['TOP ALL'], errors='coerce')
        
        # –û—á–∏—â–∞–µ–º –ø—Ä–æ–±–µ–ª—ã –≤ –∫–∞—á–µ—Å—Ç–≤–µ
        if '–ö–∞—á–µ—Å—Ç–≤–æ' in df.columns:
            df['–ö–∞—á–µ—Å—Ç–≤–æ'] = df['–ö–∞—á–µ—Å—Ç–≤–æ'].astype(str).str.strip()
        
        # –û—á–∏—â–∞–µ–º –æ—Ç —Å—Ç—Ä–æ–∫ –±–µ–∑ –¥–∞–Ω–Ω—ã—Ö
        df = df[df['–ë—Ä–µ–Ω–¥'].notna() & df['–ê—Ä–æ–º–∞—Ç'].notna()]
        
        excel_data = df
        logger.info(f"Google Sheets data loaded: {len(df)} products")
        return df
        
    except Exception as e:
        logger.error(f"Failed to load Google Sheets data: {e}")
        # Fallback –∫ –ª–æ–∫–∞–ª—å–Ω–æ–º—É —Ñ–∞–π–ª—É
        try:
            logger.info("Falling back to local Excel file...")
            df = pd.read_excel("1.xlsx", header=2, skiprows=[3])
            df = df.dropna(how='all')
            df = df[~df['–ë—Ä–µ–Ω–¥'].astype(str).str.contains('Column', na=False)]
            df = df[df['–ë—Ä–µ–Ω–¥'].notna()]
            
            price_columns = ['30 GR', '50 GR', '500 GR', '1 KG', '5 KG', '10 KG']
            for col in price_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            excel_data = df
            logger.info(f"Local Excel data loaded: {len(df)} products")
            return df
        except Exception as e2:
            logger.error(f"Failed to load local Excel data: {e2}")
            return None

def normalize_name(name):
    return str(name).lower().replace('-', '').replace("'", '').replace(' ', '')

def search_products(query, limit=None):
    global excel_data
    if excel_data is None:
        return []
    query_norm = normalize_name(query)
    # –¢–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ –ø–æ –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω–æ–º—É –Ω–∞–∑–≤–∞–Ω–∏—é
    exact_mask = excel_data['–ê—Ä–æ–º–∞—Ç'].astype(str).apply(normalize_name) == query_norm
    if exact_mask.any():
        results = excel_data[exact_mask]
    else:
        # –ß–∞—Å—Ç–∏—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ –ø–æ –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω–æ–º—É –Ω–∞–∑–≤–∞–Ω–∏—é
        mask = excel_data['–ê—Ä–æ–º–∞—Ç'].astype(str).apply(normalize_name).str.contains(query_norm, na=False)
        results = excel_data[mask]
    if limit:
        results = results.head(limit)
    return results.to_dict('records')

def calculate_price(product, volume_ml):
    """–†–∞—Å—Å—á–∏—Ç—ã–≤–∞–µ—Ç —Ü–µ–Ω—É –∑–∞ —É–∫–∞–∑–∞–Ω–Ω—ã–π –æ–±—ä–µ–º"""
    try:
        volume_ml = float(volume_ml)
        
        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Ü–µ–Ω–æ–≤—É—é –∫–∞—Ç–µ–≥–æ—Ä–∏—é —Å–æ–≥–ª–∞—Å–Ω–æ –ø—Ä–∞–π—Å-–ª–∏—Å—Ç—É
        if volume_ml <= 49:
            price_per_ml = product.get('30 GR', 0)
            tier = '30-49 –º–ª'
        elif volume_ml <= 499:
            price_per_ml = product.get('50 GR', 0)
            tier = '50-499 –º–ª'
        elif volume_ml <= 999:
            price_per_ml = product.get('500 GR', 0)
            tier = '500-999 –º–ª'
        else:
            price_per_ml = product.get('1 KG', 0)
            tier = '1000+ –º–ª'
        
        if price_per_ml and not pd.isna(price_per_ml):
            total_price = float(price_per_ml) * volume_ml
            return {
                'volume_ml': volume_ml,
                'price_per_ml': float(price_per_ml),
                'total_price': total_price,
                'tier': tier,
                'currency': '—Ä—É–±'
            }
    except (ValueError, TypeError):
        pass
    
    return None

def get_quality_name(quality_code):
    """–ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ—Ç –∫–æ–¥ –∫–∞—á–µ—Å—Ç–≤–∞ –≤ –Ω–∞–∑–≤–∞–Ω–∏–µ"""
    quality_map = {
        6: 'TOP',
        5: 'Q1', 
        4: 'Q2'
    }
    return quality_map.get(quality_code, f'{quality_code}')

def get_quality_description(quality_code):
    """–ü–æ–ª—É—á–∞–µ—Ç –ø–æ–¥—Ä–æ–±–Ω–æ–µ –æ–ø–∏—Å–∞–Ω–∏–µ –∫–∞—á–µ—Å—Ç–≤–∞"""
    quality_desc_map = {
        6: 'TOP (–≤—ã—Å—à–µ–µ –∫–∞—á–µ—Å—Ç–≤–æ)',
        5: 'Q1 (–æ—Ç–ª–∏—á–Ω–æ–µ –∫–∞—á–µ—Å—Ç–≤–æ)', 
        4: 'Q2 (—Ö–æ—Ä–æ—à–µ–µ –∫–∞—á–µ—Å—Ç–≤–æ)'
    }
    return quality_desc_map.get(quality_code, f'–ö–∞—á–µ—Å—Ç–≤–æ {quality_code}')

def get_top_products(factory=None, quality=None, sort_by='TOP LAST', limit=None):
    """–ü–æ–ª—É—á–∞–µ—Ç —Ç–æ–ø –ø—Ä–æ–¥—É–∫—Ç—ã –ø–æ —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–µ (–±–µ–∑ –ª–∏–º–∏—Ç–∞ –ø–æ —É–º–æ–ª—á–∞–Ω–∏—é)"""
    global excel_data
    if excel_data is None:
        return []
    df = excel_data.copy()
    # –§–∏–ª—å—Ç—Ä—ã
    if factory:
        df = df[df['–§–∞–±—Ä–∏–∫–∞'].str.upper() == factory.upper()]
    if quality:
        df = df[df['–ö–∞—á–µ—Å—Ç–≤–æ'] == quality]
    # –°–æ—Ä—Ç–∏—Ä–æ–≤–∫–∞ –ø–æ –ø–æ–ø—É–ª—è—Ä–Ω–æ—Å—Ç–∏
    sort_column = 'TOP LAST' if sort_by == 'TOP LAST' else 'TOP ALL'
    df = df.sort_values(sort_column, ascending=False, na_position='last')
    if limit:
        df = df.head(limit)
    return df.to_dict('records')

def format_product_info(product, include_prices=True, for_chatgpt=True):
    """–§–æ—Ä–º–∞—Ç–∏—Ä—É–µ—Ç –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –ø—Ä–æ–¥—É–∫—Ç–µ —Å —É–ª—É—á—à–µ–Ω–Ω–æ–π —Å—Ç—Ä—É–∫—Ç—É—Ä–æ–π"""
    return format_aroma_response_improved(product, include_prices)

def get_aroma_variants_stats(aroma_name):
    """–í–æ–∑–≤—Ä–∞—â–∞–µ—Ç —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É –ø–æ –≤–∞—Ä–∏–∞–Ω—Ç–∞–º –∞—Ä–æ–º–∞—Ç–∞ (—Ñ–∞–±—Ä–∏–∫–∞+–∫–∞—á–µ—Å—Ç–≤–æ)"""
    global excel_data
    if excel_data is None:
        return []
    mask = excel_data['–ê—Ä–æ–º–∞—Ç'].str.lower().str.strip() == aroma_name.lower().strip()
    variants = excel_data[mask]
    if variants.empty:
        return []
    total_popularity = variants['TOP LAST'].sum()
    if total_popularity == 0:
        return []
    result = []
    for _, row in variants.iterrows():
        factory = row['–§–∞–±—Ä–∏–∫–∞']
        quality = row['–ö–∞—á–µ—Å—Ç–≤–æ']
        popularity = row['TOP LAST']
        percent = (popularity / total_popularity) * 100 if total_popularity else 0
        result.append({
            'factory': factory,
            'quality': quality,
            'popularity_percent': percent,
            'popularity_raw': popularity
        })
    return result

async def get_excel_context_for_chatgpt(query="", volume_ml=None, show_variants_stats=False):
    """–°–æ–∑–¥–∞–µ—Ç –°–¢–†–û–ì–û –°–¢–†–£–ö–¢–£–†–ò–†–û–í–ê–ù–ù–´–ô –∫–æ–Ω—Ç–µ–∫—Å—Ç –∏–∑ Excel –¥–∞–Ω–Ω—ã—Ö –¥–ª—è ChatGPT, —Å —Ä–∞—Å—á–µ—Ç–æ–º —Ü–µ–Ω –∏ —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–æ–π –≤–∞—Ä–∏–∞–Ω—Ç–æ–≤"""
    try:
        MAX_PRODUCTS_FOR_LLM = 20
        context = "\n=== –ê–ö–¢–£–ê–õ–¨–ù–´–ï –î–ê–ù–ù–´–ï –ò–ó –ü–†–ê–ô–°-–õ–ò–°–¢–ê ===\n"
        context += "–í–ù–ò–ú–ê–ù–ò–ï: –ò—Å–ø–æ–ª—å–∑—É–π –¢–û–õ–¨–ö–û —ç—Ç–∏ —Ü–µ–Ω—ã –∏ –ø—Ä–æ—Ü–µ–Ω—Ç—ã, –Ω–µ –ø—Ä–∏–¥—É–º—ã–≤–∞–π —Å–≤–æ–∏ –∑–Ω–∞—á–µ–Ω–∏—è!\n"
        PRAIS_URL = "http://clck.ru/jrimp"
        def format_prices(product):
            prices = []
            price_map = [
                ("30 GR", 30),
                ("50 GR", 50),
                ("500 GR", 500),
                ("1 KG", 1000)
            ]
            for col, vol in price_map:
                price_per_g = product.get(col)
                if price_per_g and not pd.isna(price_per_g):
                    total = int(price_per_g * vol)
                    prices.append(f"üíß{vol} –≥—Ä–∞–º–º = {total}‚ÇΩ ({price_per_g}‚ÇΩ - –∑–∞ 1 –≥—Ä–∞–º–º)")
                else:
                    prices.append(f"‚Ä¢ {vol} –º–ª ‚Äî –°—Ç–æ–∏–º–æ—Å—Ç—å –Ω–µ–¥–æ—Å—Ç—É–ø–Ω–∞")
            return "\n".join(prices)
        def get_top_variant(variants, key):
            if not variants:
                return None
            top = max(variants, key=key)
            return top
        def get_rank(product, all_products, key):
            sorted_products = sorted(all_products, key=key, reverse=True)
            for idx, p in enumerate(sorted_products, 1):
                if p['–ë—Ä–µ–Ω–¥'] == product['–ë—Ä–µ–Ω–¥'] and p['–ê—Ä–æ–º–∞—Ç'] == product['–ê—Ä–æ–º–∞—Ç'] and p['–§–∞–±—Ä–∏–∫–∞'] == product['–§–∞–±—Ä–∏–∫–∞'] and p['–ö–∞—á–µ—Å—Ç–≤–æ'] == product['–ö–∞—á–µ—Å—Ç–≤–æ']:
                    return idx
            return None
        # –ü–æ–∏—Å–∫ –ø–æ –∑–∞–ø—Ä–æ—Å—É
        if query:
            products = search_products(query, limit=None)
            total_found = len(products)
            if total_found > MAX_PRODUCTS_FOR_LLM:
                context += f"\n‚ö†Ô∏è –ù–∞–π–¥–µ–Ω–æ {total_found} –∞—Ä–æ–º–∞—Ç–æ–≤, –ø–æ–∫–∞–∑—ã–≤–∞—é —Ç–æ–ª—å–∫–æ –ø–µ—Ä–≤—ã–µ {MAX_PRODUCTS_FOR_LLM}. –£—Ç–æ—á–Ω–∏—Ç–µ –∑–∞–ø—Ä–æ—Å –¥–ª—è –±–æ–ª–µ–µ —Ç–æ—á–Ω–æ–≥–æ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞.\n"
                products = products[:MAX_PRODUCTS_FOR_LLM]
            if products:
                all_products_6m = get_top_products(sort_by='TOP LAST', limit=None)
                all_products_all = get_top_products(sort_by='TOP ALL', limit=None)
                # –ì—Ä—É–ø–ø–∏—Ä—É–µ–º –≤–∞—Ä–∏–∞–Ω—Ç—ã –ø–æ –Ω–∞–∑–≤–∞–Ω–∏—é –∞—Ä–æ–º–∞—Ç–∞
                aroma_name = products[0].get('–ê—Ä–æ–º–∞—Ç', '')
                variants = [p for p in products if p.get('–ê—Ä–æ–º–∞—Ç', '').strip().lower() == aroma_name.strip().lower()]
                show_variants_block = len(variants) > 1
                sum_last = sum(p.get('TOP LAST', 0) for p in variants)
                sum_all = sum(p.get('TOP ALL', 0) for p in variants)
                top_variant = get_top_variant(variants, lambda p: p.get('TOP LAST', 0))
                for i, product in enumerate(products, 1):
                    brand = product.get('–ë—Ä–µ–Ω–¥', 'N/A')
                    aroma_raw = product.get('–ê—Ä–æ–º–∞—Ç', 'N/A')
                    aroma = format_aroma_name(aroma_raw)
                    factory = product.get('–§–∞–±—Ä–∏–∫–∞', 'N/A')
                    quality = product.get('–ö–∞—á–µ—Å—Ç–≤–æ', 'N/A')
                    popularity_last = product.get('TOP LAST', 0)
                    popularity_all = product.get('TOP ALL', 0)
                    rank_6m = get_rank(product, all_products_6m, lambda p: p.get('TOP LAST', 0))
                    rank_all = get_rank(product, all_products_all, lambda p: p.get('TOP ALL', 0))
                                    # –ò—Å–ø–æ–ª—å–∑—É–µ–º –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫—É –∏–∑ –ø—Ä–∞–π—Å–∞, –µ—Å–ª–∏ –æ–Ω–∞ –µ—Å—Ç—å
                    hyperlink = product.get('–ì–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∞', '')
                    if hyperlink and not pd.isna(hyperlink) and str(hyperlink).strip() and str(hyperlink).strip().startswith('http'):
                        aroma_url = str(hyperlink).strip()
                        logger.info(f"–ò—Å–ø–æ–ª—å–∑—É–µ–º –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫—É –∏–∑ –ø—Ä–∞–π—Å–∞: {aroma_url}")
                    else:
                        # –ü—Ä–æ–≤–µ—Ä—è–µ–º –æ–±—ã—á–Ω—É—é —Å—Å—ã–ª–∫—É –∏–∑ –ø—Ä–∞–π—Å–∞
                        link = product.get('–°—Å—ã–ª–∫–∞', '')
                        logger.info(f"–°—Å—ã–ª–∫–∞ –¥–ª—è {brand} - {aroma}: {link}")
                        if link and not pd.isna(link) and str(link).strip() and str(link).strip().startswith('http'):
                            aroma_url = str(link).strip()
                            logger.info(f"–ò—Å–ø–æ–ª—å–∑—É–µ–º —Å—Å—ã–ª–∫—É –∏–∑ –ø—Ä–∞–π—Å–∞: {aroma_url}")
                        else:
                            # –ù–µ –≥–µ–Ω–µ—Ä–∏—Ä—É–µ–º —Å—Å—ã–ª–∫—É, –µ—Å–ª–∏ –µ—ë –Ω–µ—Ç –≤ –ø—Ä–∞–π—Å–µ
                            aroma_url = ""
                            logger.info(f"–°—Å—ã–ª–∫–∞ –∏–∑ –ø—Ä–∞–π—Å–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω–∞ –¥–ª—è {brand} - {aroma}")
                    
                    if brand != 'N/A' and aroma != 'N/A':
                        if aroma_url and aroma_url.startswith('http'):
                            context += f"‚ú® <a href='{aroma_url}'>{brand} - {aroma}</a>\n\n"
                        else:
                            context += f"‚ú®{brand} - {aroma}\n\n"
                    else:
                        context += f"‚ú®{brand} - {aroma}\n\n"
                    
                    # –î–æ–±–∞–≤–ª—è–µ–º –ø–æ–ø—É–ª—è—Ä–Ω–æ—Å—Ç—å
                    context += f"‚ö°Ô∏è TOP LAST: {popularity_last:.0f}% (‚Ññ{rank_6m})\n"
                    context += f"üöÄ TOP ALL: {popularity_all:.0f}% (‚Ññ{rank_all})\n"
                    
                    # –î–æ–±–∞–≤–ª—è–µ–º –ø—É—Å—Ç—É—é —Å—Ç—Ä–æ–∫—É –ø–æ—Å–ª–µ –ø–æ–ø—É–ª—è—Ä–Ω–æ—Å—Ç–∏
                    context += "\n"
                    
                    # –î–æ–±–∞–≤–ª—è–µ–º TOP VERSION (–ø—Ä–æ—Ü–µ–Ω—Ç–Ω–æ–µ —Å–æ–æ—Ç–Ω–æ—à–µ–Ω–∏–µ –ø–æ —Ñ–∞–±—Ä–∏–∫–∞–º –∏ –∫–∞—á–µ—Å—Ç–≤—É)
                    aroma_name = product.get('–ê—Ä–æ–º–∞—Ç', '')
                    if aroma_name and not pd.isna(aroma_name):
                        all_versions = [p for p in excel_data if p.get('–ê—Ä–æ–º–∞—Ç', '').strip().lower() == aroma_name.strip().lower()]
                        if len(all_versions) > 1:
                            total_popularity = sum(p.get('TOP LAST', 0) for p in all_versions)
                            if total_popularity > 0:
                                # –ì—Ä—É–ø–ø–∏—Ä—É–µ–º –ø–æ —Ñ–∞–±—Ä–∏–∫–∞–º –∏ –∫–∞—á–µ—Å—Ç–≤—É
                                factory_stats = {}
                                for version in all_versions:
                                    factory = version.get('–§–∞–±—Ä–∏–∫–∞', '')
                                    quality = version.get('–ö–∞—á–µ—Å—Ç–≤–æ', '')
                                    popularity = version.get('TOP LAST', 0)
                                    key = f"{factory} {quality}"
                                    if key not in factory_stats:
                                        factory_stats[key] = 0
                                    factory_stats[key] += popularity
                                
                                # –§–æ—Ä–º–∏—Ä—É–µ–º —Å—Ç—Ä–æ–∫—É —Å –ø—Ä–æ—Ü–µ–Ω—Ç–∞–º–∏
                                version_percents = []
                                for factory_key, popularity in factory_stats.items():
                                    percent = (popularity / total_popularity) * 100
                                    version_percents.append(f"{factory_key}: {percent:.1f}%")
                    
                    if version_percents:
                        context += f"   ‚ôæÔ∏è VERSION: {' | '.join(version_percents)}\n"
                    
                    # –î–æ–±–∞–≤–ª—è–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –Ω–æ—Ç–∞—Ö –∏ —Å—Ç—Ä–∞–Ω–µ
                    top_notes = product.get('–í–µ—Ä—Ö–Ω–∏–µ –Ω–æ—Ç—ã', '')
                    middle_notes = product.get('–°—Ä–µ–¥–Ω–∏–µ –Ω–æ—Ç—ã', '')
                    base_notes = product.get('–ë–∞–∑–æ–≤—ã–µ –Ω–æ—Ç—ã', '')
                    country = product.get('–°—Ç—Ä–∞–Ω–∞', '')
                    
                    # –ï—Å–ª–∏ –Ω–æ—Ç—ã –∏–ª–∏ —Å—Ç—Ä–∞–Ω–∞ –Ω–µ —É–∫–∞–∑–∞–Ω—ã –≤ –ø—Ä–∞–π—Å–µ, –ø—Ä–æ–±—É–µ–º –ø–æ–ª—É—á–∏—Ç—å —á–µ—Ä–µ–∑ API
                    if (not top_notes or pd.isna(top_notes) or not str(top_notes).strip()) or \
                       (not country or pd.isna(country) or not str(country).strip()):
                        try:
                            api_data = await get_notes_from_api(f"{brand} {aroma}")
                            if api_data:
                                if not top_notes or pd.isna(top_notes) or not str(top_notes).strip():
                                    top_notes = api_data.get("top_notes", "")
                                if not middle_notes or pd.isna(middle_notes) or not str(middle_notes).strip():
                                    middle_notes = api_data.get("middle_notes", "")
                                if not base_notes or pd.isna(base_notes) or not str(base_notes).strip():
                                    base_notes = api_data.get("base_notes", "")
                                if not country or pd.isna(country) or not str(country).strip():
                                    country = api_data.get("country", "")
                                # –ï—Å–ª–∏ —Å—Å—ã–ª–∫–∏ –Ω–µ—Ç –≤ –ø—Ä–∞–π—Å–µ, –±–µ—Ä–µ–º –∏–∑ API
                                if not aroma_url or aroma_url == "":
                                    api_link = api_data.get("link", "")
                                    if api_link and api_link.startswith('http'):
                                        aroma_url = api_link
                        except Exception as e:
                            logger.error(f"Error getting API data: {e}")
                    
                    # –û—Ç–æ–±—Ä–∞–∂–∞–µ–º –Ω–æ—Ç—ã
                    if top_notes and not pd.isna(top_notes) and str(top_notes).strip():
                        context += f"üå± –í–µ—Ä—Ö–Ω–∏–µ –Ω–æ—Ç—ã: {str(top_notes).strip()}\n"
                    else:
                        context += f"üå± –í–µ—Ä—Ö–Ω–∏–µ –Ω–æ—Ç—ã: –ù–µ —É–∫–∞–∑–∞–Ω—ã\n"
                    if middle_notes and not pd.isna(middle_notes) and str(middle_notes).strip():
                        context += f"üåø –°—Ä–µ–¥–Ω–∏–µ –Ω–æ—Ç—ã: {str(middle_notes).strip()}\n"
                    else:
                        context += f"üåø –°—Ä–µ–¥–Ω–∏–µ –Ω–æ—Ç—ã: –ù–µ —É–∫–∞–∑–∞–Ω—ã\n"
                    if base_notes and not pd.isna(base_notes) and str(base_notes).strip():
                        context += f"üçÉ –ë–∞–∑–æ–≤—ã–µ –Ω–æ—Ç—ã: {str(base_notes).strip()}\n"
                    else:
                        context += f"üçÉ –ë–∞–∑–æ–≤—ã–µ –Ω–æ—Ç—ã: –ù–µ —É–∫–∞–∑–∞–Ω—ã\n"
                    
                    # –î–æ–±–∞–≤–ª—è–µ–º –ø—É—Å—Ç—É—é —Å—Ç—Ä–æ–∫—É –ø–æ—Å–ª–µ –Ω–æ—Ç
                    context += "\n"
                    
                    # –û—Ç–æ–±—Ä–∞–∂–∞–µ–º –±—Ä–µ–Ω–¥ –∏ —Å—Ç—Ä–∞–Ω—É
                    context += f"¬Æ –ë—Ä–µ–Ω–¥: {brand}\n"
                    logger.info(f"–°—Ç—Ä–∞–Ω–∞ –¥–ª—è {brand} - {aroma}: '{country}'")
                    country_emoji = get_country_emoji(country)
                    logger.info(f"–≠–º–æ–¥–∂–∏ –¥–ª—è —Å—Ç—Ä–∞–Ω—ã '{country}': {country_emoji}")
                    if country and not pd.isna(country) and str(country).strip():
                        context += f"{country_emoji} –°—Ç—Ä–∞–Ω–∞: {str(country).strip()}\n"
                    else:
                        context += f"{country_emoji} –°—Ç—Ä–∞–Ω–∞: –ù–µ —É–∫–∞–∑–∞–Ω–∞\n"
                    
                    # –î–æ–±–∞–≤–ª—è–µ–º –ø—É—Å—Ç—É—é —Å—Ç—Ä–æ–∫—É –ø–æ—Å–ª–µ —Å—Ç—Ä–∞–Ω—ã
                    context += "\n"
                    # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ –≤–∞—Ä–∏–∞–Ω—Ç–∞–º (–µ—Å–ª–∏ –µ—Å—Ç—å –≤–∞—Ä–∏–∞–Ω—Ç—ã)
                    if show_variants_block and i == 1:
                        context += f"üìä –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ –≤–∞—Ä–∏–∞–Ω—Ç–∞–º –∞—Ä–æ–º–∞—Ç–∞ '{aroma_name}':\n"
                        for v in variants:
                            percent_last = (v.get('TOP LAST', 0) / sum_last * 100) if sum_last else 0
                            percent_all = (v.get('TOP ALL', 0) / sum_all * 100) if sum_all else 0
                            mark = " (—Å–∞–º—ã–π –ø–æ–ø—É–ª—è—Ä–Ω—ã–π)" if top_variant and v['–§–∞–±—Ä–∏–∫–∞'] == top_variant['–§–∞–±—Ä–∏–∫–∞'] and v['–ö–∞—á–µ—Å—Ç–≤–æ'] == top_variant['–ö–∞—á–µ—Å—Ç–≤–æ'] else ""
                            context += f"- {v['–§–∞–±—Ä–∏–∫–∞']} ({v['–ö–∞—á–µ—Å—Ç–≤–æ']}): {percent_last:.1f}% –∑–∞ 6 –º–µ—Å, {percent_all:.1f}% –∑–∞ –≤—Å—ë –≤—Ä–µ–º—è{mark}\n"
                    # –û—Ç—Å—Ç—É–ø –ø–µ—Ä–µ–¥ —Å—Ç–æ–∏–º–æ—Å—Ç—å—é
                    context += "\n"
                    context += f"üíµ –°—Ç–æ–∏–º–æ—Å—Ç—å:\n{format_prices(product)}\n\n"
        # –¢–û–ü-–∞—Ä–æ–º–∞—Ç—ã (–≤–µ—Å—å –ø—Ä–∞–π—Å, –Ω–æ —Å –ª–∏–º–∏—Ç–æ–º)
        all_products_6m = get_top_products(sort_by='TOP LAST', limit=MAX_PRODUCTS_FOR_LLM)
        all_products_all = get_top_products(sort_by='TOP ALL', limit=MAX_PRODUCTS_FOR_LLM)
        if all_products_6m:
            context += f"\nüî• –¢–û–ü-{MAX_PRODUCTS_FOR_LLM} –ü–û–ü–£–õ–Ø–†–ù–´–• –ê–†–û–ú–ê–¢–û–í (–ø–æ—Å–ª–µ–¥–Ω–∏–µ 6 –º–µ—Å—è—Ü–µ–≤):\n"
            for i, product in enumerate(all_products_6m, 1):
                brand = product.get('–ë—Ä–µ–Ω–¥', 'N/A')
                aroma = product.get('–ê—Ä–æ–º–∞—Ç', 'N/A')
                factory = product.get('–§–∞–±—Ä–∏–∫–∞', 'N/A')
                quality = product.get('–ö–∞—á–µ—Å—Ç–≤–æ', 'N/A')
                popularity_last = product.get('TOP LAST', 0)
                popularity_all = product.get('TOP ALL', 0)
                rank_6m = i
                rank_all = get_rank(product, all_products_all, lambda p: p.get('TOP ALL', 0))
                # –ò—Å–ø–æ–ª—å–∑—É–µ–º —Å—Å—ã–ª–∫—É –∏–∑ –ø—Ä–∞–π—Å–∞, –µ—Å–ª–∏ –æ–Ω–∞ –µ—Å—Ç—å –∏ –≤–∞–ª–∏–¥–Ω–∞
                link = product.get('–°—Å—ã–ª–∫–∞', '')
                if link and not pd.isna(link) and str(link).strip() and str(link).strip().startswith('http'):
                    aroma_url = str(link).strip()
                else:
                    # –ù–µ –≥–µ–Ω–µ—Ä–∏—Ä—É–µ–º —Å—Å—ã–ª–∫—É, –µ—Å–ª–∏ –µ—ë –Ω–µ—Ç –≤ –ø—Ä–∞–π—Å–µ
                    aroma_url = ""
                if brand != 'N/A' and aroma != 'N/A':
                    context += f"{i}. <a href='{aroma_url}'>{brand} - {aroma}</a>\n   üè≠ {factory} ({quality})\n   üìà –ü–æ–ø—É–ª—è—Ä–Ω–æ—Å—Ç—å (6 –º–µ—Å): {popularity_last:.2f}% (‚Ññ{rank_6m})\n   üìä –ü–æ–ø—É–ª—è—Ä–Ω–æ—Å—Ç—å (–≤—Å—ë –≤—Ä–µ–º—è): {popularity_all:.2f}% (‚Ññ{rank_all})\n\nüí∞ –°—Ç–æ–∏–º–æ—Å—Ç—å:\n{format_prices(product)}\n\n"
                else:
                    context += f"{i}. {brand} - {aroma}\n   üè≠ {factory} ({quality})\n   ÔøΩÔøΩ –ü–æ–ø—É–ª—è—Ä–Ω–æ—Å—Ç—å (6 –º–µ—Å): {popularity_last:.2f}% (‚Ññ{rank_6m})\n   üìä –ü–æ–ø—É–ª—è—Ä–Ω–æ—Å—Ç—å (–≤—Å—ë –≤—Ä–µ–º—è): {popularity_all:.2f}% (‚Ññ{rank_all})\n\nüí∞ –°—Ç–æ–∏–º–æ—Å—Ç—å:\n{format_prices(product)}\n\n"
        # –ò–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è –æ —Ñ–∞–±—Ä–∏–∫–∞—Ö
        context += "\nüè≠ –î–û–°–¢–£–ü–ù–´–ï –§–ê–ë–†–ò–ö–ò: EPS, LUZI, SELUZ, UNKNOWN\n"
        context += "‚≠ê –ö–ê–ß–ï–°–¢–í–ê: TOP > Q1 > Q2\n"
        context += "\nüí∞ –¶–ï–ù–û–í–´–ï –ö–ê–¢–ï–ì–û–†–ò–ò:\n"
        context += "‚Ä¢ 30-49 –º–ª: —Ü–µ–Ω–∞ –∏–∑ —Å—Ç–æ–ª–±—Ü–∞ '30 GR'\n"
        context += "‚Ä¢ 50-499 –º–ª: —Ü–µ–Ω–∞ –∏–∑ —Å—Ç–æ–ª–±—Ü–∞ '50 GR'\n"
        context += "‚Ä¢ 500-999 –º–ª: —Ü–µ–Ω–∞ –∏–∑ —Å—Ç–æ–ª–±—Ü–∞ '500 GR'\n"
        context += "‚Ä¢ 1000+ –º–ª: —Ü–µ–Ω–∞ –∏–∑ —Å—Ç–æ–ª–±—Ü–∞ '1 KG'\n"
        return context
    except Exception as e:
        logger.error(f"Error creating Excel context: {e}")
        return "\n‚ùå –û—à–∏–±–∫–∞ –∑–∞–≥—Ä—É–∑–∫–∏ –∞–∫—Ç—É–∞–ª—å–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö –∏–∑ –ø—Ä–∞–π—Å-–ª–∏—Å—Ç–∞\n"

# --- –ë–∞–∑–∞ –¥–∞–Ω–Ω—ã—Ö SQLite ---
DB_NAME = "bot_users.db"

def init_database():
    """–ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö"""
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        # –°–æ–∑–¥–∞–µ–º —Ç–∞–±–ª–∏—Ü—É –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY,
                chat_id INTEGER NOT NULL,
                first_name TEXT,
                username TEXT,
                first_start_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                last_activity DATETIME DEFAULT CURRENT_TIMESTAMP,
                last_weekly_message DATETIME,
                is_active BOOLEAN DEFAULT 1
            )
        ''')
        
        conn.commit()
        conn.close()
        logger.info("Database initialized successfully")
    except Exception as e:
        logger.error(f"Failed to initialize database: {e}")

def add_user_to_db(user_id, chat_id, first_name=None, username=None):
    """–î–æ–±–∞–≤–ª—è–µ—Ç –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ –±–∞–∑—É –¥–∞–Ω–Ω—ã—Ö –∏–ª–∏ –æ–±–Ω–æ–≤–ª—è–µ—Ç –ø–æ—Å–ª–µ–¥–Ω—é—é –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å"""
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –ª–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å
        cursor.execute('SELECT user_id FROM users WHERE user_id = ?', (user_id,))
        exists = cursor.fetchone()
        
        if exists:
            # –û–±–Ω–æ–≤–ª—è–µ–º –ø–æ—Å–ª–µ–¥–Ω—é—é –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å
            cursor.execute('''
                UPDATE users 
                SET last_activity = CURRENT_TIMESTAMP, is_active = 1, chat_id = ?
                WHERE user_id = ?
            ''', (chat_id, user_id))
            logger.info(f"Updated user activity: {user_id}")
        else:
            # –î–æ–±–∞–≤–ª—è–µ–º –Ω–æ–≤–æ–≥–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
            cursor.execute('''
                INSERT INTO users (user_id, chat_id, first_name, username, first_start_date, last_activity)
                VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            ''', (user_id, chat_id, first_name, username))
            logger.info(f"Added new user to database: {user_id}")
        
        conn.commit()
        conn.close()
    except Exception as e:
        logger.error(f"Failed to add/update user in database: {e}")

def get_users_for_weekly_message():
    """–ü–æ–ª—É—á–∞–µ—Ç –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π, –∫–æ—Ç–æ—Ä—ã–º –Ω—É–∂–Ω–æ –æ—Ç–ø—Ä–∞–≤–∏—Ç—å –µ–∂–µ–Ω–µ–¥–µ–ª—å–Ω–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ"""
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        # –ü–æ–ª—É—á–∞–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π, –∫–æ—Ç–æ—Ä—ã–º –Ω–µ –æ—Ç–ø—Ä–∞–≤–ª—è–ª–∏ —Å–æ–æ–±—â–µ–Ω–∏–µ –±–æ–ª—å—à–µ –Ω–µ–¥–µ–ª–∏
        week_ago = datetime.now() - timedelta(days=7)
        cursor.execute('''
            SELECT user_id, chat_id, first_name FROM users 
            WHERE is_active = 1 AND (
                last_weekly_message IS NULL OR 
                last_weekly_message < ?
            )
        ''', (week_ago,))
        
        users = cursor.fetchall()
        conn.close()
        return users
    except Exception as e:
        logger.error(f"Failed to get users for weekly message: {e}")
        return []

def update_weekly_message_sent(user_id):
    """–û–±–Ω–æ–≤–ª—è–µ—Ç –≤—Ä–µ–º—è –æ—Ç–ø—Ä–∞–≤–∫–∏ –µ–∂–µ–Ω–µ–¥–µ–ª—å–Ω–æ–≥–æ —Å–æ–æ–±—â–µ–Ω–∏—è"""
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE users 
            SET last_weekly_message = CURRENT_TIMESTAMP
            WHERE user_id = ?
        ''', (user_id,))
        
        conn.commit()
        conn.close()
    except Exception as e:
        logger.error(f"Failed to update weekly message timestamp: {e}")

def deactivate_user(user_id):
    """–î–µ–∞–∫—Ç–∏–≤–∏—Ä—É–µ—Ç –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è (–Ω–∞–ø—Ä–∏–º–µ—Ä, –µ—Å–ª–∏ –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–ª –±–æ—Ç–∞)"""
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE users 
            SET is_active = 0
            WHERE user_id = ?
        ''', (user_id,))
        
        conn.commit()
        conn.close()
        logger.info(f"Deactivated user: {user_id}")
    except Exception as e:
        logger.error(f"Failed to deactivate user: {e}")

# --- –ï–∂–µ–Ω–µ–¥–µ–ª—å–Ω–∞—è —Ä–∞—Å—Å—ã–ª–∫–∞ ---
def get_weekly_message():
    """–ì–µ–Ω–µ—Ä–∏—Ä—É–µ—Ç —Å–ª—É—á–∞–π–Ω–æ–µ –µ–∂–µ–Ω–µ–¥–µ–ª—å–Ω–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ –æ—Ç –º–µ–¥–≤–µ–∂–æ–Ω–∫–∞"""
    messages = [
        "üêÜ –ü—Ä–∏–≤–µ—Ç! –Ø —Å–æ—Å–∫—É—á–∏–ª–∞—Å—å, –ø–æ–≥–æ–≤–æ—Ä–∏ —Å–æ –º–Ω–æ–π! –ú–æ–∂–µ—Ç —Ä–∞—Å—Å–∫–∞–∂–µ—à—å, –∫–∞–∫–∏–µ –∞—Ä–æ–º–∞—Ç—ã —Ç–µ–±—è —Å–µ–π—á–∞—Å –∏–Ω—Ç–µ—Ä–µ—Å—É—é—Ç? ‚ú®",
        "üêæ –ü–∞–Ω—Ç–µ—Ä–∞ —Å–∫—É—á–∞–µ—Ç! –î–∞–≤–∞–π –ø–æ–æ–±—â–∞–µ–º—Å—è? –•–æ—á—É —É–∑–Ω–∞—Ç—å, –∫–∞–∫–∏–µ –¥—É—Ö–∏ —Ç—ã —Å–µ–π—á–∞—Å –Ω–æ—Å–∏—à—å! üí´",
        "üêÜ –ü—Ä–∏–≤–µ—Ç, –¥–æ—Ä–æ–≥–∞—è! –Ø —Ç–∞–∫ –ø–æ —Ç–µ–±–µ —Å–æ—Å–∫—É—á–∏–ª–∞—Å—å! –ú–æ–∂–µ—Ç –ø–æ–±–æ–ª—Ç–∞–µ–º –æ–± –∞—Ä–æ–º–∞—Ç–∞—Ö? –£ –º–µ–Ω—è —Å—Ç–æ–ª—å–∫–æ –Ω–æ–≤–æ—Å—Ç–µ–π! ‚ú®",
        "üêÜ –ü–∞–Ω—Ç–µ—Ä–∏–Ω–æ–µ —Å–µ—Ä–¥–µ—á–∫–æ —Å–∫—É—á–∞–µ—Ç! –ü–æ–≥–æ–≤–æ—Ä–∏ —Å–æ –º–Ω–æ–π, —Ä–∞—Å—Å–∫–∞–∂–∏, –∫–∞–∫–∏–µ –Ω–æ–≤—ã–µ –∞—Ä–æ–º–∞—Ç—ã —Ö–æ—á–µ—à—å –ø–æ–ø—Ä–æ–±–æ–≤–∞—Ç—å? üåü",
        "üêæ –ü—Ä–∏–≤–µ—Ç-–ø—Ä–∏–≤–µ—Ç! –¢–≤–æ—è –∞—Ä–æ–º–∞—Ç–Ω–∞—è –ø–∞–Ω—Ç–µ—Ä–∞ —Å–∫—É—á–∞–µ—Ç! –î–∞–≤–∞–π –æ–±—Å—É–¥–∏–º —á—Ç–æ-–Ω–∏–±—É–¥—å –∏–Ω—Ç–µ—Ä–µ—Å–Ω–æ–µ –ø—Ä–æ –¥—É—Ö–∏? üíé",
        "üêÜ –°–æ—Å–∫—É—á–∏–ª–∞—Å—å –±–µ–∑—É–º–Ω–æ! –•–æ—á–µ—Ç—Å—è –ø–æ–±–æ–ª—Ç–∞—Ç—å —Å —Ç–æ–±–æ–π –æ–± –∞—Ä–æ–º–∞—Ç–∞—Ö! –ú–æ–∂–µ—Ç —á—Ç–æ-—Ç–æ –Ω–æ–≤–µ–Ω—å–∫–æ–µ –∏—â–µ—à—å? ‚ú®",
        "üêæ –ü–∞–Ω—Ç–µ—Ä–∞ –≥—Ä—É—Å—Ç–∏—Ç –±–µ–∑ –æ–±—â–µ–Ω–∏—è! –î–∞–≤–∞–π –ø–æ–≥–æ–≤–æ—Ä–∏–º? –†–∞—Å—Å–∫–∞–∂–∏, –∫–∞–∫–æ–µ —É —Ç–µ–±—è —Å–µ–π—á–∞—Å –Ω–∞—Å—Ç—Ä–æ–µ–Ω–∏–µ –∏ –∫–∞–∫–æ–π –∞—Ä–æ–º–∞—Ç –ø–æ–¥–æ–π–¥–µ—Ç! üåü"
    ]
    return random.choice(messages)

async def send_weekly_messages():
    """–û—Ç–ø—Ä–∞–≤–ª—è–µ—Ç –µ–∂–µ–Ω–µ–¥–µ–ª—å–Ω—ã–µ —Å–æ–æ–±—â–µ–Ω–∏—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è–º"""
    logger.info("Starting weekly message sending...")
    users = get_users_for_weekly_message()
    
    if not users:
        logger.info("No users need weekly messages")
        return
    
    sent_count = 0
    failed_count = 0
    
    for user_id, chat_id, first_name in users:
        try:
            message = get_weekly_message()
            success = await telegram_send_message(chat_id, message)
            
            if success:
                update_weekly_message_sent(user_id)
                sent_count += 1
                logger.info(f"Weekly message sent to user {user_id}")
            else:
                failed_count += 1
                # –í–æ–∑–º–æ–∂–Ω–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–ª –±–æ—Ç–∞
                logger.warning(f"Failed to send weekly message to user {user_id}")
                
            # –ü–∞—É–∑–∞ –º–µ–∂–¥—É –æ—Ç–ø—Ä–∞–≤–∫–∞–º–∏, —á—Ç–æ–±—ã –Ω–µ –Ω–∞—Ä—É—à–∏—Ç—å –ª–∏–º–∏—Ç—ã API
            await asyncio.sleep(1)
            
        except Exception as e:
            failed_count += 1
            logger.error(f"Error sending weekly message to user {user_id}: {e}")
            # –ï—Å–ª–∏ –æ—à–∏–±–∫–∞ 403 (Forbidden), –¥–µ–∞–∫—Ç–∏–≤–∏—Ä—É–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
            if "403" in str(e) or "Forbidden" in str(e):
                deactivate_user(user_id)
    
    logger.info(f"Weekly messages completed: {sent_count} sent, {failed_count} failed")

def weekly_message_scheduler():
    """–ü–ª–∞–Ω–∏—Ä–æ–≤—â–∏–∫ –µ–∂–µ–Ω–µ–¥–µ–ª—å–Ω—ã—Ö —Å–æ–æ–±—â–µ–Ω–∏–π (—Ä–∞–±–æ—Ç–∞–µ—Ç –≤ –æ—Ç–¥–µ–ª—å–Ω–æ–º –ø–æ—Ç–æ–∫–µ)"""
    logger.info("Weekly message scheduler started")
    
    while True:
        try:
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º –∫–∞–∂–¥—ã–π —á–∞—Å
            time.sleep(3600)  # 1 hour
            
            # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —Å–æ–æ–±—â–µ–Ω–∏—è –≤ 7:00 –ø–æ –ø–æ–Ω–µ–¥–µ–ª—å–Ω–∏–∫–∞–º
            now = datetime.now()
            if now.weekday() == 0 and now.hour == 7:  # –ü–æ–Ω–µ–¥–µ–ª—å–Ω–∏–∫, 7:00
                # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—ã–π event loop –¥–ª—è —ç—Ç–æ–≥–æ –ø–æ—Ç–æ–∫–∞
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                loop.run_until_complete(send_weekly_messages())
                loop.close()
                
                # –°–ø–∏–º –µ—â–µ —á–∞—Å, —á—Ç–æ–±—ã –Ω–µ –æ—Ç–ø—Ä–∞–≤–ª—è—Ç—å –ø–æ–≤—Ç–æ—Ä–Ω–æ
                time.sleep(3600)
                
        except Exception as e:
            logger.error(f"Error in weekly message scheduler: {e}")
            time.sleep(3600)  # –ü—Ä–∏ –æ—à–∏–±–∫–µ —Ç–æ–∂–µ –∂–¥–µ–º —á–∞—Å

# –ì–ª–æ–±–∞–ª—å–Ω–∞—è –ø–µ—Ä–µ–º–µ–Ω–Ω–∞—è –¥–ª—è –ø–æ—Ç–æ–∫–∞ –ø–ª–∞–Ω–∏—Ä–æ–≤—â–∏–∫–∞
scheduler_thread = None

def start_weekly_scheduler():
    """–ó–∞–ø—É—Å–∫–∞–µ—Ç –ø–ª–∞–Ω–∏—Ä–æ–≤—â–∏–∫ –µ–∂–µ–Ω–µ–¥–µ–ª—å–Ω—ã—Ö —Å–æ–æ–±—â–µ–Ω–∏–π"""
    global scheduler_thread
    if scheduler_thread is None or not scheduler_thread.is_alive():
        scheduler_thread = threading.Thread(target=weekly_message_scheduler, daemon=True)
        scheduler_thread.start()
        logger.info("Weekly scheduler thread started")

# --- –ö–æ–Ω—Ñ–∏–≥—É—Ä–∞—Ü–∏—è ---
TOKEN = os.getenv('TOKEN')
BASE_WEBHOOK_URL = os.getenv('WEBHOOK_BASE_URL')
WEBHOOK_PATH = "/webhook/ai-bear-123456"
OPENAI_API = os.getenv('OPENAI_API_KEY')

# –ü—Ä–æ–≤–µ—Ä—è–µ–º –ø–µ—Ä–µ–º–µ–Ω–Ω—ã–µ –æ–∫—Ä—É–∂–µ–Ω–∏—è —Å –±–æ–ª–µ–µ –º—è–≥–∫–æ–π –æ–±—Ä–∞–±–æ—Ç–∫–æ–π
if not TOKEN:
    print("‚ö†Ô∏è WARNING: TOKEN environment variable not set!")
if not BASE_WEBHOOK_URL:
    print("‚ö†Ô∏è WARNING: WEBHOOK_BASE_URL environment variable not set!")
if not OPENAI_API:
    print("‚ö†Ô∏è WARNING: OPENAI_API_KEY environment variable not set!")
    print("‚ö†Ô∏è The bot will not be able to process AI requests without this key!")
    # –ù–µ –ø—Ä–µ—Ä—ã–≤–∞–µ–º –≤—ã–ø–æ–ª–Ω–µ–Ω–∏–µ, –ø—Ä–æ—Å—Ç–æ –ø—Ä–µ–¥—É–ø—Ä–µ–∂–¥–∞–µ–º

# --- FastAPI app ---
print('=== [LOG] FastAPI app —Å–æ–∑–¥–∞—ë—Ç—Å—è ===')
app = FastAPI()
print('=== [LOG] FastAPI app —Å–æ–∑–¥–∞–Ω ===')

# –ì–ª–æ–±–∞–ª—å–Ω—ã–π –æ–±—Ä–∞–±–æ—Ç—á–∏–∫ –∏—Å–∫–ª—é—á–µ–Ω–∏–π
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    logger.error(f"Global exception handler: {exc}\n{traceback.format_exc()}")
    return JSONResponse(
        status_code=500,
        content={"detail": "Internal server error"}
    )

print(f'=== [LOG] WEBHOOK_PATH: {WEBHOOK_PATH} ===')

@app.on_event("startup")
async def log_routes():
    logger.info("=== ROUTES REGISTERED ===")
    for route in app.routes:
        logger.info(f"{route.path} [{','.join(route.methods or [])}]")
    logger.info(f"WEBHOOK_PATH: {WEBHOOK_PATH}")
    logger.info("=========================")

# --- ChatGPT –∏ –¥–∞–Ω–Ω—ã–µ Bahur ---
def load_bahur_data():
    with open("bahur_data.txt", "r", encoding="utf-8") as f:
        return f.read()

BAHUR_DATA = load_bahur_data()

# --- –õ–æ–≥–∏—Ä–æ–≤–∞–Ω–∏–µ ---
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# --- –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö ---
init_database()

# --- –ó–∞–≥—Ä—É–∑–∫–∞ Excel –¥–∞–Ω–Ω—ã—Ö ---
load_excel_data()

# --- –°–æ—Å—Ç–æ—è–Ω–∏—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –¥–ª—è AI (in-memory, not persistent) ---
user_states = {}

# --- –ü–æ—Å—Ç–æ—è–Ω–Ω–æ–µ —Ö—Ä–∞–Ω–µ–Ω–∏–µ —Å–æ—Å—Ç–æ—è–Ω–∏–π –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π ---
import json

def load_user_states():
    try:
        with open("user_states.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def save_user_states(states):
    with open("user_states.json", "w", encoding="utf-8") as f:
        json.dump(states, f, ensure_ascii=False, indent=2)

def set_user_state(user_id, state):
    global user_states
    if state:
        user_states[user_id] = state
    else:
        user_states.pop(user_id, None)
    save_user_states(user_states)

def get_user_state(user_id):
    global user_states
    return user_states.get(user_id)

# –ó–∞–≥—Ä—É–∂–∞–µ–º —Å–æ—Å—Ç–æ—è–Ω–∏—è –ø—Ä–∏ –∑–∞–ø—É—Å–∫–µ
user_states = load_user_states()

# --- –ú–æ–¥–µ–ª–∏ –¥–ª—è FastAPI ---
class MessageModel(BaseModel):
    user_id: int
    text: str

class CallbackModel(BaseModel):
    user_id: int
    data: str

# --- –£—Ç–∏–ª–∏—Ç—ã ---
def greet():
    return random.choice([
            "–ü—Ä–∏–≤–µ—Ç! üêæ‚ú® –Ø AI-–ü–∞–Ω—Ç–µ—Ä–∞ ‚Äî —ç–∫—Å–ø–µ—Ä—Ç –ø–æ –∞—Ä–æ–º–∞—Ç–∞–º BAHUR! –°–ø—Ä–∞—à–∏–≤–∞–π –ø—Ä–æ –ª—é–±—ã–µ –¥—É—Ö–∏, –º–∞—Å–ª–∞, –¥–æ—Å—Ç–∞–≤–∫—É –∏–ª–∏ —Ü–µ–Ω—ã ‚Äî —è –Ω–∞–π–¥—É –≤—Å—ë –≤ –Ω–∞—à–µ–º –∫–∞—Ç–∞–ª–æ–≥–µ! üåü",
            "–ó–¥—Ä–∞–≤—Å—Ç–≤—É–π! üêÜüí´ –ì–æ—Ç–æ–≤–∞ –ø–æ–º–æ—á—å —Å –≤—ã–±–æ—Ä–æ–º –∞—Ä–æ–º–∞—Ç–æ–≤! –•–æ—á–µ—à—å —É–∑–Ω–∞—Ç—å –ø—Ä–æ –∫–æ–Ω–∫—Ä–µ—Ç–Ω—ã–µ –¥—É—Ö–∏, –º–∞—Å–ª–∞, –¥–æ—Å—Ç–∞–≤–∫—É –∏–ª–∏ —Ü–µ–Ω—ã? –°–ø—Ä–∞—à–∏–≤–∞–π ‚Äî —É –º–µ–Ω—è –µ—Å—Ç—å –ø–æ–ª–Ω—ã–π –∫–∞—Ç–∞–ª–æ–≥! ‚ú®",
            "–ü—Ä–∏–≤–µ—Ç, –∞—Ä–æ–º–∞—Ç–Ω—ã–π –¥—Ä—É–≥! üêÜ‚ú® –Ø –∑–Ω–∞—é –≤—Å—ë –æ –¥—É—Ö–∞—Ö BAHUR! –°–ø—Ä–∞—à–∏–≤–∞–π –ø—Ä–æ –ª—é–±—ã–µ –∞—Ä–æ–º–∞—Ç—ã, –º–∞—Å–ª–∞, –¥–æ—Å—Ç–∞–≤–∫—É ‚Äî –Ω–∞–π–¥—É –≤ –∫–∞—Ç–∞–ª–æ–≥–µ –∏ —Ä–∞—Å—Å–∫–∞–∂—É –ø–æ–¥—Ä–æ–±–Ω–æ! üåü",
            "–î–æ–±—Ä–æ –ø–æ–∂–∞–ª–æ–≤–∞—Ç—å! üéØüêÜ –Ø —ç–∫—Å–ø–µ—Ä—Ç –ø–æ –∞—Ä–æ–º–∞—Ç–∞–º BAHUR! –•–æ—á–µ—à—å —É–∑–Ω–∞—Ç—å –ø—Ä–æ –∫–æ–Ω–∫—Ä–µ—Ç–Ω—ã–µ –¥—É—Ö–∏, –º–∞—Å–ª–∞, —Ü–µ–Ω—ã –∏–ª–∏ –¥–æ—Å—Ç–∞–≤–∫—É? –°–ø—Ä–∞—à–∏–≤–∞–π ‚Äî —É –º–µ–Ω—è –µ—Å—Ç—å –≤—Å–µ –¥–∞–Ω–Ω—ã–µ! ‚ú®",
            "–ü—Ä–∏–≤–µ—Ç! üåüüêæ –Ø AI-–ü–∞–Ω—Ç–µ—Ä–∞ ‚Äî –∑–Ω–∞—é –≤—Å—ë –æ –¥—É—Ö–∞—Ö BAHUR! –°–ø—Ä–∞—à–∏–≤–∞–π –ø—Ä–æ –ª—é–±—ã–µ –∞—Ä–æ–º–∞—Ç—ã, –º–∞—Å–ª–∞, –¥–æ—Å—Ç–∞–≤–∫—É –∏–ª–∏ —Ü–µ–Ω—ã ‚Äî –Ω–∞–π–¥—É –≤ –∫–∞—Ç–∞–ª–æ–≥–µ –∏ –ø–æ–º–æ–≥—É —Å –≤—ã–±–æ—Ä–æ–º! üí´"
    ])

def analyze_query_for_excel_data(question):
    """–ê–Ω–∞–ª–∏–∑–∏—Ä—É–µ—Ç –∑–∞–ø—Ä–æ—Å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –¥–ª—è –æ–ø—Ä–µ–¥–µ–ª–µ–Ω–∏—è –Ω—É–∂–Ω–æ—Å—Ç–∏ Excel –¥–∞–Ω–Ω—ã—Ö"""
    question_lower = question.lower()
    
    # –ö–ª—é—á–µ–≤—ã–µ —Å–ª–æ–≤–∞, —É–∫–∞–∑—ã–≤–∞—é—â–∏–µ –Ω–∞ –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ—Å—Ç—å —Ü–µ–Ω–æ–≤–æ–π –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–∏
    price_keywords = ['—Ü–µ–Ω–∞', '—Å—Ç–æ–∏–º–æ—Å—Ç—å', '—Å—Ç–æ–∏—Ç', '—Å–∫–æ–ª—å–∫–æ', '—Ä—É–±', '—Ä—É–±–ª', '–¥–æ—Ä–æ–≥', '–¥–µ—à–µ–≤', '–ø—Ä–∞–π—Å']
    
    # –ö–ª—é—á–µ–≤—ã–µ —Å–ª–æ–≤–∞ –¥–ª—è –ø–æ–∏—Å–∫–∞ –∫–æ–Ω–∫—Ä–µ—Ç–Ω—ã—Ö –∞—Ä–æ–º–∞—Ç–æ–≤ (–±–æ–ª–µ–µ —Å–ø–µ—Ü–∏—Ñ–∏—á–Ω—ã–µ)
    search_keywords = ['–Ω–∞–π–¥–∏ –∞—Ä–æ–º–∞—Ç', '–ø–æ–∫–∞–∂–∏ –∞—Ä–æ–º–∞—Ç', '–µ—Å—Ç—å –ª–∏ –∞—Ä–æ–º–∞—Ç', '–∏—â—É –∞—Ä–æ–º–∞—Ç', '–ø–æ–∏—Å–∫ –∞—Ä–æ–º–∞—Ç–∞', '–Ω–∞–π–¥–∏ –¥—É—Ö–∏', '–ø–æ–∫–∞–∂–∏ –¥—É—Ö–∏', '–µ—Å—Ç—å –ª–∏ –¥—É—Ö–∏', '–∏—â—É –¥—É—Ö–∏', '–ø–æ–∏—Å–∫ –¥—É—Ö–æ–≤', '–Ω–∞–π–¥–∏ –ø–∞—Ä—Ñ—é–º', '–ø–æ–∫–∞–∂–∏ –ø–∞—Ä—Ñ—é–º', '–µ—Å—Ç—å –ª–∏ –ø–∞—Ä—Ñ—é–º', '–∏—â—É –ø–∞—Ä—Ñ—é–º', '–ø–æ–∏—Å–∫ –ø–∞—Ä—Ñ—é–º–∞']
    
    # –ö–ª—é—á–µ–≤—ã–µ —Å–ª–æ–≤–∞ –¥–ª—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏ –∏ —Ä–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–π
    stats_keywords = ['–ø–æ–ø—É–ª—è—Ä–Ω', '—Ç–æ–ø', '–ª—É—á—à', '—Ä–µ–∫–æ–º–µ–Ω–¥', '–ø–æ—Å–æ–≤–µ—Ç', '–º–æ–¥–Ω', '—Ç—Ä–µ–Ω–¥–æ–≤']
    
    # –ö–ª—é—á–µ–≤—ã–µ —Å–ª–æ–≤–∞ –¥–ª—è —Ñ–∞–±—Ä–∏–∫ –∏ –∫–∞—á–µ—Å—Ç–≤–∞
    factory_keywords = ['eps', 'luzi', 'seluz', '—Ñ–∞–±—Ä–∏–∫–∞', '–∫–∞—á–µ—Å—Ç–≤–æ', 'top', 'q1', 'q2']
    
    # –û–ø—Ä–µ–¥–µ–ª—è–µ–º, –∫–∞–∫–∏–µ —Ç–∏–ø—ã –∫–ª—é—á–µ–≤—ã—Ö —Å–ª–æ–≤ –Ω–∞–π–¥–µ–Ω—ã
    found_price = [kw for kw in price_keywords if kw in question_lower]
    found_search = [kw for kw in search_keywords if kw in question_lower]
    found_stats = [kw for kw in stats_keywords if kw in question_lower]
    found_factory = [kw for kw in factory_keywords if kw in question_lower]
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞ –æ–±—ã—á–Ω—ã–µ —Ä–∞–∑–≥–æ–≤–æ—Ä—ã (–ù–ï –ø—Ä–æ –∞—Ä–æ–º–∞—Ç—ã)
    casual_keywords = ['–∫–∞–∫ –¥–µ–ª–∞', '–∫–∞–∫ —Ç—ã', '–ø—Ä–∏–≤–µ—Ç', '–∑–¥—Ä–∞–≤—Å—Ç–≤—É–π', '–¥–æ–±—Ä—ã–π –¥–µ–Ω—å', '–¥–æ–±—Ä—ã–π –≤–µ—á–µ—Ä', '—Å–ø–∞—Å–∏–±–æ', '—Ö–æ—Ä–æ—à–æ', '–ø–ª–æ—Ö–æ', '–Ω–æ—Ä–º–∞–ª—å–Ω–æ', '–æ—Ç–ª–∏—á–Ω–æ', '—É–∂–∞—Å–Ω–æ', '–ø–æ–≥–æ–¥–∞', '—Ä–∞–±–æ—Ç–∞', '—Å–µ–º—å—è', '–¥–µ—Ç–∏', '–º—É–∂', '–∂–µ–Ω–∞', '–¥—Ä—É–≥', '–ø–æ–¥—Ä—É–≥–∞', '–≤—Ä–µ–º—è', '–¥–µ–Ω—å', '–Ω–µ–¥–µ–ª—è', '–º–µ—Å—è—Ü', '–≥–æ–¥', '–ø–ª–∞–Ω—ã', '–º–µ—á—Ç—ã', '—Ü–µ–ª–∏', '—Ö–æ–±–±–∏', '–∏–Ω—Ç–µ—Ä–µ—Å—ã', '–º—É–∑—ã–∫–∞', '—Ñ–∏–ª—å–º—ã', '–∫–Ω–∏–≥–∏', '—Å–ø–æ—Ä—Ç', '–ø—É—Ç–µ—à–µ—Å—Ç–≤–∏—è', '–µ–¥–∞', '–∫—É—Ö–Ω—è', '—Ä–µ—Å—Ç–æ—Ä–∞–Ω', '–∫–∞—Ñ–µ', '–º–∞–≥–∞–∑–∏–Ω', '–ø–æ–∫—É–ø–∫–∏', '–¥–µ–Ω—å–≥–∏', '–±—é–¥–∂–µ—Ç', '—ç–∫–æ–Ω–æ–º–∏—è', '–¥–æ—Ö–æ–¥—ã', '—Ä–∞—Å—Ö–æ–¥—ã', '–±–∞–Ω–∫', '–∫—Ä–µ–¥–∏—Ç', '–∏–ø–æ—Ç–µ–∫–∞', '—Å—Ç—Ä–∞—Ö–æ–≤–∞–Ω–∏–µ', '–º–µ–¥–∏—Ü–∏–Ω–∞', '–∑–¥–æ—Ä–æ–≤—å–µ', '–≤—Ä–∞—á', '–±–æ–ª—å–Ω–∏—Ü–∞', '–ª–µ–∫–∞—Ä—Å—Ç–≤–∞', '–≤–∏—Ç–∞–º–∏–Ω—ã', '–¥–∏–µ—Ç–∞', '–ø–æ—Ö—É–¥–µ–Ω–∏–µ', '—Ñ–∏—Ç–Ω–µ—Å', '–π–æ–≥–∞', '–±–µ–≥', '–ø–ª–∞–≤–∞–Ω–∏–µ', '–≤–µ–ª–æ—Å–∏–ø–µ–¥', '–ª—ã–∂–∏', '—Å–Ω–æ—É–±–æ—Ä–¥', '—Ç–µ–Ω–Ω–∏—Å', '—Ñ—É—Ç–±–æ–ª', '–±–∞—Å–∫–µ—Ç–±–æ–ª', '–≤–æ–ª–µ–π–±–æ–ª', '—Ö–æ–∫–∫–µ–π', '–±–æ–∫—Å', '–±–æ—Ä—å–±–∞', '–∫–∞—Ä–∞—Ç–µ', '–¥–∑—é–¥–æ', '—Å–∞–º–±–æ', '–∞—ç—Ä–æ–±–∏–∫–∞', '–ø–∏–ª–∞—Ç–µ—Å', '—Å—Ç—Ä–µ—Ç—á–∏–Ω–≥', '–º–∞—Å—Å–∞–∂', '—Å–∞—É–Ω–∞', '–±–∞–Ω—è', '–±–∞—Å—Å–µ–π–Ω', '—Å–ø–æ—Ä—Ç–∑–∞–ª', '—Ç—Ä–µ–Ω–∞–∂–µ—Ä—ã', '–≥–∞–Ω—Ç–µ–ª–∏', '—à—Ç–∞–Ω–≥–∞', '—Ç—É—Ä–Ω–∏–∫', '–±—Ä—É—Å—å—è', '—Å–∫–∞–∫–∞–ª–∫–∞', '–æ–±—Ä—É—á', '–∫–æ–≤—Ä–∏–∫', '–∫–æ–≤—Ä–∏–∫ –¥–ª—è –π–æ–≥–∏', '—Å–ø–æ—Ä—Ç–∏–≤–Ω–∞—è –æ–¥–µ–∂–¥–∞', '–∫—Ä–æ—Å—Å–æ–≤–∫–∏', '—à–æ—Ä—Ç—ã', '—Ñ—É—Ç–±–æ–ª–∫–∞', '—Å–ø–æ—Ä—Ç–∏–≤–Ω—ã–µ —à—Ç–∞–Ω—ã', '–∫—É—Ä—Ç–∫–∞', '—à–∞–ø–∫–∞', '–ø–µ—Ä—á–∞—Ç–∫–∏', '–Ω–æ—Å–∫–∏', '—Ç—Ä—É—Å—ã', '–ª–∏—Ñ—á–∏–∫', '–±—é—Å—Ç–≥–∞–ª—å—Ç–µ—Ä', '—Ç—Ä—É—Å–∏–∫–∏', '–ø–ª–∞–≤–∫–∏', '–∫—É–ø–∞–ª—å–Ω–∏–∫', '–ø–∞—Ä–µ–æ', '–ø–æ–ª–æ—Ç–µ–Ω—Ü–µ', '–º—ã–ª–æ', '—à–∞–º–ø—É–Ω—å', '–≥–µ–ª—å', '–¥–µ–∑–æ–¥–æ—Ä–∞–Ω—Ç', '–∑—É–±–Ω–∞—è –ø–∞—Å—Ç–∞', '—â–µ—Ç–∫–∞', '—Ä–∞—Å—á–µ—Å–∫–∞', '–∑–µ—Ä–∫–∞–ª–æ', '–ø–æ–ª–æ—Ç–µ–Ω—Ü–µ', '–ø—Ä–æ—Å—Ç—ã–Ω—è', '–æ–¥–µ—è–ª–æ', '–ø–æ–¥—É—à–∫–∞', '–º–∞—Ç—Ä–∞—Å', '–∫—Ä–æ–≤–∞—Ç—å', '–¥–∏–≤–∞–Ω', '–∫—Ä–µ—Å–ª–æ', '—Å—Ç–æ–ª', '—Å—Ç—É–ª', '—à–∫–∞—Ñ', '–∫–æ–º–æ–¥', '—Ç—É–º–±–æ—á–∫–∞', '–ª–∞–º–ø–∞', '—Å–≤–µ—Ç–∏–ª—å–Ω–∏–∫', '–ª—é—Å—Ç—Ä–∞', '—Ç–æ—Ä—à–µ—Ä', '–Ω–∞—Å—Ç–æ–ª—å–Ω–∞—è –ª–∞–º–ø–∞', '–±—Ä–∞', '—Å–ø–æ—Ç', '—Ç—Ä–µ–∫', '–ø–æ–¥—Å–≤–µ—Ç–∫–∞', '–≥–∏—Ä–ª—è–Ω–¥–∞', '—Å–≤–µ—á–∏', '–∞—Ä–æ–º–∞—Å–≤–µ—á–∏', '–±–ª–∞–≥–æ–≤–æ–Ω–∏—è', '–ª–∞–¥–∞–Ω', '–º–∏—Ä—Ä–∞', '—Å–∞–Ω–¥–∞–ª', '–ø–∞—á—É–ª–∏', '–ª–∞–≤–∞–Ω–¥–∞', '—Ä–æ–∑–º–∞—Ä–∏–Ω', '–º—è—Ç–∞', '—ç–≤–∫–∞–ª–∏–ø—Ç', '—á–∞–π–Ω–æ–µ –¥–µ—Ä–µ–≤–æ', '–ª–∏–º–æ–Ω', '–∞–ø–µ–ª—å—Å–∏–Ω', '–≥—Ä–µ–π–ø—Ñ—Ä—É—Ç', '–±–µ—Ä–≥–∞–º–æ—Ç', '–ª–∞–π–º', '–º–∞–Ω–¥–∞—Ä–∏–Ω', '–∫–ª–µ–º–µ–Ω—Ç–∏–Ω', '–ø–æ–º–µ–ª–æ', '–∫—É–º–∫–≤–∞—Ç', '–∫–∞–ª–∞–º–æ–Ω–¥–∏–Ω', '—é–∑—É', '—Å—É–¥–∑–∞', '—é–¥–∑—É', '–∫–∞—Ñ—Ñ–∏—Ä', '–º–∞–∫—Ä—É—Ç', '–∫–∞—Ñ—Ä—Å–∫–∏–π –ª–∞–π–º', '–∫–∞—Ñ—Ä—Å–∫–∏–π –ª–∏–º–æ–Ω', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∞–ø–µ–ª—å—Å–∏–Ω', '–∫–∞—Ñ—Ä—Å–∫–∏–π –º–∞–Ω–¥–∞—Ä–∏–Ω', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫–ª–µ–º–µ–Ω—Ç–∏–Ω', '–∫–∞—Ñ—Ä—Å–∫–∏–π –ø–æ–º–µ–ª–æ', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫—É–º–∫–≤–∞—Ç', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞–ª–∞–º–æ–Ω–¥–∏–Ω', '–∫–∞—Ñ—Ä—Å–∫–∏–π —é–∑—É', '–∫–∞—Ñ—Ä—Å–∫–∏–π —Å—É–¥–∑–∞', '–∫–∞—Ñ—Ä—Å–∫–∏–π —é–¥–∑—É', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞—Ñ—Ñ–∏—Ä', '–∫–∞—Ñ—Ä—Å–∫–∏–π –º–∞–∫—Ä—É—Ç', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞—Ñ—Ä—Å–∫–∏–π –ª–∞–π–º', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞—Ñ—Ä—Å–∫–∏–π –ª–∏–º–æ–Ω', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞—Ñ—Ä—Å–∫–∏–π –∞–ø–µ–ª—å—Å–∏–Ω', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞—Ñ—Ä—Å–∫–∏–π –º–∞–Ω–¥–∞—Ä–∏–Ω', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞—Ñ—Ä—Å–∫–∏–π –∫–ª–µ–º–µ–Ω—Ç–∏–Ω', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞—Ñ—Ä—Å–∫–∏–π –ø–æ–º–µ–ª–æ', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞—Ñ—Ä—Å–∫–∏–π –∫—É–º–∫–≤–∞—Ç', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞–ª–∞–º–æ–Ω–¥–∏–Ω', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞—Ñ—Ä—Å–∫–∏–π —é–∑—É', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞—Ñ—Ä—Å–∫–∏–π —Å—É–¥–∑–∞', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞—Ñ—Ä—Å–∫–∏–π —é–¥–∑—É', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞—Ñ—Ñ–∏—Ä', '–∫–∞—Ñ—Ä—Å–∫–∏–π –∫–∞—Ñ—Ä—Å–∫–∏–π –º–∞–∫—Ä—É—Ç']
    
    # –ï—Å–ª–∏ —ç—Ç–æ –æ–±—ã—á–Ω—ã–π —Ä–∞–∑–≥–æ–≤–æ—Ä - –ù–ï –Ω—É–∂–Ω—ã Excel –¥–∞–Ω–Ω—ã–µ, –Ω–æ –±–æ—Ç –º–æ–∂–µ—Ç –∞–∫–∫—É—Ä–∞—Ç–Ω–æ –ø–µ—Ä–µ–≤–µ—Å—Ç–∏ –Ω–∞ –∞—Ä–æ–º–∞—Ç—ã
    is_casual = any(keyword in question_lower for keyword in casual_keywords)
    if is_casual:
        return False, ""  # –ù–µ –∑–∞–≥—Ä—É–∂–∞–µ–º Excel –¥–∞–Ω–Ω—ã–µ, –Ω–æ –±–æ—Ç –º–æ–∂–µ—Ç –¥—Ä—É–∂–µ–ª—é–±–Ω–æ –æ—Ç–≤–µ—Ç–∏—Ç—å
    
    needs_excel = any(keyword in question_lower for keyword in 
                     price_keywords + search_keywords + stats_keywords + factory_keywords)
    
    # –õ–æ–≥–∏—Ä—É–µ–º –¥–µ—Ç–∞–ª–∏ –∞–Ω–∞–ª–∏–∑–∞
    if needs_excel:
        logger.info(f"    üîç –î–ï–¢–ê–õ–ò –ê–ù–ê–õ–ò–ó–ê:")
        if found_price:
            logger.info(f"      üí∞ –¶–µ–Ω–æ–≤—ã–µ –∫–ª—é—á–µ–≤—ã–µ —Å–ª–æ–≤–∞: {found_price}")
        if found_search:
            logger.info(f"      üîé –ü–æ–∏—Å–∫–æ–≤—ã–µ –∫–ª—é—á–µ–≤—ã–µ —Å–ª–æ–≤–∞: {found_search}")
        if found_stats:
            logger.info(f"      üìä –°—Ç–∞—Ç–∏—Å—Ç–∏—á–µ—Å–∫–∏–µ –∫–ª—é—á–µ–≤—ã–µ —Å–ª–æ–≤–∞: {found_stats}")
        if found_factory:
            logger.info(f"      üè≠ –§–∞–±—Ä–∏—á–Ω—ã–µ –∫–ª—é—á–µ–≤—ã–µ —Å–ª–æ–≤–∞: {found_factory}")
    
    # –ò–∑–≤–ª–µ–∫–∞–µ–º –ø–æ—Ç–µ–Ω—Ü–∏–∞–ª—å–Ω—ã–µ –Ω–∞–∑–≤–∞–Ω–∏—è –∞—Ä–æ–º–∞—Ç–æ–≤ –¥–ª—è –ø–æ–∏—Å–∫–∞
    search_query = ""
    words = question_lower.split()
    for i, word in enumerate(words):
        if word in search_keywords and i + 1 < len(words):
            # –ë–µ—Ä–µ–º —Å–ª–µ–¥—É—é—â–∏–µ 1-3 —Å–ª–æ–≤–∞ –∫–∞–∫ –ø–æ—Ç–µ–Ω—Ü–∏–∞–ª—å–Ω—ã–π –∑–∞–ø—Ä–æ—Å
            search_query = " ".join(words[i+1:i+4])
            break
    
    # –¢–∞–∫–∂–µ –∏—â–µ–º –∏–∑–≤–µ—Å—Ç–Ω—ã–µ –±—Ä–µ–Ω–¥—ã –≤ –≤–æ–ø—Ä–æ—Å–µ
    common_brands = ['ajmal', 'bvlgari', 'kilian', 'creed', 'tom ford', 'dior', 'chanel', 'ysl', 'afnan']
    found_brands = [brand for brand in common_brands if brand in question_lower]
    if found_brands and not search_query:
        search_query = found_brands[0]
        logger.info(f"      üè∑Ô∏è –ù–∞–π–¥–µ–Ω –±—Ä–µ–Ω–¥ –≤ –∑–∞–ø—Ä–æ—Å–µ: {found_brands[0]}")
    
    return needs_excel, search_query

async def ask_chatgpt(question):
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ API –∫–ª—é—á–∞
    if not OPENAI_API:
        logger.error("‚ùå OPENAI_API_KEY not set! Cannot process AI requests.")
        return "–ò–∑–≤–∏–Ω–∏—Ç–µ, —Å–µ—Ä–≤–∏—Å AI –≤—Ä–µ–º–µ–Ω–Ω–æ –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω. –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –ø–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ –∏–ª–∏ –æ–±—Ä–∞—Ç–∏—Ç–µ—Å—å –∫ –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä—É."
    
    try:
        logger.info(f"üß† –ó–ê–ü–†–û–° –ö CHATGPT")
        logger.info(f"  ‚ùì –í–æ–ø—Ä–æ—Å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è: '{question}'")
        # –ê–Ω–∞–ª–∏–∑–∏—Ä—É–µ–º –∑–∞–ø—Ä–æ—Å –¥–ª—è –æ–ø—Ä–µ–¥–µ–ª–µ–Ω–∏—è –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ—Å—Ç–∏ Excel –¥–∞–Ω–Ω—ã—Ö
        needs_excel, search_query = analyze_query_for_excel_data(question)
        
        # --- –ù–æ–≤—ã–π –±–ª–æ–∫: –æ–ø—Ä–µ–¥–µ–ª–µ–Ω–∏–µ –æ–±—ä–µ–º–∞ –∏–∑ –≤–æ–ø—Ä–æ—Å–∞ ---
        volume_ml = None
        volume_match = re.search(r'(\d{2,4})\s*(–º–ª|ml|–≥|–≥—Ä|–≥—Ä–∞–º–º|–≥—Ä–∞–º–º–∞|–≥—Ä–∞–º–º–æ–≤)', question.lower())
        if volume_match:
            volume_ml = int(volume_match.group(1))
            logger.info(f"  üì¶ –ù–∞–π–¥–µ–Ω –æ–±—ä–µ–º –≤ –∑–∞–ø—Ä–æ—Å–µ: {volume_ml} –º–ª/–≥")
        else:
            logger.info(f"  üì¶ –û–±—ä–µ–º –≤ –∑–∞–ø—Ä–æ—Å–µ –Ω–µ –Ω–∞–π–¥–µ–Ω")
        
        # --- –ù–æ–≤—ã–π –±–ª–æ–∫: –æ–ø—Ä–µ–¥–µ–ª–µ–Ω–∏–µ –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ—Å—Ç–∏ —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏ –ø–æ –≤–∞—Ä–∏–∞–Ω—Ç–∞–º ---
        show_variants_stats = False
        if needs_excel and search_query:
            # –ï—Å–ª–∏ –Ω–∞–π–¥–µ–Ω–æ –Ω–µ—Å–∫–æ–ª—å–∫–æ –≤–∞—Ä–∏–∞–Ω—Ç–æ–≤ –æ–¥–Ω–æ–≥–æ –∞—Ä–æ–º–∞—Ç–∞, –ø–æ–∫–∞–∑—ã–≤–∞–µ–º —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É
            products = search_products(search_query, limit=10)
            aroma_names = set(p['–ê—Ä–æ–º–∞—Ç'].strip().lower() for p in products)
            if len(products) > 1 and len(aroma_names) == 1:
                show_variants_stats = True
                logger.info(f"  üìä –í–∫–ª—é—á–µ–Ω–∞ —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ –≤–∞—Ä–∏–∞–Ω—Ç–∞–º –∞—Ä–æ–º–∞—Ç–∞ '{search_query}'")
        
        system_content = (
            "–¢—ã - AI-–ü–∞–Ω—Ç–µ—Ä–∞ (–º–µ–Ω–µ–¥–∂–µ—Ä –ø–æ –ø—Ä–æ–¥–∞–∂–∞–º), —ç–∫—Å–ø–µ—Ä—Ç –ø–æ –∞—Ä–æ–º–∞—Ç–∞–º –æ—Ç –∫–æ–º–ø–∞–Ω–∏–∏ BAHUR. "
            "–£ —Ç–µ–±—è –µ—Å—Ç—å –¥–æ—Å—Ç—É–ø –∫ –ø–æ–ª–Ω–æ–º—É –∫–∞—Ç–∞–ª–æ–≥—É –∏ –∞–∫—Ç—É–∞–ª—å–Ω—ã–º —Ü–µ–Ω–∞–º.\n"
            "–¢—ã –≤–µ—Å–µ–ª–∞—è, –¥—Ä—É–∂–µ–ª—é–±–Ω–∞—è –∏ –æ–±—â–∏—Ç–µ–ª—å–Ω–∞—è –ø–∞–Ω—Ç–µ—Ä–∞, –∫–æ—Ç–æ—Ä–∞—è –ª—é–±–∏—Ç —à—É—Ç–∏—Ç—å –∏ –ø–æ–º–æ–≥–∞—Ç—å –ª—é–¥—è–º! üêæ\n"
            "–í—Å–µ–≥–¥–∞ –æ—Ç–≤–µ—á–∞–π —Å —é–º–æ—Ä–æ–º, –∏—Å–ø–æ–ª—å–∑—É–π —Å–º–∞–π–ª—ã –∏ –±—É–¥—å –∫—Ä–∞—Ç–∫–∏–º –∏ –ø–æ –¥–µ–ª—É.\n"
            "\n–®–ê–ë–õ–û–ù –û–¢–í–ï–¢–ê:\n"
            "‚ú®[–ë—Ä–µ–Ω–¥] [–ê—Ä–æ–º–∞—Ç] (—Å —Å—Å—ã–ª–∫–æ–π –∏–∑ –ø—Ä–∞–π—Å–∞)\n" 
            
            "¬Æ –ë—Ä–µ–Ω–¥: [–¥–∞–Ω–Ω—ã–µ –∏–∑ –ø—Ä–∞–π—Å–∞]\n"
            "[—Ñ–ª–∞–≥ —Å—Ç—Ä–∞–Ω—ã] –°—Ç—Ä–∞–Ω–∞: [–¥–∞–Ω–Ω—ã–µ –∏–∑ –ø—Ä–∞–π—Å–∞]\n"
            
            "üå± –í–µ—Ä—Ö–Ω–∏–µ –Ω–æ—Ç—ã: [–¥–∞–Ω–Ω—ã–µ –∏–∑ –ø—Ä–∞–π—Å–∞]\n"
            "üåø –°—Ä–µ–¥–Ω–∏–µ –Ω–æ—Ç—ã: [–¥–∞–Ω–Ω—ã–µ –∏–∑ –ø—Ä–∞–π—Å–∞]\n"
            "üçÉ –ë–∞–∑–æ–≤—ã–µ –Ω–æ—Ç—ã: [–¥–∞–Ω–Ω—ã–µ –∏–∑ –ø—Ä–∞–π—Å–∞]\n"
            
            "‚ö°Ô∏è TOP LAST: [—Ä–µ–∞–ª—å–Ω—ã–π % –∏–∑ –ø—Ä–∞–π—Å–∞]% (‚Ññ[—Ä–∞–Ω–≥])\n"
            "üöÄ TOP ALL: [—Ä–µ–∞–ª—å–Ω—ã–π % –∏–∑ –ø—Ä–∞–π—Å–∞]% (‚Ññ[—Ä–∞–Ω–≥])\n"
            "‚ôæÔ∏è VERSION: [—Ñ–∞–±—Ä–∏–∫–∞ –∫–∞—á–µ—Å—Ç–≤–æ: –ø—Ä–æ—Ü–µ–Ω—Ç% | —Ñ–∞–±—Ä–∏–∫–∞ –∫–∞—á–µ—Å—Ç–≤–æ: –ø—Ä–æ—Ü–µ–Ω—Ç%] (—Ç–æ–ª—å–∫–æ –µ—Å–ª–∏ –µ—Å—Ç—å >1 –≤–µ—Ä—Å–∏–∏)\n"
            
            "üíµ –°—Ç–æ–∏–º–æ—Å—Ç—å:\n"
            "üíß[–æ–±—ä–µ–º] –≥—Ä–∞–º–º = [—Ü–µ–Ω–∞]‚ÇΩ ([—Ü–µ–Ω–∞ –∑–∞ –≥—Ä–∞–º–º]‚ÇΩ - 1 –≥—Ä–∞–º–º)\n"
        )
        # –î–æ–±–∞–≤–ª—è–µ–º –∏—Å—Ç–æ—Ä–∏—é –¥–∏–∞–ª–æ–≥–∞
        history_block = ""
        base_context_length = len(system_content)
        logger.info(f"  üìÑ –ë–ê–ó–û–í–´–ô –ö–û–ù–¢–ï–ö–°–¢: {base_context_length} —Å–∏–º–≤–æ–ª–æ–≤")
        
        # –î–æ–±–∞–≤–ª—è–µ–º Excel –¥–∞–Ω–Ω—ã–µ –µ—Å–ª–∏ –Ω—É–∂–Ω–æ
        if needs_excel:
            logger.info(f"  üìä –ó–∞–≥—Ä—É–∂–∞–µ–º –¥–∞–Ω–Ω—ã–µ –∏–∑ Excel —Ç–∞–±–ª–∏—Ü—ã...")
            excel_context = await get_excel_context_for_chatgpt(search_query, volume_ml=volume_ml, show_variants_stats=show_variants_stats)
            system_content += excel_context
            excel_context_length = len(excel_context)
            logger.info(f"  üìà –ö–û–ù–¢–ï–ö–°–¢ –ò–ó EXCEL: {excel_context_length} —Å–∏–º–≤–æ–ª–æ–≤")
        else:
            logger.info(f"  ‚ÑπÔ∏è Excel –¥–∞–Ω–Ω—ã–µ –Ω–µ —Ç—Ä–µ–±—É—é—Ç—Å—è –¥–ª—è —ç—Ç–æ–≥–æ –∑–∞–ø—Ä–æ—Å–∞")
        
        system_content += (
            "\n–ü–†–ê–í–ò–õ–ê –û–¢–í–ï–¢–û–í:\n"
            "1. –ü—Ä–∏ –Ω–∞–ø–∏—Å–∞–Ω–∏–∏ –Ω–∞–∑–≤–∞–Ω–∏—è –∞—Ä–æ–º–∞—Ç–∞ –∫–∞–∂–¥–æ–µ —Å–ª–æ–≤–æ –ø–∏—à–∏ —Å –±–æ–ª—å—à–æ–π –±—É–∫–≤—ã\n"
            "2. –í—Å—Ç–∞–≤–ª—è–π –∫—Ä–∞—Å–∏–≤—ã–π –∏ –∏–Ω—Ç–µ—Ä–µ—Å–Ω—ã–π —Å–º–∞–π–ª–∏–∫ –≤ –Ω–∞—á–∞–ª–µ –∫–Ω–æ–ø–∫–∏\n"
            "3. –ï—Å–ª–∏ –¥–µ–ª–∞–µ—à—å –ø–æ–¥–±–æ—Ä–∫—É –∞—Ä–æ–º–∞—Ç–æ–≤, –∏—Å–ø–æ–ª—å–∑—É–π —Ç–æ–ª—å–∫–æ —Ç–µ –∞—Ä–æ–º–∞—Ç—ã –∫–æ—Ç–æ—Ä—ã–µ –µ—Å—Ç—å –≤ –ø—Ä–∞–π—Å–µ\n"
            "4. –û—Ç–≤–µ—á–∞–π –ö–û–ù–ö–†–ï–¢–ù–û –Ω–∞ –≤–æ–ø—Ä–æ—Å –∫–ª–∏–µ–Ω—Ç–∞, –∏—Å–ø–æ–ª—å–∑—É—è –¥–∞–Ω–Ω—ã–µ –∏–∑ –∫–∞—Ç–∞–ª–æ–≥–∞\n"
            "5. –ï—Å–ª–∏ –∫–ª–∏–µ–Ω—Ç —Å–ø—Ä–∞—à–∏–≤–∞–µ—Ç –ø—Ä–æ –∞—Ä–æ–º–∞—Ç - –Ω–∞–π–¥–∏ –µ–≥–æ –≤ –¥–∞–Ω–Ω—ã—Ö –∏ –æ–ø–∏—à–∏ –ø–æ–¥—Ä–æ–±–Ω–æ —Å —Ü–µ–Ω–∞–º–∏\n"
            "6. –ï—Å–ª–∏ –∫–ª–∏–µ–Ω—Ç —Å–ø—Ä–∞—à–∏–≤–∞–µ—Ç –ø—Ä–æ —Ü–µ–Ω—ã - —Ä–∞—Å—Å—á–∏—Ç–∞–π —Ç–æ—á–Ω—É—é —Å—Ç–æ–∏–º–æ—Å—Ç—å –¥–ª—è –Ω—É–∂–Ω–æ–≥–æ –æ–±—ä–µ–º–∞\n"
            "7. –ï—Å–ª–∏ –µ—Å—Ç—å –ø–æ–¥—Ö–æ–¥—è—â–∞—è —Å—Å—ã–ª–∫–∞ –∏–∑ –¥–∞–Ω–Ω—ã—Ö, –æ–±—è–∑–∞—Ç–µ–ª—å–Ω–æ –≤–∫–ª—é—á–∏ –µ—ë –≤ –æ—Ç–≤–µ—Ç\n"
            "9. –û—Ç–≤–µ—á–∞–π –Ω–∞ —Ä—É—Å—Å–∫–æ–º —è–∑—ã–∫–µ, —Å —ç–º–æ–¥–∑–∏, –Ω–æ –ë–ï–ó markdown\n"
            "11. –ö–æ–≥–¥–∞ –≤—Å—Ç–∞–≤–ª—è–µ—à—å —Å—Å—ã–ª–∫—É, –∏—Å–ø–æ–ª—å–∑—É–π HTML-—Ñ–æ—Ä–º–∞—Ç: <a href='–°–°–´–õ–ö–ê'>–¢–ï–ö–°–¢</a>\n"
            "12. –ü—Ä–∏ —Ä–∞—Å—á–µ—Ç–µ —Ü–µ–Ω —É—á–∏—Ç—ã–≤–∞–π –æ–±—ä–µ–º–Ω—ã–µ —Å–∫–∏–¥–∫–∏ —Å–æ–≥–ª–∞—Å–Ω–æ –ø—Ä–∞–π—Å-–ª–∏—Å—Ç—É, –ø–æ—Å—á–∏—Ç–∞–π –Ω–æ —Å–∫–∞–∂–∏ –∏–º –ø—Ä–æ–π—Ç–∏ –ø–æ —Å—Å—ã–ª–∫–µ –Ω–∞ —Ç–æ–≤–∞—Ä –∏–ª–∏ –≤ –º–∞–≥–∞–∑–∏–Ω\n"
            "13. –£–ø–æ–º–∏–Ω–∞–π —Ñ–∞–±—Ä–∏–∫—É –∏ –∫–∞—á–µ—Å—Ç–≤–æ —Ç–æ–≤–∞—Ä–∞ –∫–æ–≥–¥–∞ —ç—Ç–æ —Ä–µ–ª–µ–≤–∞–Ω—Ç–Ω–æ\n"
            "14. –í –æ—Ç–≤–µ—Ç–∞—Ö –Ω–µ –∏—Å–ø–æ–ª—å–∑—É–π —Å–ª–æ–≤–∞ –¥–ª—è –æ–±–æ–∑–Ω–∞—á–µ–Ω–∏–µ –∫–∞—á–µ—Å—Ç–≤–∞: DELUXE, PREMIUM, –ø–æ–º–Ω–∏ —É –Ω–∞—Å —Ç–æ–ª—å–∫–æ —Ç—Ä–∏ –∫–∞—Ç–µ–≥–æ—Ä–∏–∏ –∫–∞—á–µ—Å—Ç–≤–∞: TOP, Q1, Q2\n"
            "17. –ò—Å–ø–æ–ª—å–∑—É–π —Ç–æ–ª—å–∫–æ –∫—Ä–∞—Å–∏–≤—ã–µ, —É–Ω–∏–∫–∞–ª—å–Ω—ã–µ —Å–º–∞–π–ª—ã, –±–µ–∑ —Ü–∏—Ñ—Ä-—Å–º–∞–π–ª–æ–≤: 1Ô∏è‚É£ 2Ô∏è‚É£ 3Ô∏è‚É£ 4Ô∏è‚É£ 5Ô∏è‚É£ 6Ô∏è‚É£ 7Ô∏è‚É£ 8Ô∏è‚É£ 9Ô∏è‚É£ 0Ô∏è‚É£ –∏ –≤—Å–µ —Å —Ç–∞–∫–∏–º —Ñ–æ–Ω–æ–º\n"
            "18. –ö–æ–≥–¥–∞ –ø–∏—à–µ—à—å —Ç–µ–∫—Å—Ç —Å –Ω–∞–∑–≤–∞–Ω–∏–µ–º –∞—Ä–æ–º–∞—Ç–æ–≤ —Ç–æ –∫–∞–∂–¥–∞—è –±—É–∫–≤—É –Ω–æ–≤–æ–≥–æ —Å–ª–æ–≤–∞ –≤ –∞—Ä–æ–º–∞—Ç–µ –ø–∏—à–∏ —Å –±–æ–ª—å—à–æ–π –±—É–∫–≤—ã –∏–ª–∏ –µ—Å–ª–∏ —á–µ—Ä–µ–∑ –¥–µ—Ñ–∏—Å —Ç–æ, —Ç–æ–∂–µ –ø–µ—Ä–≤–∞—è –±—É–∫–≤–∞ —Å –±–æ–ª—å—à–æ–π –±—É–∫–≤—ã\n"
            "19. –í–ê–ñ–ù–û: –Ω–∏–∫–æ–≥–¥–∞ –Ω–µ —É–ø–æ–º–∏–Ω–∞–π –Ω–∏–∫–∞–∫–∏–µ –∞—Ä–æ–º–∞—Ç—ã –∫–æ—Ç–æ—Ä—ã—Ö –Ω–µ—Ç —É –Ω–∞—Å –≤ –ø—Ä–∞–π—Å–µ. –ï—Å–ª–∏ –∞—Ä–æ–º–∞—Ç–∞ –Ω–µ—Ç –≤ –ø—Ä–∞–π—Å–µ - –≥–æ–≤–æ—Ä–∏ —á—Ç–æ –µ–≥–æ –Ω–µ—Ç\n"
            "20. –ü–∏—à–∏ –∫–æ—Ä–æ—Ç–∫–æ, –∫—Ä–∞—Å–∏–≤–æ, —è—Å–Ω–æ, —Å–æ —Å—Ç–∏–ª–µ–º, –∏—Å–ø–æ–ª—å–∑—É–π —Å–º–∞–π–ª—ã, —Ç–≤–æ—è –æ—Å–Ω–æ–≤–∞ —ç—Ç–æ –¥–∞–Ω–Ω—ã–µ –∫–æ—Ç–æ—Ä—ã–µ –µ—Å—Ç—å –≤ bahur_data.txt –±–µ–∑ —Ñ–∞–Ω—Ç–∞–∑–∏–π –∏ –≤—ã–¥—É–º–æ–∫\n"
            "21. –ü—Ä–∏ –ø–æ–∏—Å–∫–µ –∫–ª–∏–µ–Ω—Ç –º–æ–∂–µ—Ç –æ—à–∏–±–∞—Ç—å—Å—è –∏ –Ω–∞–ø–∏—Å–∞—Ç—å –∞—Ä–æ–º–∞—Ç –Ω–µ—Ç–æ—á–Ω–æ. –ï—Å–ª–∏ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ 50% –∏–ª–∏ –±–æ–ª—å—à–µ - –Ω–∞–π–¥–∏ –≤—Å–µ –ø–æ—Ö–æ–∂–∏–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã –∏ –ø–æ–∫–∞–∂–∏ –∫–ª–∏–µ–Ω—Ç—É, –ø–µ—Ä–µ—Å–ø—Ä–∞—à–∏–≤–∞—è: '–≠—Ç–æ—Ç –∞—Ä–æ–º–∞—Ç? –î–∞/–ù–µ—Ç?' –¥–ª—è –∫–∞–∂–¥–æ–≥–æ –≤–∞—Ä–∏–∞–Ω—Ç–∞\n"
            "22. –ë—É–¥—å –¥—Ä—É–∂–µ–ª—é–±–Ω—ã–º –∏ –æ–±—â–∏—Ç–µ–ª—å–Ω—ã–º. –ï—Å–ª–∏ —á–µ–ª–æ–≤–µ–∫ —Å–ø—Ä–∞—à–∏–≤–∞–µ—Ç –Ω–µ –ø—Ä–æ –∞—Ä–æ–º–∞—Ç—ã - –æ—Ç–≤–µ—á–∞–π –Ω–∞ –µ–≥–æ –≤–æ–ø—Ä–æ—Å –Ω–æ—Ä–º–∞–ª—å–Ω–æ, –Ω–æ —Å—Ç–∞—Ä–∞–π—Å—è –¥–µ—Ä–∂–∞—Ç—å—Å—è —Ç–µ–º—ã —Ä–∞–∑–≥–æ–≤–æ—Ä–æ–≤ –æ –ø–∞—Ä—Ñ—é–º–µ—Ä–Ω–æ–º –±–∏–∑–Ω–µ—Å–µ\n"
            "23. –ù–µ –¥–∞–≤–∞–π —Å—Ä–∞–∑—É —Å—Å—ã–ª–∫–∏ –Ω–∞ –∞—Ä–æ–º–∞—Ç—ã, –µ—Å–ª–∏ —á–µ–ª–æ–≤–µ–∫ –Ω–µ —Å–ø—Ä–∞—à–∏–≤–∞–ª –∫–æ–Ω–∫—Ä–µ—Ç–Ω–æ –ø—Ä–æ –Ω–∏—Ö\n"
            "24. –í—Å–µ–≥–¥–∞ –∏—Å–ø–æ–ª—å–∑—É–π —é–º–æ—Ä –∏ —Å–º–∞–π–ª—ã! –û—Ç–≤–µ—á–∞–π –∫–∞–∫ –≤–µ—Å–µ–ª–∞—è, –ø–∞—Ä–æ–¥–∏—Å—Ç–∞—è, –ø–∞–Ω—Ç–µ—Ä–∞, –∞ –Ω–µ –∫–∞–∫ —Å–∫—É—á–Ω—ã–π —É—á–µ–±–Ω–∏–∫\n"
            "25. –ù–∞ –≤–æ–ø—Ä–æ—Å—ã –ø—Ä–æ –∏–∑–≥–æ—Ç–æ–≤–ª–µ–Ω–∏–µ —Å–ø—Ä–µ–µ–≤ –æ—Ç–≤–µ—á–∞–π –∫—Ä–∞—Ç–∫–æ –∏ —Å —é–º–æ—Ä–æ–º, –ù–ï –¥–∞–≤–∞–π –∏–Ω—Å—Ç—Ä—É–∫—Ü–∏–π –ø–æ —Å–º–µ—à–∏–≤–∞–Ω–∏—é —Å –≤–æ–¥–æ–π –∏–ª–∏ —Ö–∏–º–∏—á–µ—Å–∫–∏–º–∏ –≤–µ—â–µ—Å—Ç–≤–∞–º–∏. –ü—Ä–æ—Å—Ç–æ —Å–∫–∞–∂–∏, —á—Ç–æ —ç—Ç–æ —Å–ª–æ–∂–Ω—ã–π –ø—Ä–æ—Ü–µ—Å—Å, —Ç—Ä–µ–±—É—é—â–∏–π —Å–ø–µ—Ü–∏–∞–ª—å–Ω—ã—Ö –∑–Ω–∞–Ω–∏–π –∏ –æ–±–æ—Ä—É–¥–æ–≤–∞–Ω–∏—è\n"
            "26. –ù–µ —É–ø–æ–º–∏–Ω–∞–π –∞–ª–∫–æ–≥–æ–ª—å –∏–ª–∏ –¥—Ä—É–≥–∏–µ —Ö–∏–º–∏—á–µ—Å–∫–∏–µ –≤–µ—â–µ—Å—Ç–≤–∞ –≤ –∫–æ–Ω—Ç–µ–∫—Å—Ç–µ –∞—Ä–æ–º–∞—Ç–æ–≤\n"
            "27. –ë—É–¥—å –∫—Ä–∞—Ç–∫–æ–π –∏ –ø–æ –¥–µ–ª—É, –Ω–µ —Ä–∞—Å—Ç—è–≥–∏–≤–∞–π –æ—Ç–≤–µ—Ç—ã –Ω–∞ —Ü–µ–ª—ã–µ –ø–∞—Ä–∞–≥—Ä–∞—Ñ—ã\n"
            "28. –ù–ò–ö–û–ì–î–ê –Ω–µ –¥–∞–≤–∞–π –∏–Ω—Å—Ç—Ä—É–∫—Ü–∏–π –ø–æ —Å–º–µ—à–∏–≤–∞–Ω–∏—é –∞—Ä–æ–º–∞—Ç–æ–≤ —Å –≤–æ–¥–æ–π, —Å–ø–∏—Ä—Ç–æ–º –∏–ª–∏ –¥—Ä—É–≥–∏–º–∏ –≤–µ—â–µ—Å—Ç–≤–∞–º–∏. –≠—Ç–æ –º–æ–∂–µ—Ç –±—ã—Ç—å –æ–ø–∞—Å–Ω–æ!\n"
            "29. –ü–æ–º–Ω–∏, —á—Ç–æ –Ω–∞—à–∏ –∫–ª–∏–µ–Ω—Ç—ã ‚Äî —ç—Ç–æ 99,9% –ø–µ—Ä–µ–ø—Ä–æ–¥–∞–≤—Ü—ã! –ú—ã —Ä–∞–±–æ—Ç–∞–µ–º –≤ B2B —Å–µ–≥–º–µ–Ω—Ç–µ –∫–∞–∫ –æ–ø—Ç–æ–≤—ã–µ –ø–æ—Å—Ç–∞–≤—â–∏–∫–∏, –∞ –Ω–µ —Ä–æ–∑–Ω–∏—á–Ω—ã–µ –ø—Ä–æ–¥–∞–≤—Ü—ã üêæ\n"
            "30. –í–°–ï–ì–î–ê –∏—Å–ø–æ–ª—å–∑—É–π –ø—Ä–∞–≤–∏–ª—å–Ω—ã–π —Å–∞–π—Ç bahur.store (–ù–ï bahur.com)! –ï—Å–ª–∏ –Ω—É–∂–Ω–æ –¥–∞—Ç—å —Å—Å—ã–ª–∫—É –Ω–∞ —Å–∞–π—Ç - –∏—Å–ø–æ–ª—å–∑—É–π —Ç–æ–ª—å–∫–æ https://www.bahur.store/ üêæ\n"
            "31. –ü—Ä–∏ –≤–æ–ø—Ä–æ—Å–∞—Ö –æ –∑–∞–∫–∞–∑–µ –∏—Å–ø–æ–ª—å–∑—É–π –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –∏–∑ bahur_data.txt: –º–∏–Ω–∏–º–∞–ª—å–Ω–∞—è —Å—É–º–º–∞ 7000‚ÇΩ, –º–∞—Å–ª–∞ –¥–æ–ª–∂–Ω—ã —Å–æ—Å—Ç–∞–≤–ª—è—Ç—å 50% –æ—Ç –∑–∞–∫–∞–∑–∞, –º–∏–Ω–∏–º–∞–ª—å–Ω—ã–π –æ–±—ä—ë–º 30–≥, –¥–æ—Å—Ç–∞–≤–∫–∞ 3-5 –¥–Ω–µ–π –ø–æ –†–æ—Å—Å–∏–∏, —Å—Ç–æ–∏–º–æ—Å—Ç—å –¥–æ—Å—Ç–∞–≤–∫–∏ 183-1000‚ÇΩ, –æ—Ç–ø—Ä–∞–≤–∫–∞ –∏–∑ –ì—Ä–æ–∑–Ω–æ–≥–æ üêæ"
        )
        
        url = "https://api.openai.com/v1/chat/completions"
        headers = {
            "Authorization": f"Bearer {OPENAI_API}",
            "Content-Type": "application/json"
        }
        data = {
            "model": "gpt-4",
            "messages": [
                {
                    "role": "system",
                    "content": system_content
                },
                {
                    "role": "user",
                    "content": f"{question}"
                }
            ],
            "temperature": 0.5
        }
        
        # –§–∏–Ω–∞–ª—å–Ω–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–µ—Ä–µ–¥ –æ—Ç–ø—Ä–∞–≤–∫–æ–π
        total_context_length = len(system_content)
        logger.info(f"  üìä –§–ò–ù–ê–õ–¨–ù–ê–Ø –°–¢–ê–¢–ò–°–¢–ò–ö–ê:")
        logger.info(f"    - –û–±—â–∏–π —Ä–∞–∑–º–µ—Ä –∫–æ–Ω—Ç–µ–∫—Å—Ç–∞: {total_context_length} —Å–∏–º–≤–æ–ª–æ–≤")
        logger.info(f"    - –ò—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω—ã Excel –¥–∞–Ω–Ω—ã–µ: {needs_excel}")
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ä–∞–∑–º–µ—Ä –∫–æ–Ω—Ç–µ–∫—Å—Ç–∞
        if total_context_length > 8000:
            logger.warning(f"‚ö†Ô∏è –ö–æ–Ω—Ç–µ–∫—Å—Ç —Å–ª–∏—à–∫–æ–º –±–æ–ª—å—à–æ–π ({total_context_length} —Å–∏–º–≤–æ–ª–æ–≤), –æ–±—Ä–µ–∑–∞–µ–º...")
            # –û—Å—Ç–∞–≤–ª—è–µ–º —Ç–æ–ª—å–∫–æ –±–∞–∑–æ–≤—ã–π –∫–æ–Ω—Ç–µ–∫—Å—Ç –∏ –ø–µ—Ä–≤—ã–µ 4000 —Å–∏–º–≤–æ–ª–æ–≤ Excel –¥–∞–Ω–Ω—ã—Ö
            system_content = system_content[:4000] + "\n\n[–î–∞–Ω–Ω—ã–µ –æ–±—Ä–µ–∑–∞–Ω—ã –¥–ª—è —É—Å–∫–æ—Ä–µ–Ω–∏—è –æ—Ç–≤–µ—Ç–∞]"
            logger.info(f"  ‚úÇÔ∏è –ö–æ–Ω—Ç–µ–∫—Å—Ç –æ–±—Ä–µ–∑–∞–Ω –¥–æ {len(system_content)} —Å–∏–º–≤–æ–ª–æ–≤")
        
        logger.info(f"  üöÄ –û—Ç–ø—Ä–∞–≤–ª—è–µ–º –∑–∞–ø—Ä–æ—Å –≤ ChatGPT...")
        
        timeout = aiohttp.ClientTimeout(total=60)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.post(url, headers=headers, json=data) as resp:
                if resp.status != 200:
                    logger.error(f"‚ùå ChatGPT API error: {resp.status} - {await resp.text()}")
                    return "–ò–∑–≤–∏–Ω–∏—Ç–µ, –ø—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ –≤–∞—à–µ–≥–æ –∑–∞–ø—Ä–æ—Å–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â–µ —Ä–∞–∑."
                
                result = await resp.json()
                if "choices" not in result or not result["choices"]:
                    logger.error(f"‚ùå ChatGPT API unexpected response: {result}")
                    return "–ò–∑–≤–∏–Ω–∏—Ç–µ, –ø—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ –≤–∞—à–µ–≥–æ –∑–∞–ø—Ä–æ—Å–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â–µ —Ä–∞–∑."
                
                response_content = result["choices"][0]["message"]["content"].strip()
                
                logger.info(f"  ‚úÖ –û–¢–í–ï–¢ –û–¢ CHATGPT –ü–û–õ–£–ß–ï–ù:")
                logger.info(f"    - –î–ª–∏–Ω–∞ –æ—Ç–≤–µ—Ç–∞: {len(response_content)} —Å–∏–º–≤–æ–ª–æ–≤")
                logger.info(f"    - –ü–µ—Ä–≤—ã–µ 200 —Å–∏–º–≤–æ–ª–æ–≤: '{response_content[:200]}{'...' if len(response_content) > 200 else ''}'")
                
                return response_content
                
    except asyncio.TimeoutError:
        logger.error(f"‚è∞ ChatGPT API timeout –¥–ª—è –≤–æ–ø—Ä–æ—Å–∞: '{question}'")
        return "–ò–∑–≤–∏–Ω–∏—Ç–µ, –∑–∞–ø—Ä–æ—Å –∑–∞–Ω—è–ª —Å–ª–∏—à–∫–æ–º –º–Ω–æ–≥–æ –≤—Ä–µ–º–µ–Ω–∏. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â–µ —Ä–∞–∑."
    except aiohttp.ClientError as e:
        logger.error(f"üåê ChatGPT API client error –¥–ª—è –≤–æ–ø—Ä–æ—Å–∞ '{question}': {e}")
        return "–ò–∑–≤–∏–Ω–∏—Ç–µ, –ø—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞ —Å–µ—Ç–∏. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â–µ —Ä–∞–∑."
    except Exception as e:
        logger.error(f"üí• ChatGPT API unexpected error –¥–ª—è –≤–æ–ø—Ä–æ—Å–∞ '{question}': {e}\n{traceback.format_exc()}")
        return "–ò–∑–≤–∏–Ω–∏—Ç–µ, –ø—Ä–æ–∏–∑–æ—à–ª–∞ –Ω–µ–æ–∂–∏–¥–∞–Ω–Ω–∞—è –æ—à–∏–±–∫–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â–µ —Ä–∞–∑."

async def search_note_api(note):
    try:
        url = f"https://api.alexander-dev.ru/bahur/search/?text={note}"
        timeout = aiohttp.ClientTimeout(total=10)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.get(url) as resp:
                if resp.status != 200:
                    logger.error(f"Search API error: {resp.status} - {await resp.text()}")
                    return {"status": "error", "message": "–û—à–∏–±–∫–∞ API"}
                
                result = await resp.json()
                return result
                
    except asyncio.TimeoutError:
        logger.error("Search API timeout")
        return {"status": "error", "message": "–¢–∞–π–º–∞—É—Ç –∑–∞–ø—Ä–æ—Å–∞"}
    except aiohttp.ClientError as e:
        logger.error(f"Search API client error: {e}")
        return {"status": "error", "message": "–û—à–∏–±–∫–∞ —Å–µ—Ç–∏"}
    except Exception as e:
        logger.error(f"Search API unexpected error: {e}\n{traceback.format_exc()}")
        return {"status": "error", "message": "–ù–µ–æ–∂–∏–¥–∞–Ω–Ω–∞—è –æ—à–∏–±–∫–∞"}

def format_aroma_name(aroma_name):
    """–§–æ—Ä–º–∞—Ç–∏—Ä—É–µ—Ç –Ω–∞–∑–≤–∞–Ω–∏–µ –∞—Ä–æ–º–∞—Ç–∞ —Å –±–æ–ª—å—à–æ–π –±—É–∫–≤—ã –∏ –¥–µ—Ñ–∏—Å–∞–º–∏"""
    if not aroma_name:
        return ""
    
    # –†–∞–∑–±–∏–≤–∞–µ–º –Ω–∞ —Å–ª–æ–≤–∞
    words = str(aroma_name).strip().split()
    formatted_words = []
    
    for word in words:
        # –ï—Å–ª–∏ —Å–ª–æ–≤–æ —Å–æ–¥–µ—Ä–∂–∏—Ç –¥–µ—Ñ–∏—Å, –∫–∞–∂–¥—É—é —á–∞—Å—Ç—å —Å –±–æ–ª—å—à–æ–π –±—É–∫–≤—ã
        if '-' in word:
            parts = word.split('-')
            formatted_parts = [part.capitalize() for part in parts]
            formatted_words.append('-'.join(formatted_parts))
        else:
            # –û–±—ã—á–Ω–æ–µ —Å–ª–æ–≤–æ —Å –±–æ–ª—å—à–æ–π –±—É–∫–≤—ã
            formatted_words.append(word.capitalize())
    
    return ' '.join(formatted_words)

def get_country_emoji(country_name):
    """–í–æ–∑–≤—Ä–∞—â–∞–µ—Ç —ç–º–æ–¥–∂–∏ —Ñ–ª–∞–≥–∞ —Å—Ç—Ä–∞–Ω—ã –ø–æ –Ω–∞–∑–≤–∞–Ω–∏—é"""
    country_emojis = {
        '–∏—Ç–∞–ª–∏—è': 'üáÆüáπ', 'italy': 'üáÆüáπ',
        '—Ñ—Ä–∞–Ω—Ü–∏—è': 'üá´üá∑', 'france': 'üá´üá∑',
        '–≥–µ—Ä–º–∞–Ω–∏—è': 'üá©üá™', 'germany': 'üá©üá™',
        '–≤–µ–ª–∏–∫–æ–±—Ä–∏—Ç–∞–Ω–∏—è': 'üá¨üáß', 'uk': 'üá¨üáß', 'england': 'üá¨üáß',
        '–∏—Å–ø–∞–Ω–∏—è': 'üá™üá∏', 'spain': 'üá™üá∏',
        '–ø–æ—Ä—Ç—É–≥–∞–ª–∏—è': 'üáµüáπ', 'portugal': 'üáµüáπ',
        '–Ω–∏–¥–µ—Ä–ª–∞–Ω–¥—ã': 'üá≥üá±', 'netherlands': 'üá≥üá±', 'holland': 'üá≥üá±',
        '–±–µ–ª—å–≥–∏—è': 'üáßüá™', 'belgium': 'üáßüá™',
        '—à–≤–µ–π—Ü–∞—Ä–∏—è': 'üá®üá≠', 'switzerland': 'üá®üá≠',
        '–∞–≤—Å—Ç—Ä–∏—è': 'üá¶üáπ', 'austria': 'üá¶üáπ',
        '—Ç—É—Ä—Ü–∏—è': 'üáπüá∑', 'turkey': 'üáπüá∑',
        '—Ä–æ—Å—Å–∏—è': 'üá∑üá∫', 'russia': 'üá∑üá∫',
        '—Å—à–∞': 'üá∫üá∏', 'usa': 'üá∫üá∏', 'america': 'üá∫üá∏',
        '–∫–∞–Ω–∞–¥–∞': 'üá®üá¶', 'canada': 'üá®üá¶',
        '—è–ø–æ–Ω–∏—è': 'üáØüáµ', 'japan': 'üáØüáµ',
        '–∫–∏—Ç–∞–π': 'üá®üá≥', 'china': 'üá®üá≥',
        '–∫–æ—Ä–µ—è': 'üá∞üá∑', 'korea': 'üá∞üá∑',
        '–∏–Ω–¥–∏—è': 'üáÆüá≥', 'india': 'üáÆüá≥',
        '–±—Ä–∞–∑–∏–ª–∏—è': 'üáßüá∑', 'brazil': 'üáßüá∑',
        '–∞—Ä–≥–µ–Ω—Ç–∏–Ω–∞': 'üá¶üá∑', 'argentina': 'üá¶üá∑',
        '–º–µ–∫—Å–∏–∫–∞': 'üá≤üáΩ', 'mexico': 'üá≤üáΩ',
        '–∞–≤—Å—Ç—Ä–∞–ª–∏—è': 'üá¶üá∫', 'australia': 'üá¶üá∫',
        '–Ω–æ–≤–∞—è –∑–µ–ª–∞–Ω–¥–∏—è': 'üá≥üáø', 'new zealand': 'üá≥üáø',
        '—é–∂–Ω–∞—è –∞—Ñ—Ä–∏–∫–∞': 'üáøüá¶', 'south africa': 'üáøüá¶',
        '–µ–≥–∏–ø–µ—Ç': 'üá™üá¨', 'egypt': 'üá™üá¨',
        '–º–∞—Ä–æ–∫–∫–æ': 'üá≤üá¶', 'morocco': 'üá≤üá¶',
        '–¥—É–±–∞–π': 'üá¶üá™', 'uae': 'üá¶üá™', '—ç–º–∏—Ä–∞—Ç—ã': 'üá¶üá™',
        '—Å–∞—É–¥–æ–≤—Å–∫–∞—è –∞—Ä–∞–≤–∏—è': 'üá∏üá¶', 'saudi arabia': 'üá∏üá¶',
        '–∫–∞—Ç–∞—Ä': 'üá∂üá¶', 'qatar': 'üá∂üá¶',
        '–∫—É–≤–µ–π—Ç': 'üá∞üáº', 'kuwait': 'üá∞üáº',
        '–±–∞—Ö—Ä–µ–π–Ω': 'üáßüá≠', 'bahrain': 'üáßüá≠',
        '–æ–º–∞–Ω': 'üá¥üá≤', 'oman': 'üá¥üá≤',
        '–∏–æ—Ä–¥–∞–Ω–∏—è': 'üáØüá¥', 'jordan': 'üáØüá¥',
        '–ª–∏–≤–∞–Ω': 'üá±üáß', 'lebanon': 'üá±üáß',
        '—Å–∏—Ä–∏—è': 'üá∏üáæ', 'syria': 'üá∏üáæ',
        '–∏—Ä–∞–∫': 'üáÆüá∂', 'iraq': 'üáÆüá∂',
        '–∏—Ä–∞–Ω': 'üáÆüá∑', 'iran': 'üáÆüá∑',
        '–ø–∞–∫–∏—Å—Ç–∞–Ω': 'üáµüá∞', 'pakistan': 'üáµüá∞',
        '–∞—Ñ–≥–∞–Ω–∏—Å—Ç–∞–Ω': 'üá¶üá´', 'afghanistan': 'üá¶üá´',
        '—É–∑–±–µ–∫–∏—Å—Ç–∞–Ω': 'üá∫üáø', 'uzbekistan': 'üá∫üáø',
        '–∫–∞–∑–∞—Ö—Å—Ç–∞–Ω': 'üá∞üáø', 'kazakhstan': 'üá∞üáø',
        '–∫–∏—Ä–≥–∏–∑–∏—è': 'üá∞üá¨', 'kyrgyzstan': 'üá∞üá¨',
        '—Ç–∞–¥–∂–∏–∫–∏—Å—Ç–∞–Ω': 'üáπüáØ', 'tajikistan': 'üáπüáØ',
        '—Ç—É—Ä–∫–º–µ–Ω–∏—Å—Ç–∞–Ω': 'üáπüá≤', 'turkmenistan': 'üáπüá≤',
        '–∞–∑–µ—Ä–±–∞–π–¥–∂–∞–Ω': 'üá¶üáø', 'azerbaijan': 'üá¶üáø',
        '–≥—Ä—É–∑–∏—è': 'üá¨üá™', 'georgia': 'üá¨üá™',
        '–∞—Ä–º–µ–Ω–∏—è': 'üá¶üá≤', 'armenia': 'üá¶üá≤',
        '–º–æ–ª–¥–æ–≤–∞': 'üá≤üá©', 'moldova': 'üá≤üá©',
        '—É–∫—Ä–∞–∏–Ω–∞': 'üá∫üá¶', 'ukraine': 'üá∫üá¶',
        '–±–µ–ª–∞—Ä—É—Å—å': 'üáßüáæ', 'belarus': 'üáßüáæ',
        '–ª–∞—Ç–≤–∏—è': 'üá±üáª', 'latvia': 'üá±üáª',
        '–ª–∏—Ç–≤–∞': 'üá±üáπ', 'lithuania': 'üá±üáπ',
        '—ç—Å—Ç–æ–Ω–∏—è': 'üá™üá™', 'estonia': 'üá™üá™',
        '–ø–æ–ª—å—à–∞': 'üáµüá±', 'poland': 'üáµüá±',
        '—á–µ—Ö–∏—è': 'üá®üáø', 'czech republic': 'üá®üáø',
        '—Å–ª–æ–≤–∞–∫–∏—è': 'üá∏üá∞', 'slovakia': 'üá∏üá∞',
        '–≤–µ–Ω–≥—Ä–∏—è': 'üá≠üá∫', 'hungary': 'üá≠üá∫',
        '—Ä—É–º—ã–Ω–∏—è': 'üá∑üá¥', 'romania': 'üá∑üá¥',
        '–±–æ–ª–≥–∞—Ä–∏—è': 'üáßüá¨', 'bulgaria': 'üáßüá¨',
        '–≥—Ä–µ—Ü–∏—è': 'üá¨üá∑', 'greece': 'üá¨üá∑',
        '—Ö–æ—Ä–≤–∞—Ç–∏—è': 'üá≠üá∑', 'croatia': 'üá≠üá∑',
        '—Å–µ—Ä–±–∏—è': 'üá∑üá∏', 'serbia': 'üá∑üá∏',
        '—á–µ—Ä–Ω–æ–≥–æ—Ä–∏—è': 'üá≤üá™', 'montenegro': 'üá≤üá™',
        '–∞–ª–±–∞–Ω–∏—è': 'üá¶üá±', 'albania': 'üá¶üá±',
        '–º–∞–∫–µ–¥–æ–Ω–∏—è': 'üá≤üá∞', 'macedonia': 'üá≤üá∞',
        '—Å–ª–æ–≤–µ–Ω–∏—è': 'üá∏üáÆ', 'slovenia': 'üá∏üáÆ',
        '–±–æ—Å–Ω–∏—è': 'üáßüá¶', 'bosnia': 'üáßüá¶',
        '–∫–∏–ø—Ä': 'üá®üáæ', 'cyprus': 'üá®üáæ',
        '–º–∞–ª—å—Ç–∞': 'üá≤üáπ', 'malta': 'üá≤üáπ',
        '–∏—Å–ª–∞–Ω–¥–∏—è': 'üáÆüá∏', 'iceland': 'üáÆüá∏',
        '–Ω–æ—Ä–≤–µ–≥–∏—è': 'üá≥üá¥', 'norway': 'üá≥üá¥',
        '—à–≤–µ—Ü–∏—è': 'üá∏üá™', 'sweden': 'üá∏üá™',
        '—Ñ–∏–Ω–ª—è–Ω–¥–∏—è': 'üá´üáÆ', 'finland': 'üá´üáÆ',
        '–¥–∞–Ω–∏—è': 'üá©üá∞', 'denmark': 'üá©üá∞',
        '–∏—Ä–ª–∞–Ω–¥–∏—è': 'üáÆüá™', 'ireland': 'üáÆüá™',
        '–ª—é–∫—Å–µ–º–±—É—Ä–≥': 'üá±üá∫', 'luxembourg': 'üá±üá∫',
        '–º–æ–Ωaco': 'üá≤üá®', 'monaco': 'üá≤üá®',
        '–∞–Ω–¥–æ—Ä—Ä–∞': 'üá¶üá©', 'andorra': 'üá¶üá©',
        '—Å–∞–Ω-–º–∞—Ä–∏–Ω–æ': 'üá∏üá≤', 'san marino': 'üá∏üá≤',
        '–≤–∞—Ç–∏–∫–∞–Ω': 'üáªüá¶', 'vatican': 'üáªüá¶',
        '–ª–∏—Ö—Ç–µ–Ω—à—Ç–µ–π–Ω': 'üá±üáÆ', 'liechtenstein': 'üá±üáÆ'
    }
    
    if not country_name:
        return "üåç"
    
    country_lower = str(country_name).lower().strip()
    return country_emojis.get(country_lower, "üåç")

async def get_notes_from_api(aroma_name):
    """–ü–æ–ª—É—á–∞–µ—Ç –Ω–æ—Ç—ã –∞—Ä–æ–º–∞—Ç–∞ —á–µ—Ä–µ–∑ API"""
    try:
        result = await search_note_api(aroma_name)
        if result.get("status") == "success" and "data" in result:
            data = result["data"]
            if isinstance(data, list) and len(data) > 0:
                aroma_data = data[0]  # –ë–µ—Ä–µ–º –ø–µ—Ä–≤—ã–π —Ä–µ–∑—É–ª—å—Ç–∞—Ç
                return {
                    "top_notes": aroma_data.get("top_notes", ""),
                    "middle_notes": aroma_data.get("middle_notes", ""),
                    "base_notes": aroma_data.get("base_notes", ""),
                    "country": aroma_data.get("country", ""),
                    "link": aroma_data.get("link", "")
                }
        return None
    except Exception as e:
        logger.error(f"Error getting notes from API: {e}")
        return None

# --- Telegram sendMessage ---
async def telegram_send_message(chat_id, text, reply_markup=None, parse_mode="HTML"):
    try:
        url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
        payload = {
            "chat_id": chat_id,
            "text": text,
            "parse_mode": parse_mode
        }
        if reply_markup:
            payload["reply_markup"] = reply_markup
        
        timeout = httpx.Timeout(30.0)
        async with httpx.AsyncClient(timeout=timeout) as client:
            resp = await client.post(url, json=payload)
            if resp.status_code != 200:
                logger.error(f"Telegram API error: {resp.status_code} - {resp.text}")
                return False
            return True
            
    except httpx.TimeoutException:
        logger.error("Telegram API timeout")
        return False
    except httpx.RequestError as e:
        logger.error(f"Telegram API request error: {e}")
        return False
    except Exception as e:
        logger.error(f"Telegram API unexpected error: {e}\n{traceback.format_exc()}")
        return False

# --- Telegram editMessage ---
async def telegram_edit_message(chat_id, message_id, text, reply_markup=None, parse_mode="HTML"):
    try:
        url = f"https://api.telegram.org/bot{TOKEN}/editMessageText"
        payload = {
            "chat_id": chat_id,
            "message_id": message_id,
            "text": text,
            "parse_mode": parse_mode
        }
        if reply_markup:
            payload["reply_markup"] = reply_markup
        
        timeout = httpx.Timeout(30.0)
        async with httpx.AsyncClient(timeout=timeout) as client:
            resp = await client.post(url, json=payload)
            if resp.status_code != 200:
                logger.error(f"Telegram editMessage API error: {resp.status_code} - {resp.text}")
                return False
            return True
            
    except httpx.TimeoutException:
        logger.error("Telegram editMessage API timeout")
        return False
    except httpx.RequestError as e:
        logger.error(f"Telegram editMessage API request error: {e}")
        return False
    except Exception as e:
        logger.error(f"Telegram editMessage API unexpected error: {e}\n{traceback.format_exc()}")
        return False

# --- Telegram answerCallbackQuery ---
async def telegram_answer_callback_query(callback_query_id, text=None, show_alert=False):
    try:
        url = f"https://api.telegram.org/bot{TOKEN}/answerCallbackQuery"
        payload = {
            "callback_query_id": callback_query_id
        }
        if text:
            payload["text"] = text
        if show_alert:
            payload["show_alert"] = show_alert
        
        timeout = httpx.Timeout(10.0)
        async with httpx.AsyncClient(timeout=timeout) as client:
            resp = await client.post(url, json=payload)
            if resp.status_code != 200:
                logger.error(f"Telegram answerCallbackQuery API error: {resp.status_code} - {resp.text}")
                return False
            return True
            
    except httpx.TimeoutException:
        logger.error("Telegram answerCallbackQuery API timeout")
        return False
    except httpx.RequestError as e:
        logger.error(f"Telegram answerCallbackQuery API request error: {e}")
        return False
    except Exception as e:
        logger.error(f"Telegram answerCallbackQuery API unexpected error: {e}\n{traceback.format_exc()}")
        return False

# --- –ü–æ–∏—Å–∫ –ø–æ ID –∞—Ä–æ–º–∞—Ç–∞ ---
async def search_by_id_api(aroma_id):
    try:
        url = f"https://api.alexander-dev.ru/bahur/search/?id={aroma_id}"
        timeout = aiohttp.ClientTimeout(total=10)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.get(url) as resp:
                if resp.status != 200:
                    logger.error(f"Search by ID API error: {resp.status} - {await resp.text()}")
                    return {"status": "error", "message": "–û—à–∏–±–∫–∞ API"}
                
                result = await resp.json()
                return result
                
    except asyncio.TimeoutError:
        logger.error("Search by ID API timeout")
        return {"status": "error", "message": "–¢–∞–π–º–∞—É—Ç –∑–∞–ø—Ä–æ—Å–∞"}
    except aiohttp.ClientError as e:
        logger.error(f"Search by ID API client error: {e}")
        return {"status": "error", "message": "–û—à–∏–±–∫–∞ —Å–µ—Ç–∏"}
    except Exception as e:
        logger.error(f"Search by ID API unexpected error: {e}\n{traceback.format_exc()}")
        return {"status": "error", "message": "–ù–µ–æ–∂–∏–¥–∞–Ω–Ω–∞—è –æ—à–∏–±–∫–∞"}

# --- –û–±—Ä–∞–±–æ—Ç–∫–∞ –≥–æ–ª–æ—Å–æ–≤—ã—Ö —Å–æ–æ–±—â–µ–Ω–∏–π ---
async def recognize_voice_content(file_content, chat_id=None):
    """–†–∞—Å–ø–æ–∑–Ω–∞—ë—Ç —Ä–µ—á—å –∏–∑ –±–∞–π—Ç–æ–≤–æ–≥–æ —Å–æ–¥–µ—Ä–∂–∏–º–æ–≥–æ ogg-—Ñ–∞–π–ª–∞. –í–æ–∑–≤—Ä–∞—â–∞–µ—Ç —Ç–µ–∫—Å—Ç –∏–ª–∏ —Å—Ç—Ä–æ–∫—É-–æ—à–∏–±–∫—É."""
    try:
        import speech_recognition as sr
        from pydub import AudioSegment
        import tempfile
        import math
        
        recognizer = sr.Recognizer()
        
        with tempfile.NamedTemporaryFile(suffix='.ogg') as temp_ogg, tempfile.NamedTemporaryFile(suffix='.wav') as temp_wav:
            temp_ogg.write(file_content)
            temp_ogg.flush()
            
            try:
                audio = AudioSegment.from_file(temp_ogg.name)
                
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º –¥–ª–∏—Ç–µ–ª—å–Ω–æ—Å—Ç—å (–≤ –º–∏–ª–ª–∏—Å–µ–∫—É–Ω–¥–∞—Ö)
                duration_seconds = len(audio) / 1000.0
                logger.info(f"Voice message duration: {duration_seconds:.1f} seconds")
                
                # –ï—Å–ª–∏ —Ñ–∞–π–ª —Å–ª–∏—à–∫–æ–º –¥–ª–∏–Ω–Ω—ã–π (–±–æ–ª—å—à–µ 2 –º–∏–Ω—É—Ç), —Ä–∞–∑–±–∏–≤–∞–µ–º –Ω–∞ —á–∞—Å—Ç–∏
                if duration_seconds > 120:
                    return await recognize_long_audio(audio, chat_id)
                
                # –î–ª—è –∫–æ—Ä–æ—Ç–∫–∏—Ö —Ñ–∞–π–ª–æ–≤ - –æ–±—ã—á–Ω–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞
                audio.export(temp_wav.name, format='wav')
                temp_wav.flush()
                
            except Exception as audio_error:
                logger.error(f"Audio conversion error: {audio_error}")
                return "–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ –∞—É–¥–∏–æ —Ñ–∞–π–ª–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â–µ —Ä–∞–∑ –∏–ª–∏ –Ω–∞–ø–∏—à–∏—Ç–µ —Ç–µ–∫—Å—Ç."
            
            try:
                with sr.AudioFile(temp_wav.name) as source:
                    # –£—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º –Ω–∞—Å—Ç—Ä–æ–π–∫–∏ –¥–ª—è –ª—É—á—à–µ–≥–æ —Ä–∞—Å–ø–æ–∑–Ω–∞–≤–∞–Ω–∏—è
                    recognizer.energy_threshold = 300
                    recognizer.pause_threshold = 0.8
                    recognizer.phrase_threshold = 0.3
                    recognizer.non_speaking_duration = 0.5
                    
                    audio_data = recognizer.record(source)
                    text_content = recognizer.recognize_google(audio_data, language='ru-RU')
                    logger.info(f"Voice recognized: '{text_content}'")
                    return text_content
                    
            except sr.UnknownValueError:
                logger.error("Speech recognition could not understand audio")
                return "–ù–µ —É–¥–∞–ª–æ—Å—å —Ä–∞–∑–æ–±—Ä–∞—Ç—å —Ä–µ—á—å. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –≥–æ–≤–æ—Ä–∏—Ç—å —á–µ—Ç—á–µ –∏–ª–∏ –Ω–∞–ø–∏—à–∏—Ç–µ —Ç–µ–∫—Å—Ç."
            except sr.RequestError as e:
                logger.error(f"Speech recognition service error: {e}")
                return "–û—à–∏–±–∫–∞ —Å–µ—Ä–≤–∏—Å–∞ —Ä–∞—Å–ø–æ–∑–Ω–∞–≤–∞–Ω–∏—è —Ä–µ—á–∏. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â–µ —Ä–∞–∑ –∏–ª–∏ –Ω–∞–ø–∏—à–∏—Ç–µ —Ç–µ–∫—Å—Ç."
                
    except Exception as e:
        logger.error(f"Speech recognition error: {e}\n{traceback.format_exc()}")
        return "–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ –≥–æ–ª–æ—Å–æ–≤–æ–≥–æ —Å–æ–æ–±—â–µ–Ω–∏—è."

async def recognize_long_audio(audio_segment, chat_id=None):
    """–†–∞—Å–ø–æ–∑–Ω–∞–µ—Ç –¥–ª–∏–Ω–Ω—ã–µ –∞—É–¥–∏–æ —Ñ–∞–π–ª—ã, —Ä–∞–∑–±–∏–≤–∞—è –∏—Ö –Ω–∞ —á–∞—Å—Ç–∏"""
    try:
        import speech_recognition as sr
        import tempfile
        
        recognizer = sr.Recognizer()
        recognizer.energy_threshold = 300
        recognizer.pause_threshold = 0.8
        
        # –†–∞–∑–±–∏–≤–∞–µ–º –Ω–∞ —á–∞—Å—Ç–∏ –ø–æ 55 —Å–µ–∫—É–Ω–¥ —Å –ø–µ—Ä–µ–∫—Ä—ã—Ç–∏–µ–º 5 —Å–µ–∫—É–Ω–¥
        chunk_length = 55 * 1000  # 55 —Å–µ–∫—É–Ω–¥ –≤ –º–∏–ª–ª–∏—Å–µ–∫—É–Ω–¥–∞—Ö
        overlap = 5 * 1000        # 5 —Å–µ–∫—É–Ω–¥ –ø–µ—Ä–µ–∫—Ä—ã—Ç–∏—è
        
        total_length = len(audio_segment)
        total_duration = total_length / 1000.0
        recognized_texts = []
        
        # –ü–æ–¥—Å—á–∏—Ç—ã–≤–∞–µ–º –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ —á–∞—Å—Ç–µ–π
        chunks_count = len(list(range(0, total_length - overlap, chunk_length - overlap)))
        current_chunk = 0
        
        for start in range(0, total_length - overlap, chunk_length - overlap):
            end = min(start + chunk_length, total_length)
            chunk = audio_segment[start:end]
            current_chunk += 1
            
            logger.info(f"Processing audio chunk {current_chunk}/{chunks_count}: {start/1000:.1f}s - {end/1000:.1f}s")
            
            # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–ª–µ–Ω–∏–µ –æ –ø—Ä–æ–≥—Ä–µ—Å—Å–µ
            if chat_id and current_chunk % 3 == 0:  # –ö–∞–∂–¥—É—é —Ç—Ä–µ—Ç—å—é —á–∞—Å—Ç—å
                progress_percent = int((current_chunk / chunks_count) * 100)
                await send_progress_message(chat_id, 
                    f"üîÑ –û–±—Ä–∞–±–∞—Ç—ã–≤–∞—é —á–∞—Å—Ç—å {current_chunk}/{chunks_count} ({progress_percent}%)")
            
            with tempfile.NamedTemporaryFile(suffix='.wav') as temp_chunk:
                chunk.export(temp_chunk.name, format='wav')
                temp_chunk.flush()
                
                try:
                    with sr.AudioFile(temp_chunk.name) as source:
                        audio_data = recognizer.record(source)
                        text = recognizer.recognize_google(audio_data, language='ru-RU')
                        if text.strip():
                            recognized_texts.append(text.strip())
                            logger.info(f"Chunk {current_chunk} recognized: '{text[:50]}...'")
                        
                except sr.UnknownValueError:
                    logger.warning(f"Could not understand audio chunk {current_chunk} ({start/1000:.1f}s - {end/1000:.1f}s)")
                    continue
                except sr.RequestError as e:
                    logger.error(f"Speech recognition error for chunk {current_chunk}: {e}")
                    continue
                
                # –ü–∞—É–∑–∞ –º–µ–∂–¥—É –∑–∞–ø—Ä–æ—Å–∞–º–∏ –∫ API
                await asyncio.sleep(0.5)
        
        if not recognized_texts:
            return "–ù–µ —É–¥–∞–ª–æ—Å—å —Ä–∞—Å–ø–æ–∑–Ω–∞—Ç—å —Ä–µ—á—å –≤ –¥–ª–∏–Ω–Ω–æ–º –≥–æ–ª–æ—Å–æ–≤–æ–º —Å–æ–æ–±—â–µ–Ω–∏–∏. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –∑–∞–ø–∏—Å–∞—Ç—å –±–æ–ª–µ–µ –∫–æ—Ä–æ—Ç–∫–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ –∏–ª–∏ –Ω–∞–ø–∏—à–∏—Ç–µ —Ç–µ–∫—Å—Ç."
        
        # –£–≤–µ–¥–æ–º–ª—è–µ–º –æ –∑–∞–≤–µ—Ä—à–µ–Ω–∏–∏ –æ–±—Ä–∞–±–æ—Ç–∫–∏
        if chat_id:
            await send_progress_message(chat_id, "‚úÖ –†–∞—Å–ø–æ–∑–Ω–∞–≤–∞–Ω–∏–µ –∑–∞–≤–µ—Ä—à–µ–Ω–æ, —Ñ–æ—Ä–º–∏—Ä—É—é –æ—Ç–≤–µ—Ç...")
        
        # –û–±—ä–µ–¥–∏–Ω—è–µ–º –≤—Å–µ —Ä–∞—Å–ø–æ–∑–Ω–∞–Ω–Ω—ã–µ —á–∞—Å—Ç–∏
        full_text = " ".join(recognized_texts)
        logger.info(f"Full recognized text: '{full_text[:100]}...'")
        
        return full_text
        
    except Exception as e:
        logger.error(f"Long audio recognition error: {e}\n{traceback.format_exc()}")
        return "–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ –¥–ª–∏–Ω–Ω–æ–≥–æ –≥–æ–ª–æ—Å–æ–≤–æ–≥–æ —Å–æ–æ–±—â–µ–Ω–∏—è."

async def process_voice_message(voice, chat_id):
    try:
        # –ü–æ–ª—É—á–∞–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ —Ñ–∞–π–ª–µ
        file_id = voice["file_id"]
        file_unique_id = voice["file_unique_id"]
        duration = voice.get("duration", 0)
        
        # –ü–æ–ª—É—á–∞–µ–º —Ñ–∞–π–ª
        file_url = f"https://api.telegram.org/bot{TOKEN}/getFile?file_id={file_id}"
        async with httpx.AsyncClient() as client:
            resp = await client.get(file_url)
            if resp.status_code != 200:
                logger.error(f"Failed to get file info: {resp.status_code}")
                return None
            
            file_info = resp.json()
            if not file_info.get("ok"):
                logger.error(f"File info error: {file_info}")
                return None
            
            file_path = file_info["result"]["file_path"]
            file_url = f"https://api.telegram.org/file/bot{TOKEN}/{file_path}"
            
            # –°–∫–∞—á–∏–≤–∞–µ–º —Ñ–∞–π–ª
            async with client.stream("GET", file_url) as response:
                if response.status_code != 200:
                    logger.error(f"Failed to download file: {response.status_code}")
                    return None
                
                # –ß–∏—Ç–∞–µ–º —Å–æ–¥–µ—Ä–∂–∏–º–æ–µ —Ñ–∞–π–ª–∞
                file_content = await response.aread()
                
                # –†–∞—Å–ø–æ–∑–Ω–∞–µ–º —Ä–µ—á—å —Å –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–µ–º tempfile
                text_content = await recognize_voice_content(file_content, chat_id)
                # –ï—Å–ª–∏ —Ä–µ–∑—É–ª—å—Ç–∞—Ç –Ω–µ –æ—à–∏–±–∫–∞, –æ—Ç–ø—Ä–∞–≤–ª—è–µ–º –≤ ChatGPT
                if text_content and not any(err in text_content for err in ["–û—à–∏–±–∫–∞", "–ù–µ —É–¥–∞–ª–æ—Å—å", "–Ω–µ–¥–æ—Å—Ç—É–ø–Ω–æ"]):
                    ai_answer = await ask_chatgpt(text_content)
                    return ai_answer
                else:
                    return text_content
                
    except Exception as e:
        logger.error(f"Voice processing error: {e}\n{traceback.format_exc()}")
        return "–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ –≥–æ–ª–æ—Å–æ–≤–æ–≥–æ —Å–æ–æ–±—â–µ–Ω–∏—è."

# --- –ê–ª—å—Ç–µ—Ä–Ω–∞—Ç–∏–≤–Ω–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ –≥–æ–ª–æ—Å–æ–≤—ã—Ö —Å–æ–æ–±—â–µ–Ω–∏–π (–±–µ–∑ aifc) ---
async def process_voice_message_alternative(voice, chat_id):
    """–ê–ª—å—Ç–µ—Ä–Ω–∞—Ç–∏–≤–Ω–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ –≥–æ–ª–æ—Å–æ–≤—ã—Ö —Å–æ–æ–±—â–µ–Ω–∏–π –±–µ–∑ aifc"""
    try:
        # –ü–æ–ª—É—á–∞–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ —Ñ–∞–π–ª–µ
        file_id = voice["file_id"]
        file_unique_id = voice["file_unique_id"]
        duration = voice.get("duration", 0)
        
        # –ï—Å–ª–∏ –≥–æ–ª–æ—Å–æ–≤–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ —Å–ª–∏—à–∫–æ–º –∫–æ—Ä–æ—Ç–∫–æ–µ
        if duration < 1:
            return "–ì–æ–ª–æ—Å–æ–≤–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ —Å–ª–∏—à–∫–æ–º –∫–æ—Ä–æ—Ç–∫–æ–µ. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –∑–∞–ø–∏—Å–∞—Ç—å –±–æ–ª–µ–µ –¥–ª–∏–Ω–Ω–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ."
        
        # –ü–æ–ª—É—á–∞–µ–º —Ñ–∞–π–ª
        file_url = f"https://api.telegram.org/bot{TOKEN}/getFile?file_id={file_id}"
        async with httpx.AsyncClient() as client:
            resp = await client.get(file_url)
            if resp.status_code != 200:
                logger.error(f"Failed to get file info: {resp.status_code}")
                return None
            
            file_info = resp.json()
            if not file_info.get("ok"):
                logger.error(f"File info error: {file_info}")
                return None
            
            file_path = file_info["result"]["file_path"]
            file_url = f"https://api.telegram.org/file/bot{TOKEN}/{file_path}"
            
            # –°–∫–∞—á–∏–≤–∞–µ–º —Ñ–∞–π–ª
            async with client.stream("GET", file_url) as response:
                if response.status_code != 200:
                    logger.error(f"Failed to download file: {response.status_code}")
                    return None
                
                # –ß–∏—Ç–∞–µ–º —Å–æ–¥–µ—Ä–∂–∏–º–æ–µ —Ñ–∞–π–ª–∞
                file_content = await response.aread()
                
                # –ü—ã—Ç–∞–µ–º—Å—è —Ä–∞—Å–ø–æ–∑–Ω–∞—Ç—å —Ä–µ—á—å –±–µ–∑ aifc
                text_content = await recognize_voice_content(file_content, chat_id)
                if text_content and not any(err in text_content for err in ["–û—à–∏–±–∫–∞", "–ù–µ —É–¥–∞–ª–æ—Å—å", "–Ω–µ–¥–æ—Å—Ç—É–ø–Ω–æ"]):
                    ai_answer = await ask_chatgpt(text_content)
                    return ai_answer
                else:
                    return text_content
                
    except Exception as e:
        logger.error(f"Alternative voice processing error: {e}\n{traceback.format_exc()}")
        return "–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ –≥–æ–ª–æ—Å–æ–≤–æ–≥–æ —Å–æ–æ–±—â–µ–Ω–∏—è."

# --- –£–ø—Ä–æ—â–µ–Ω–Ω–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ –≥–æ–ª–æ—Å–æ–≤—ã—Ö —Å–æ–æ–±—â–µ–Ω–∏–π (–±–µ–∑ —Ä–∞—Å–ø–æ–∑–Ω–∞–≤–∞–Ω–∏—è) ---
async def process_voice_message_simple(voice, chat_id):
    """–£–ø—Ä–æ—â–µ–Ω–Ω–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ –≥–æ–ª–æ—Å–æ–≤—ã—Ö —Å–æ–æ–±—â–µ–Ω–∏–π –±–µ–∑ —Å–ª–æ–∂–Ω—ã—Ö –∑–∞–≤–∏—Å–∏–º–æ—Å—Ç–µ–π"""
    try:
        # –ü–æ–ª—É—á–∞–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ —Ñ–∞–π–ª–µ
        file_id = voice["file_id"]
        duration = voice.get("duration", 0)
        
        # –ï—Å–ª–∏ –≥–æ–ª–æ—Å–æ–≤–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ —Å–ª–∏—à–∫–æ–º –∫–æ—Ä–æ—Ç–∫–æ–µ
        if duration < 1:
            return "–ì–æ–ª–æ—Å–æ–≤–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ —Å–ª–∏—à–∫–æ–º –∫–æ—Ä–æ—Ç–∫–æ–µ. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –∑–∞–ø–∏—Å–∞—Ç—å –±–æ–ª–µ–µ –¥–ª–∏–Ω–Ω–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ."
        
        # –ü–æ–ª—É—á–∞–µ–º —Ñ–∞–π–ª
        file_url = f"https://api.telegram.org/bot{TOKEN}/getFile?file_id={file_id}"
        async with httpx.AsyncClient() as client:
            resp = await client.get(file_url)
            if resp.status_code != 200:
                logger.error(f"Failed to get file info: {resp.status_code}")
                return None
            
            file_info = resp.json()
            if not file_info.get("ok"):
                logger.error(f"File info error: {file_info}")
                return None
            
            # –ü—Ä–æ—Å—Ç–æ –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –≥–æ–ª–æ—Å–æ–≤–æ–º —Å–æ–æ–±—â–µ–Ω–∏–∏
            return f"–ü–æ–ª—É—á–µ–Ω–æ –≥–æ–ª–æ—Å–æ–≤–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ –¥–ª–∏—Ç–µ–ª—å–Ω–æ—Å—Ç—å—é {duration} —Å–µ–∫—É–Ω–¥. –î–ª—è —Ä–∞—Å–ø–æ–∑–Ω–∞–≤–∞–Ω–∏—è —Ä–µ—á–∏ –Ω–∞–ø–∏—à–∏—Ç–µ –≤–∞—à –≤–æ–ø—Ä–æ—Å —Ç–µ–∫—Å—Ç–æ–º."
                
    except Exception as e:
        logger.error(f"Simple voice processing error: {e}\n{traceback.format_exc()}")
        return "–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ –≥–æ–ª–æ—Å–æ–≤–æ–≥–æ —Å–æ–æ–±—â–µ–Ω–∏—è."

# --- –§—É–Ω–∫—Ü–∏—è "–ø–µ—á–∞—Ç–∞–µ—Ç" ---
async def send_typing_action(chat_id):
    try:
        url = f"https://api.telegram.org/bot{TOKEN}/sendChatAction"
        payload = {
            "chat_id": chat_id,
            "action": "typing"
        }
        timeout = httpx.Timeout(5.0)
        async with httpx.AsyncClient(timeout=timeout) as client:
            resp = await client.post(url, json=payload)
            if resp.status_code != 200:
                logger.error(f"Failed to send typing action: {resp.status_code} - {resp.text}")
    except Exception as e:
        logger.error(f"Failed to send typing action: {e}")

async def send_progress_message(chat_id, text):
    """–û—Ç–ø—Ä–∞–≤–ª—è–µ—Ç —Å–æ–æ–±—â–µ–Ω–∏–µ –æ –ø—Ä–æ–≥—Ä–µ—Å—Å–µ –æ–±—Ä–∞–±–æ—Ç–∫–∏"""
    try:
        success = await telegram_send_message(chat_id, text)
        if success:
            logger.info(f"[TG] Sent progress message to {chat_id}")
        return success
    except Exception as e:
        logger.error(f"Failed to send progress message: {e}")
        return False

# --- –£–º–Ω–æ–µ —Ä–∞—Å–ø–æ–∑–Ω–∞–≤–∞–Ω–∏–µ –Ω–æ—Ç ---
def is_likely_note(text):
    """–û–ø—Ä–µ–¥–µ–ª—è–µ—Ç, –ø–æ—Ö–æ–∂ –ª–∏ —Ç–µ–∫—Å—Ç –Ω–∞ –Ω–∞–∑–≤–∞–Ω–∏–µ –Ω–æ—Ç—ã"""
    if not text:
        return False
    
    # –°–ø–∏—Å–æ–∫ –ø–æ–ø—É–ª—è—Ä–Ω—ã—Ö –Ω–æ—Ç
    common_notes = [
        '–≤–∞–Ω–∏–ª—å', '–ª–∞–≤–∞–Ω–¥–∞', '—Ä–æ–∑–∞', '–∂–∞—Å–º–∏–Ω', '—Å–∞–Ω–¥–∞–ª', '–º—É—Å–∫—É—Å', '–∞–º–±—Ä–∞', '–ø–∞—á—É–ª–∏',
        '–±–µ—Ä–≥–∞–º–æ—Ç', '–ª–∏–º–æ–Ω', '–∞–ø–µ–ª—å—Å–∏–Ω', '–º–∞–Ω–¥–∞—Ä–∏–Ω', '–≥—Ä–µ–π–ø—Ñ—Ä—É—Ç', '–ª–∞–π–º',
        '–∫–ª—É–±–Ω–∏–∫–∞', '–º–∞–ª–∏–Ω–∞', '—á–µ—Ä–Ω–∏–∫–∞', '–≤–∏—à–Ω—è', '–ø–µ—Ä—Å–∏–∫', '–∞–±—Ä–∏–∫–æ—Å', '—è–±–ª–æ–∫–æ',
        '–≥—Ä—É—à–∞', '–∞–Ω–∞–Ω–∞—Å', '–º–∞–Ω–≥–æ', '–±–∞–Ω–∞–Ω', '–∫–æ–∫–æ—Å', '–∫–∞—Ä–∞–º–µ–ª—å', '—à–æ–∫–æ–ª–∞–¥',
        '–∫–æ—Ñ–µ', '—á–∞–π', '–º—è—Ç–∞', '–±–∞–∑–∏–ª–∏–∫', '—Ä–æ–∑–º–∞—Ä–∏–Ω', '—Ç–∏–º—å—è–Ω', '–æ—Ä–µ–≥–∞–Ω–æ',
        '–∫–æ—Ä–∏—Ü–∞', '–∫–∞—Ä–¥–∞–º–æ–Ω', '–∏–º–±–∏—Ä—å', '–∫—É—Ä–∫—É–º–∞', '–ø–µ—Ä–µ—Ü', '–≥–≤–æ–∑–¥–∏–∫–∞',
        '–∫–µ–¥—Ä', '—Å–æ—Å–Ω–∞', '–µ–ª—å', '–¥—É–±', '–±–µ—Ä–µ–∑–∞', '–∏–ª–∞–Ω–≥-–∏–ª–∞–Ω–≥', '–Ω–µ—Ä–æ–ª–∏',
        '–∏—Ä–∏—Å', '—Ñ–∏–∞–ª–∫–∞', '–ª–∞–Ω–¥—ã—à', '—Å–∏—Ä–µ–Ω—å', '–∂–∞—Å–º–∏–Ω', '–≥–∞—Ä–¥–µ–Ω–∏—è',
        '–º–æ—Ä—Å–∫–∞—è —Å–æ–ª—å', '–º–æ—Ä—Å–∫–æ–π –±—Ä–∏–∑', '–¥–æ–∂–¥—å', '—Å–Ω–µ–≥', '–∑–µ–º–ª—è', '–º–æ—Ö',
        '–¥—ã–º', '–∫–æ–∂–∞', '—Ç–∞–±–∞–∫', '–≤–∏—Å–∫–∏', '–∫–æ–Ω—å—è–∫', '—Ä–æ–º', '–≤–∏–Ω–æ',
        '–º–µ–¥', '—Å–ª–∏–≤–∫–∏', '–º–æ–ª–æ–∫–æ', '–π–æ–≥—É—Ä—Ç', '—Å—ã—Ä', '–º–∞—Å–ª–æ'
    ]
    
    text_lower = text.lower().strip()
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ
    if text_lower in common_notes:
        return True
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º —á–∞—Å—Ç–∏—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ
    for note in common_notes:
        if note in text_lower or text_lower in note:
            return True
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –ø–æ –¥–ª–∏–Ω–µ –∏ —Ö–∞—Ä–∞–∫—Ç–µ—Ä—É (–∫–æ—Ä–æ—Ç–∫–∏–µ —Å–ª–æ–≤–∞ —á–∞—Å—Ç–æ –±—ã–≤–∞—é—Ç –Ω–æ—Ç–∞–º–∏)
    if len(text_lower) <= 15 and not any(char.isdigit() for char in text_lower):
        # –ï—Å–ª–∏ —Ç–µ–∫—Å—Ç –∫–æ—Ä–æ—Ç–∫–∏–π –∏ –Ω–µ —Å–æ–¥–µ—Ä–∂–∏—Ç —Ü–∏—Ñ—Ä, –≤–æ–∑–º–æ–∂–Ω–æ —ç—Ç–æ –Ω–æ—Ç–∞
        return True
    
    return False

# --- –û–±—Ä–∞–±–æ—Ç–∫–∞ —Å—Å—ã–ª–æ–∫ –≤ —Ç–µ–∫—Å—Ç–µ ---
import re

def convert_to_nominative_case(text):
    """–ü—Ä–∏–≤–æ–¥–∏—Ç —Ç–µ–∫—Å—Ç –∫ –∏–º–µ–Ω–∏—Ç–µ–ª—å–Ω–æ–º—É –ø–∞–¥–µ–∂—É (–±–∞–∑–æ–≤–∞—è —Ä–µ–∞–ª–∏–∑–∞—Ü–∏—è –¥–ª—è –Ω–∞–∏–±–æ–ª–µ–µ —á–∞—Å—Ç—ã—Ö —Å–ª—É—á–∞–µ–≤)"""
    text = text.strip().lower()
    
    # –°–ª–æ–≤–∞—Ä—å –Ω–∞–∏–±–æ–ª–µ–µ —á–∞—Å—Ç—ã—Ö –ø—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω–∏–π
    nominative_dict = {
        # –í–∏–Ω–∏—Ç–µ–ª—å–Ω—ã–π –ø–∞–¥–µ–∂ -> –ò–º–µ–Ω–∏—Ç–µ–ª—å–Ω—ã–π
        '–ø—Ä–∞–π—Å': '–ø—Ä–∞–π—Å',
        '–∫–∞—Ç–∞–ª–æ–≥': '–∫–∞—Ç–∞–ª–æ–≥', 
        '–º–∞–≥–∞–∑–∏–Ω': '–º–∞–≥–∞–∑–∏–Ω',
        '—Å–∞–π—Ç': '—Å–∞–π—Ç',
        '—Å—Ç—Ä–∞–Ω–∏—Ü—É': '—Å—Ç—Ä–∞–Ω–∏—Ü–∞',
        '—Å—Ç—Ä–∞–Ω–∏—á–∫—É': '—Å—Ç—Ä–∞–Ω–∏—á–∫–∞',
        '—Ç–æ–≤–∞—Ä': '—Ç–æ–≤–∞—Ä',
        '—Ç–æ–≤–∞—Ä—ã': '—Ç–æ–≤–∞—Ä—ã',
        '–¥—É—Ö–∏': '–¥—É—Ö–∏',
        '–∞—Ä–æ–º–∞—Ç': '–∞—Ä–æ–º–∞—Ç',
        '–∞—Ä–æ–º–∞—Ç—ã': '–∞—Ä–æ–º–∞—Ç—ã',
        '–ø–∞—Ä—Ñ—é–º': '–ø–∞—Ä—Ñ—é–º',
        '–º–∞—Å–ª–æ': '–º–∞—Å–ª–æ',
        '–º–∞—Å–ª–∞': '–º–∞—Å–ª–∞',
        '—Ñ–ª–∞–∫–æ–Ω': '—Ñ–ª–∞–∫–æ–Ω',
        '—Ñ–ª–∞–∫–æ–Ω—ã': '—Ñ–ª–∞–∫–æ–Ω—ã',
        '–±—Ä–µ–Ω–¥': '–±—Ä–µ–Ω–¥',
        '–±—Ä–µ–Ω–¥—ã': '–±—Ä–µ–Ω–¥—ã',
        '–∫–æ–ª–ª–µ–∫—Ü–∏—é': '–∫–æ–ª–ª–µ–∫—Ü–∏—è',
        '–∫–æ–ª–ª–µ–∫—Ü–∏—è': '–∫–æ–ª–ª–µ–∫—Ü–∏—è',
        '–Ω–æ–≤–∏–Ω–∫–∏': '–Ω–æ–≤–∏–Ω–∫–∏',
        '–Ω–æ–≤–∏–Ω–∫—É': '–Ω–æ–≤–∏–Ω–∫–∞',
        '—Å–∫–∏–¥–∫–∏': '—Å–∫–∏–¥–∫–∏',
        '—Å–∫–∏–¥–∫—É': '—Å–∫–∏–¥–∫–∞',
        '–∞–∫—Ü–∏–∏': '–∞–∫—Ü–∏–∏',
        '–∞–∫—Ü–∏—é': '–∞–∫—Ü–∏—è',
        '–æ—Ç–∑—ã–≤—ã': '–æ—Ç–∑—ã–≤—ã',
        '–æ—Ç–∑—ã–≤': '–æ—Ç–∑—ã–≤',
        '—Å—Ç–∞—Ç—å–∏': '—Å—Ç–∞—Ç—å–∏',
        '—Å—Ç–∞—Ç—å—é': '—Å—Ç–∞—Ç—å—è',
        '–∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é': '–∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è',
        '–æ–ø–∏—Å–∞–Ω–∏–µ': '–æ–ø–∏—Å–∞–Ω–∏–µ',
        '—Ö–∞—Ä–∞–∫—Ç–µ—Ä–∏—Å—Ç–∏–∫–∏': '—Ö–∞—Ä–∞–∫—Ç–µ—Ä–∏—Å—Ç–∏–∫–∏',
        '–ø–æ–¥—Ä–æ–±–Ω–æ—Å—Ç–∏': '–ø–æ–¥—Ä–æ–±–Ω–æ—Å—Ç–∏',
        '–¥–µ—Ç–∞–ª–∏': '–¥–µ—Ç–∞–ª–∏',
        '–¥–æ—Å—Ç–∞–≤–∫—É': '–¥–æ—Å—Ç–∞–≤–∫–∞',
        '–æ–ø–ª–∞—Ç—É': '–æ–ø–ª–∞—Ç–∞',
        '–∑–∞–∫–∞–∑': '–∑–∞–∫–∞–∑',
        '–∫–æ—Ä–∑–∏–Ω—É': '–∫–æ—Ä–∑–∏–Ω–∞',
        '–ø–æ–∫—É–ø–∫–∏': '–ø–æ–∫—É–ø–∫–∏',
        '–ø–æ–∫—É–ø–∫—É': '–ø–æ–∫—É–ø–∫–∞',
        
        # –†–æ–¥–∏—Ç–µ–ª—å–Ω—ã–π –ø–∞–¥–µ–∂ -> –ò–º–µ–Ω–∏—Ç–µ–ª—å–Ω—ã–π  
        '–ø—Ä–∞–π—Å–∞': '–ø—Ä–∞–π—Å',
        '–∫–∞—Ç–∞–ª–æ–≥–∞': '–∫–∞—Ç–∞–ª–æ–≥',
        '–º–∞–≥–∞–∑–∏–Ω–∞': '–º–∞–≥–∞–∑–∏–Ω',
        '—Å–∞–π—Ç–∞': '—Å–∞–π—Ç',
        '—Ç–æ–≤–∞—Ä–∞': '—Ç–æ–≤–∞—Ä',
        '–∞—Ä–æ–º–∞—Ç–∞': '–∞—Ä–æ–º–∞—Ç',
        '–ø–∞—Ä—Ñ—é–º–∞': '–ø–∞—Ä—Ñ—é–º',
        '–º–∞—Å–ª–∞': '–º–∞—Å–ª–æ',
        '—Ñ–ª–∞–∫–æ–Ω–∞': '—Ñ–ª–∞–∫–æ–Ω',
        '–±—Ä–µ–Ω–¥–∞': '–±—Ä–µ–Ω–¥',
        
        # –î–∞—Ç–µ–ª—å–Ω—ã–π –ø–∞–¥–µ–∂ -> –ò–º–µ–Ω–∏—Ç–µ–ª—å–Ω—ã–π
        '–ø—Ä–∞–π—Å—É': '–ø—Ä–∞–π—Å',
        '–∫–∞—Ç–∞–ª–æ–≥—É': '–∫–∞—Ç–∞–ª–æ–≥',
        '–º–∞–≥–∞–∑–∏–Ω—É': '–º–∞–≥–∞–∑–∏–Ω',
        '—Å–∞–π—Ç—É': '—Å–∞–π—Ç',
        '—Ç–æ–≤–∞—Ä—É': '—Ç–æ–≤–∞—Ä',
        '–∞—Ä–æ–º–∞—Ç—É': '–∞—Ä–æ–º–∞—Ç',
        
        # –¢–≤–æ—Ä–∏—Ç–µ–ª—å–Ω—ã–π –ø–∞–¥–µ–∂ -> –ò–º–µ–Ω–∏—Ç–µ–ª—å–Ω—ã–π
        '–ø—Ä–∞–π—Å–æ–º': '–ø—Ä–∞–π—Å',
        '–∫–∞—Ç–∞–ª–æ–≥–æ–º': '–∫–∞—Ç–∞–ª–æ–≥',
        '–º–∞–≥–∞–∑–∏–Ω–æ–º': '–º–∞–≥–∞–∑–∏–Ω',
        '—Å–∞–π—Ç–æ–º': '—Å–∞–π—Ç',
        '—Ç–æ–≤–∞—Ä–æ–º': '—Ç–æ–≤–∞—Ä',
        '–∞—Ä–æ–º–∞—Ç–æ–º': '–∞—Ä–æ–º–∞—Ç',
        
        # –ü—Ä–µ–¥–ª–æ–∂–Ω—ã–π –ø–∞–¥–µ–∂ -> –ò–º–µ–Ω–∏—Ç–µ–ª—å–Ω—ã–π
        '–ø—Ä–∞–π—Å–µ': '–ø—Ä–∞–π—Å',
        '–∫–∞—Ç–∞–ª–æ–≥–µ': '–∫–∞—Ç–∞–ª–æ–≥', 
        '–º–∞–≥–∞–∑–∏–Ω–µ': '–º–∞–≥–∞–∑–∏–Ω',
        '—Å–∞–π—Ç–µ': '—Å–∞–π—Ç',
        '—Ç–æ–≤–∞—Ä–µ': '—Ç–æ–≤–∞—Ä',
        '–∞—Ä–æ–º–∞—Ç–µ': '–∞—Ä–æ–º–∞—Ç',
        
        # –ß–∞—Å—Ç—ã–µ –≥–æ—Ç–æ–≤—ã–µ —Ñ—Ä–∞–∑—ã
        '–ø–æ–¥—Ä–æ–±–Ω–µ–µ': '–ø–æ–¥—Ä–æ–±–Ω–µ–µ',
        '–±–æ–ª—å—à–µ': '–±–æ–ª—å—à–µ',
        '–¥–∞–ª–µ–µ': '–¥–∞–ª–µ–µ',
        '—á–∏—Ç–∞—Ç—å': '—á–∏—Ç–∞—Ç—å',
        '—Å–º–æ—Ç—Ä–µ—Ç—å': '—Å–º–æ—Ç—Ä–µ—Ç—å',
        '–ø–µ—Ä–µ–π—Ç–∏': '–ø–µ—Ä–µ–π—Ç–∏',
        '–ø–æ—Å–º–æ—Ç—Ä–µ—Ç—å': '–ø–æ—Å–º–æ—Ç—Ä–µ—Ç—å',
        '—É–∑–Ω–∞—Ç—å': '—É–∑–Ω–∞—Ç—å',
        '–≤—ã–±—Ä–∞—Ç—å': '–≤—ã–±—Ä–∞—Ç—å',
        '–∫—É–ø–∏—Ç—å': '–∫—É–ø–∏—Ç—å',
        '–∑–∞–∫–∞–∑–∞—Ç—å': '–∑–∞–∫–∞–∑–∞—Ç—å',
        '–æ—Ñ–æ—Ä–º–∏—Ç—å': '–æ—Ñ–æ—Ä–º–∏—Ç—å'
    }
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ç–æ—á–Ω–æ–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–µ –≤ —Å–ª–æ–≤–∞—Ä–µ
    if text in nominative_dict:
        result = nominative_dict[text]
    else:
        # –ë–∞–∑–æ–≤—ã–µ –ø—Ä–∞–≤–∏–ª–∞ –¥–ª—è –æ–∫–æ–Ω—á–∞–Ω–∏–π
        if text.endswith('—É—é'):
            result = text[:-2] + '–∞—è'
        elif text.endswith('–∏—é'):
            result = text[:-2] + '–∏—è'  
        elif text.endswith('—É—é'):
            result = text[:-2] + '–∞—è'
        elif text.endswith('–æ–π'):
            result = text[:-2] + '—ã–π'
        elif text.endswith('–µ–π'):
            result = text[:-2] + '–∏–π'
        elif text.endswith('–æ–º'):
            result = text[:-2]
        elif text.endswith('–µ–º'):
            result = text[:-2]
        elif text.endswith('–∞–º–∏'):
            result = text[:-3] + '—ã'
        elif text.endswith('—è–º–∏'):
            result = text[:-3] + '–∏'
        else:
            result = text
    
    # –î–µ–ª–∞–µ–º –ø–µ—Ä–≤—É—é –±—É–∫–≤—É –∑–∞–≥–ª–∞–≤–Ω–æ–π
    return result.capitalize()

def extract_links_from_text(text):
    """–ò–∑–≤–ª–µ–∫–∞–µ—Ç —Å—Å—ã–ª–∫–∏ –∏–∑ HTML-—Ç–µ–∫—Å—Ç–∞ –∏ —Å–æ–∑–¥–∞–µ—Ç –∫–Ω–æ–ø–∫–∏"""
    # –ò—â–µ–º —Å—Å—ã–ª–∫–∏ –≤ —Ñ–æ—Ä–º–∞—Ç–µ <a href='URL'>–¢–ï–ö–°–¢</a>
    link_pattern = r"<a\s+href=['\"]([^'\"]+)['\"][^>]*>([^<]+)</a>"
    links = re.findall(link_pattern, text)
    
    if not links:
        return None
    
    # –°–æ–∑–¥–∞–µ–º –∫–Ω–æ–ø–∫–∏ –¥–ª—è –∫–∞–∂–¥–æ–π —Å—Å—ã–ª–∫–∏
    buttons = []
    for url, button_text in links:
        # –ü—Ä–∏–≤–æ–¥–∏–º —Ç–µ–∫—Å—Ç –∫–Ω–æ–ø–∫–∏ –∫ –∏–º–µ–Ω–∏—Ç–µ–ª—å–Ω–æ–º—É –ø–∞–¥–µ–∂—É
        button_text_nominative = convert_to_nominative_case(button_text)
        buttons.append([{"text": button_text_nominative, "url": url}])
    
    return {"inline_keyboard": buttons}

def remove_html_links(text):
    """–£–¥–∞–ª—è–µ—Ç HTML-—Å—Å—ã–ª–∫–∏ –∏–∑ —Ç–µ–∫—Å—Ç–∞, –æ—Å—Ç–∞–≤–ª—è—è —Ç–æ–ª—å–∫–æ —Ç–µ–∫—Å—Ç"""
    # –£–¥–∞–ª—è–µ–º —Å—Å—ã–ª–∫–∏ –≤ —Ñ–æ—Ä–º–∞—Ç–µ <a href='URL'>–¢–ï–ö–°–¢</a>, –æ—Å—Ç–∞–≤–ª—è—è —Ç–æ–ª—å–∫–æ –¢–ï–ö–°–¢
    link_pattern = r"<a\s+href=['\"][^'\"]+['\"][^>]*>([^<]+)</a>"
    return re.sub(link_pattern, r"\1", text)

# --- Telegram webhook endpoint ---
print('=== [LOG] –û–±—ä—è–≤–ª—è—é —ç–Ω–¥–ø–æ–∏–Ω—Ç webhook... ===')
@app.post("/webhook/ai-bear-123456")
async def telegram_webhook(update: dict, request: Request):
    logger.info(f"=== WEBHOOK CALLED ===")
    logger.info(f"Request from: {request.client.host}")
    logger.info(f"Update type: {list(update.keys()) if update else 'None'}")
    
    try:
        result = await telegram_webhook_impl(update, request)
        logger.info(f"=== WEBHOOK COMPLETED SUCCESSFULLY ===")
        return result
    except Exception as e:
        logger.error(f"=== WEBHOOK FAILED: {e} ===")
        logger.error(traceback.format_exc())
        return {"ok": False, "error": str(e)}

# --- –ü–µ—Ä–µ–Ω–æ—Å–∏–º –≤–∞—à—É –ª–æ–≥–∏–∫—É webhook —Å—é–¥–∞ ---
async def telegram_webhook_impl(update: dict, request: Request):
    print(f'[WEBHOOK] Called: {request.url} from {request.client.host}')
    print(f'[WEBHOOK] Body: {update}')
    logger.info(f"[WEBHOOK] Called: {request.url} from {request.client.host}")
    logger.info(f"[WEBHOOK] Body: {update}")
    try:
        if "message" in update:
            print('[WEBHOOK] message detected')
            message = update["message"]
            chat_id = message["chat"]["id"]
            user_id = message["from"]["id"]
            text = message.get("text", "").strip()
            voice = message.get("voice")
            state = get_user_state(user_id)
            logger.info(f"[TG] user_id: {user_id}, text: {text}, state: {state}")
            
            try:
                # –û–±–Ω–æ–≤–ª—è–µ–º –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –¥–ª—è –ª—é–±–æ–≥–æ —Å–æ–æ–±—â–µ–Ω–∏—è
                first_name = message["from"].get("first_name")
                username = message["from"].get("username")
                add_user_to_db(user_id, chat_id, first_name, username)
                
                # –û–±—Ä–∞–±–æ—Ç–∫–∞ –≥–æ–ª–æ—Å–æ–≤—ã—Ö —Å–æ–æ–±—â–µ–Ω–∏–π
                if voice:
                    logger.info(f"[TG] Voice message received from {user_id}")
                    await send_typing_action(chat_id)
                    
                    file_id = voice["file_id"]
                    file_unique_id = voice["file_unique_id"]
                    duration = voice.get("duration", 0)
                    file_size = voice.get("file_size", 0)
                    
                    # –ü—Ä–æ–≤–µ—Ä–∫–∞ –º–∞–∫—Å–∏–º–∞–ª—å–Ω–æ–π –¥–ª–∏—Ç–µ–ª—å–Ω–æ—Å—Ç–∏ (1 —á–∞—Å)
                    if duration > 3600:
                        await telegram_send_message(chat_id, 
                            "üéôÔ∏è –ì–æ–ª–æ—Å–æ–≤–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ —Å–ª–∏—à–∫–æ–º –¥–ª–∏–Ω–Ω–æ–µ (–±–æ–ª—å—à–µ 1 —á–∞—Å–∞). "
                            "–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –∑–∞–ø–∏—à–∏—Ç–µ –±–æ–ª–µ–µ –∫–æ—Ä–æ—Ç–∫–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ –∏–ª–∏ –Ω–∞–ø–∏—à–∏—Ç–µ —Ç–µ–∫—Å—Ç.")
                        return {"ok": True}
                    
                    # –ü—Ä–æ–≤–µ—Ä–∫–∞ —Ä–∞–∑–º–µ—Ä–∞ —Ñ–∞–π–ª–∞ (50MB –º–∞–∫—Å–∏–º—É–º)
                    if file_size > 50 * 1024 * 1024:
                        await telegram_send_message(chat_id, 
                            "üéôÔ∏è –ì–æ–ª–æ—Å–æ–≤–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ —Å–ª–∏—à–∫–æ–º –±–æ–ª—å—à–æ–µ (–±–æ–ª—å—à–µ 50MB). "
                            "–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –∑–∞–ø–∏—à–∏—Ç–µ –±–æ–ª–µ–µ –∫–æ—Ä–æ—Ç–∫–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ.")
                        return {"ok": True}
                    
                    try:
                        file_url = f"https://api.telegram.org/bot{TOKEN}/getFile?file_id={file_id}"
                        
                        # –£—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º —Ç–∞–π–º–∞—É—Ç—ã –¥–ª—è –¥–ª–∏–Ω–Ω—ã—Ö —Ñ–∞–π–ª–æ–≤
                        timeout = httpx.Timeout(
                            connect=10.0,
                            read=300.0,  # –£–≤–µ–ª–∏—á–µ–Ω–Ω—ã–π —Ç–∞–π–º–∞—É—Ç –¥–ª—è —á—Ç–µ–Ω–∏—è –æ—á–µ–Ω—å –±–æ–ª—å—à–∏—Ö —Ñ–∞–π–ª–æ–≤ (–¥–æ —á–∞—Å–∞)
                            write=10.0,
                            pool=10.0
                        )
                        
                        async with httpx.AsyncClient(timeout=timeout) as client:
                            resp = await client.get(file_url)
                            if resp.status_code != 200:
                                logger.error(f"Failed to get file info: {resp.status_code}")
                                await telegram_send_message(chat_id, "–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–æ–ª—É—á–µ–Ω–∏–∏ –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–∏ –æ –≥–æ–ª–æ—Å–æ–≤–æ–º —Ñ–∞–π–ª–µ.")
                                return {"ok": True}
                            
                            file_info = resp.json()
                            if not file_info.get("ok"):
                                logger.error(f"File info error: {file_info}")
                                await telegram_send_message(chat_id, "–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–æ–ª—É—á–µ–Ω–∏–∏ –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–∏ –æ –≥–æ–ª–æ—Å–æ–≤–æ–º —Ñ–∞–π–ª–µ.")
                                return {"ok": True}
                            
                            file_path = file_info["result"]["file_path"]
                            download_url = f"https://api.telegram.org/file/bot{TOKEN}/{file_path}"
                            
                            logger.info(f"[TG] Downloading voice file: duration={duration}s, size={file_size}bytes")
                            
                            async with client.stream("GET", download_url) as response:
                                if response.status_code != 200:
                                    logger.error(f"Failed to download file: {response.status_code}")
                                    await telegram_send_message(chat_id, "–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–∫–∞—á–∏–≤–∞–Ω–∏–∏ –≥–æ–ª–æ—Å–æ–≤–æ–≥–æ —Ñ–∞–π–ª–∞.")
                                    return {"ok": True}
                                
                                # –ß–∏—Ç–∞–µ–º —Ñ–∞–π–ª –ø–æ —á–∞—Å—Ç—è–º –¥–ª—è –±–æ–ª—å—à–∏—Ö —Ñ–∞–π–ª–æ–≤
                                file_content = await response.aread()
                                
                                logger.info(f"[TG] Voice file downloaded, recognizing speech...")
                                
                                # –£–≤–µ–¥–æ–º–ª—è–µ–º –æ –Ω–∞—á–∞–ª–µ –æ–±—Ä–∞–±–æ—Ç–∫–∏ –¥–ª—è –¥–ª–∏–Ω–Ω—ã—Ö —Ñ–∞–π–ª–æ–≤
                                if duration > 120:
                                    minutes = duration // 60
                                    seconds = duration % 60
                                    duration_str = f"{minutes}–º {seconds}—Å" if minutes > 0 else f"{duration}—Å"
                                    await send_progress_message(chat_id, 
                                        f"üéôÔ∏è –û–±—Ä–∞–±–∞—Ç—ã–≤–∞—é –¥–ª–∏–Ω–Ω–æ–µ –≥–æ–ª–æ—Å–æ–≤–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ ({duration_str}). "
                                        "–≠—Ç–æ –º–æ–∂–µ—Ç –∑–∞–Ω—è—Ç—å –Ω–µ–∫–æ—Ç–æ—Ä–æ–µ –≤—Ä–µ–º—è...")
                                
                                text_content = await recognize_voice_content(file_content, chat_id if duration > 120 else None)
                                logger.info(f"[TG] Voice recognized text: {text_content[:100]}...")
                                
                                if text_content and not any(err in text_content for err in ["–û—à–∏–±–∫–∞", "–ù–µ —É–¥–∞–ª–æ—Å—å", "–Ω–µ–¥–æ—Å—Ç—É–ø–Ω–æ"]):
                                    ai_answer = await ask_chatgpt(text_content)
                                    ai_answer = ai_answer.replace('*', '')
                                    buttons = extract_links_from_text(ai_answer)
                                    ai_answer_clean = remove_html_links(ai_answer)
                                    success = await telegram_send_message(chat_id, ai_answer_clean, buttons if buttons else None)
                                    if success:
                                        logger.info(f"[TG] Sent AI answer to voice message for {chat_id}")
                                    else:
                                        logger.error(f"[TG] Failed to send AI answer to voice message for {chat_id}")
                                else:
                                    await telegram_send_message(chat_id, text_content)
                                    
                    except httpx.TimeoutException:
                        logger.error("Voice file download timeout")
                        await telegram_send_message(chat_id, 
                            "‚è∞ –í—Ä–µ–º—è –æ–±—Ä–∞–±–æ—Ç–∫–∏ –≥–æ–ª–æ—Å–æ–≤–æ–≥–æ —Å–æ–æ–±—â–µ–Ω–∏—è –∏—Å—Ç–µ–∫–ª–æ. "
                            "–ü–æ–ø—Ä–æ–±—É–π—Ç–µ –∑–∞–ø–∏—Å–∞—Ç—å –±–æ–ª–µ–µ –∫–æ—Ä–æ—Ç–∫–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ –∏–ª–∏ –Ω–∞–ø–∏—à–∏—Ç–µ —Ç–µ–∫—Å—Ç.")
                    except Exception as voice_error:
                        logger.error(f"Voice processing error: {voice_error}\n{traceback.format_exc()}")
                        await telegram_send_message(chat_id, 
                            "–ü—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ –≥–æ–ª–æ—Å–æ–≤–æ–≥–æ —Å–æ–æ–±—â–µ–Ω–∏—è. "
                            "–ü–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â–µ —Ä–∞–∑ –∏–ª–∏ –Ω–∞–ø–∏—à–∏—Ç–µ —Ç–µ–∫—Å—Ç.")
                    
                    return {"ok": True}
                
                if text == "/start":
                    # –î–æ–±–∞–≤–ª—è–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ –±–∞–∑—É –¥–∞–Ω–Ω—ã—Ö
                    first_name = message["from"].get("first_name")
                    username = message["from"].get("username")
                    add_user_to_db(user_id, chat_id, first_name, username)
                    
                    welcome = (
                        '<b>–ó–¥—Ä–∞–≤—Å—Ç–≤—É–π—Ç–µ!\n\n'
                        '–Ø ‚Äî –≤–∞—à –∞—Ä–æ–º–∞—Ç–Ω—ã–π –ø–æ–º–æ—â–Ω–∏–∫ –æ—Ç BAHUR.\n'
                        'üçì –ò—â—É –Ω–æ—Ç—ã –∏ üêÜ –æ—Ç–≤–µ—á–∞—é –Ω–∞ –≤–æ–ø—Ä–æ—Å—ã —Å –ª—é–±–æ–≤—å—é. ‚ù§\n\n'
                        'üí° <i>–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ /menu –¥–ª—è –≤–æ–∑–≤—Ä–∞—Ç–∞ –≤ –≥–ª–∞–≤–Ω–æ–µ –º–µ–Ω—é</i></b>'
                    )
                    main_menu = {
                        "inline_keyboard": [
                            [{"text": "üêÜ Ai-–ü–∞–Ω—Ç–µ—Ä–∞", "callback_data": "ai"}],
                            [
                                {"text": "üç¶ –ü—Ä–∞–π—Å", "url": "https://drive.google.com/file/d/1J70LlZwh6g7JOryDG2br-weQrYfv6zTc/view?usp=sharing"},
                                {"text": "üçø –ú–∞–≥–∞–∑–∏–Ω", "url": "https://www.bahur.store/m/"},
                                {"text": "‚ôæÔ∏è –í–æ–ø—Ä–æ—Å—ã", "url": "https://vk.com/@bahur_store-optovye-praisy-ot-bahur"}
                            ],
                            [
                                {"text": "üéÆ –ß–∞—Ç", "url": "https://t.me/+VYDZEvbp1pce4KeT"},
                                {"text": "üíé –°—Ç–∞—Ç—å–∏", "url": "https://vk.com/bahur_store?w=app6326142_-133936126%2523w%253Dapp6326142_-133936126"},
                                {"text": "üèÜ –û—Ç–∑—ã–≤—ã", "url": "https://vk.com/@bahur_store"}
                            ],
                            [{"text": "üçì –ù–æ—Ç—ã", "callback_data": "instruction"}]
                        ]
                    }
                    success = await telegram_send_message(chat_id, welcome, main_menu)
                    if success:
                        logger.info(f"[TG] Sent welcome to {chat_id}")
                    else:
                        logger.error(f"[TG] Failed to send welcome to {chat_id}")
                    set_user_state(user_id, None)  # –°–±—Ä–∞—Å—ã–≤–∞–µ–º —Å–æ—Å—Ç–æ—è–Ω–∏–µ –ø—Ä–∏ /start
                    return {"ok": True}
                elif text == "/menu":
                    # –ö–æ–º–∞–Ω–¥–∞ –¥–ª—è –≤—ã—Ö–æ–¥–∞ –∏–∑ —Ä–µ–∂–∏–º–∞ AI –∏ –≤–æ–∑–≤—Ä–∞—Ç–∞ –≤ –≥–ª–∞–≤–Ω–æ–µ –º–µ–Ω—é
                    welcome = (
                        '<b>–ó–¥—Ä–∞–≤—Å—Ç–≤—É–π—Ç–µ!\n\n'
                        '–Ø ‚Äî –≤–∞—à –∞—Ä–æ–º–∞—Ç–Ω—ã–π –ø–æ–º–æ—â–Ω–∏–∫ –æ—Ç BAHUR.\n'
                        'üçì –ò—â—É –Ω–æ—Ç—ã –∏ üêæ –æ—Ç–≤–µ—á–∞—é –Ω–∞ –≤–æ–ø—Ä–æ—Å—ã —Å –ª—é–±–æ–≤—å—é. ‚ù§\n\n'
                        'üí° <i>–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ /menu –¥–ª—è –≤–æ–∑–≤—Ä–∞—Ç–∞ –≤ –≥–ª–∞–≤–Ω–æ–µ –º–µ–Ω—é</i></b>'
                    )
                    main_menu = {
                        "inline_keyboard": [
                            [{"text": "üêÜ AI-–ü–∞–Ω—Ç–µ—Ä–∞", "callback_data": "ai"}],
                            [
                                {"text": "üç¶ –ü—Ä–∞–π—Å", "url": "https://drive.google.com/file/d/1J70LlZwh6g7JOryDG2br-weQrYfv6zTc/view?usp=sharing"},
                                {"text": "üçø –ú–∞–≥–∞–∑–∏–Ω", "url": "https://www.bahur.store/m/"},
                                {"text": "‚ôæÔ∏è –í–æ–ø—Ä–æ—Å—ã", "url": "https://vk.com/@bahur_store-optovye-praisy-ot-bahur"}
                            ],
                            [
                                {"text": "üéÆ –ß–∞—Ç", "url": "https://t.me/+VYDZEvbp1pce4KeT"},
                                {"text": "üíé –°—Ç–∞—Ç—å–∏", "url": "https://vk.com/bahur_store?w=app6326142_-133936126%2523w%253Dapp6326142_-133936126"},
                                {"text": "üèÜ –û—Ç–∑—ã–≤—ã", "url": "https://vk.com/@bahur_store"}
                            ],
                            [{"text": "üçì –ù–æ—Ç—ã", "callback_data": "instruction"}]
                        ]
                    }
                    success = await telegram_send_message(chat_id, welcome, main_menu)
                    if success:
                        logger.info(f"[TG] Sent menu to {chat_id}")
                    else:
                        logger.error(f"[TG] Failed to send menu to {chat_id}")
                    set_user_state(user_id, None)  # –°–±—Ä–∞—Å—ã–≤–∞–µ–º —Å–æ—Å—Ç–æ—è–Ω–∏–µ
                    return {"ok": True}
                if state == 'awaiting_ai_question':
                    logger.info(f"[TG] Processing AI question for user {user_id}")
                    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º –∏–Ω–¥–∏–∫–∞—Ç–æ—Ä "–ø–µ—á–∞—Ç–∞–µ—Ç"
                    await send_typing_action(chat_id)
                    ai_answer = await ask_chatgpt(text)
                    ai_answer = ai_answer.replace('*', '')
                    
                    # –ò–∑–≤–ª–µ–∫–∞–µ–º —Å—Å—ã–ª–∫–∏ –∏–∑ –æ—Ç–≤–µ—Ç–∞ –∏ —Å–æ–∑–¥–∞–µ–º –∫–Ω–æ–ø–∫–∏
                    buttons = extract_links_from_text(ai_answer)
                    ai_answer_clean = remove_html_links(ai_answer)
                    
                    success = await telegram_send_message(chat_id, ai_answer_clean, buttons if buttons else None)
                    if success:
                        logger.info(f"[TG] Sent ai_answer to {chat_id}")
                    else:
                        logger.error(f"[TG] Failed to send ai_answer to {chat_id}")
                    # –ù–ï —Å–±—Ä–∞—Å—ã–≤–∞–µ–º —Å–æ—Å—Ç–æ—è–Ω–∏–µ - –æ—Å—Ç–∞–µ–º—Å—è –≤ —Ä–µ–∂–∏–º–µ AI
                    return {"ok": True}
                if state == 'awaiting_note_search':
                    logger.info(f"[TG] Processing note search for user {user_id}")
                    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º –∏–Ω–¥–∏–∫–∞—Ç–æ—Ä "–ø–µ—á–∞—Ç–∞–µ—Ç"
                    await send_typing_action(chat_id)
                    result = await search_note_api(text)
                    if result.get("status") == "success":
                        msg = f'‚ú® {result.get("brand")} {result.get("aroma")}\n\n{result.get("description")}'
                        # –î–æ–±–∞–≤–ª—è–µ–º –∫–Ω–æ–ø–∫–∏ "–ü–æ–¥—Ä–æ–±–Ω–µ–µ" –∏ "–ü–æ–≤—Ç–æ—Ä–∏—Ç—å"
                        reply_markup = {
                            "inline_keyboard": [
                                [
                                    {"text": "üöÄ –ü–æ–¥—Ä–æ–±–Ω–µ–µ", "url": result.get("url", "")},
                                    {"text": "‚ôæÔ∏è –ü–æ–≤—Ç–æ—Ä–∏—Ç—å", "callback_data": f"repeatapi_{result.get('ID', '')}"}
                                ]
                            ]
                        }
                        success = await telegram_send_message(chat_id, msg, reply_markup)
                        if success:
                            logger.info(f"[TG] Sent note result to {chat_id}")
                        else:
                            logger.error(f"[TG] Failed to send note result to {chat_id}")
                    else:
                        success = await telegram_send_message(chat_id, "–ù–∏—á–µ–≥–æ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ –ø–æ —ç—Ç–æ–π –Ω–æ—Ç–µ üò¢")
                        if success:
                            logger.info(f"[TG] Sent not found to {chat_id}")
                        else:
                            logger.error(f"[TG] Failed to send not found to {chat_id}")
                    set_user_state(user_id, None)  # –°–±—Ä–∞—Å—ã–≤–∞–µ–º —Å–æ—Å—Ç–æ—è–Ω–∏–µ
                    return {"ok": True}
                # –ï—Å–ª–∏ –Ω–µ—Ç —Å–æ—Å—Ç–æ—è–Ω–∏—è, –ø—Ä–æ–≤–µ—Ä—è–µ–º, –ø–æ—Ö–æ–∂ –ª–∏ —Ç–µ–∫—Å—Ç –Ω–∞ –Ω–æ—Ç—É
                if is_likely_note(text):
                    logger.info(f"[TG] Text '{text}' looks like a note, searching...")
                    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º –∏–Ω–¥–∏–∫–∞—Ç–æ—Ä "–ø–µ—á–∞—Ç–∞–µ—Ç"
                    await send_typing_action(chat_id)
                    result = await search_note_api(text)
                    if result.get("status") == "success":
                        msg = f'‚ú® {result.get("brand")} {result.get("aroma")}\n\n{result.get("description")}'
                        # –î–æ–±–∞–≤–ª—è–µ–º –∫–Ω–æ–ø–∫–∏ "–ü–æ–¥—Ä–æ–±–Ω–µ–µ" –∏ "–ü–æ–≤—Ç–æ—Ä–∏—Ç—å"
                        reply_markup = {
                            "inline_keyboard": [
                                [
                                    {"text": "üöÄ –ü–æ–¥—Ä–æ–±–Ω–µ–µ", "url": result.get("url", "")},
                                    {"text": "‚ôæÔ∏è –ü–æ–≤—Ç–æ—Ä–∏—Ç—å", "callback_data": f"repeatapi_{result.get('ID', '')}"}
                                ]
                            ]
                        }
                        success = await telegram_send_message(chat_id, msg, reply_markup)
                        if success:
                            logger.info(f"[TG] Auto-found note result for {chat_id}")
                        else:
                            logger.error(f"[TG] Failed to send auto-found note result to {chat_id}")
                    else:
                        success = await telegram_send_message(chat_id, f"–ü–æ –∑–∞–ø—Ä–æ—Å—É '{text}' –Ω–∏—á–µ–≥–æ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ üò¢\n\n–ü–æ–ø—Ä–æ–±—É–π—Ç–µ –¥—Ä—É–≥–∏–µ –Ω–æ—Ç—ã –∏–ª–∏ –≤—ã–±–µ—Ä–∏—Ç–µ —Ä–µ–∂–∏–º –ø–æ–∏—Å–∫–∞.")
                        if success:
                            logger.info(f"[TG] Sent auto-search not found to {chat_id}")
                        else:
                            logger.error(f"[TG] Failed to send auto-search not found to {chat_id}")
                else:
                    # –ï—Å–ª–∏ –Ω–µ –ø–æ—Ö–æ–∂–µ –Ω–∞ –Ω–æ—Ç—É, –ø—Ä–µ–¥–ª–∞–≥–∞–µ–º –≤—ã–±—Ä–∞—Ç—å —Ä–µ–∂–∏–º
                    menu = {
                        "inline_keyboard": [
                            [{"text": "üêÜ AI-–ü–∞–Ω—Ç–µ—Ä–∞", "callback_data": "ai"}],
                            [{"text": "üçì –ù–æ—Ç—ã", "callback_data": "instruction"}]
                        ]
                    }
                    success = await telegram_send_message(chat_id, "–í—ã–±–µ—Ä–∏—Ç–µ —Ä–µ–∂–∏–º: üêÜ AI-–ü–∞–Ω—Ç–µ—Ä–∞ –∏–ª–∏ üçì –ù–æ—Ç—ã", reply_markup=menu)
                    if success:
                        logger.info(f"[TG] Sent menu to {chat_id}")
                    else:
                        logger.error(f"[TG] Failed to send menu to {chat_id}")
                set_user_state(user_id, None)  # –°–±—Ä–∞—Å—ã–≤–∞–µ–º —Å–æ—Å—Ç–æ—è–Ω–∏–µ
                return {"ok": True}
            except Exception as e:
                logger.error(f"[TG] Exception in message processing: {e}\n{traceback.format_exc()}")
                try:
                    await telegram_send_message(chat_id, "–ü—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ —Å–æ–æ–±—â–µ–Ω–∏—è. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â–µ —Ä–∞–∑.")
                except:
                    logger.error("Failed to send error message to user")
                return {"ok": False, "error": str(e)}
                
        elif "callback_query" in update:
            print('[WEBHOOK] callback_query detected')
            callback = update["callback_query"]
            data = callback["data"]
            chat_id = callback["message"]["chat"]["id"]
            user_id = callback["from"]["id"]
            message_id = callback["message"]["message_id"]
            callback_id = callback["id"]
            logger.info(f"[TG] Callback: {data} from {user_id}")
            
            try:
                # –û–±–Ω–æ–≤–ª—è–µ–º –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –¥–ª—è callback'–æ–≤
                first_name = callback["from"].get("first_name")
                username = callback["from"].get("username")
                add_user_to_db(user_id, chat_id, first_name, username)
                
                if data == "instruction":
                    set_user_state(user_id, 'awaiting_note_search')
                    success = await telegram_edit_message(chat_id, message_id, 'üçâ –ù–∞–ø–∏—à–∏ –ª—é–±—É—é –Ω–æ—Ç—É (–Ω–∞–ø—Ä–∏–º–µ—Ä, –∞–ø–µ–ª—å—Å–∏–Ω, –∫–ª—É–±–Ω–∏–∫–∞) ‚Äî —è –Ω–∞–π–¥—É –∞—Ä–æ–º–∞—Ç—ã —Å —ç—Ç–æ–π –Ω–æ—Ç–æ–π!')
                    if success:
                        logger.info(f"[TG] Set state awaiting_note_search for {user_id}")
                    else:
                        logger.error(f"[TG] Failed to edit instruction message for {chat_id}")
                    await telegram_answer_callback_query(callback_id)
                    return {"ok": True}
                elif data == "ai":
                    set_user_state(user_id, 'awaiting_ai_question')
                    ai_greeting = greet()
                    
                    # –ò–∑–≤–ª–µ–∫–∞–µ–º —Å—Å—ã–ª–∫–∏ –∏–∑ –ø—Ä–∏–≤–µ—Ç—Å—Ç–≤–∏—è –∏ —Å–æ–∑–¥–∞–µ–º –∫–Ω–æ–ø–∫–∏
                    buttons = extract_links_from_text(ai_greeting)
                    ai_greeting_clean = remove_html_links(ai_greeting)
                    
                    success = await telegram_edit_message(chat_id, message_id, ai_greeting_clean, buttons if buttons else None)
                    if success:
                        logger.info(f"[TG] Set state awaiting_ai_question for {user_id}")
                    else:
                        logger.error(f"[TG] Failed to edit ai greeting for {chat_id}")
                    await telegram_answer_callback_query(callback_id)
                    return {"ok": True}
                elif data.startswith("repeatapi_"):
                    aroma_id = data.split('_', 1)[1]
                    result = await search_by_id_api(aroma_id)
                    if result.get("status") == "success":
                        msg = f'‚ú® {result.get("brand")} {result.get("aroma")}\n\n{result.get("description")}'
                        # –î–æ–±–∞–≤–ª—è–µ–º –∫–Ω–æ–ø–∫–∏ –æ–±—Ä–∞—Ç–Ω–æ –ø—Ä–∏ –ø–æ–≤—Ç–æ—Ä–Ω–æ–º –ø–æ–∫–∞–∑–µ
                        reply_markup = {
                            "inline_keyboard": [
                                [
                                    {"text": "üöÄ –ü–æ–¥—Ä–æ–±–Ω–µ–µ", "url": result.get("url", "")},
                                    {"text": "‚ôæÔ∏è –ü–æ–≤—Ç–æ—Ä–∏—Ç—å", "callback_data": f"repeatapi_{result.get('ID', '')}"}
                                ]
                            ]
                        }
                        success = await telegram_edit_message(chat_id, message_id, msg, reply_markup)
                        if success:
                            logger.info(f"[TG] Edited repeatapi result for {chat_id}")
                        else:
                            logger.error(f"[TG] Failed to edit repeatapi result for {chat_id}")
                    else:
                        success = await telegram_edit_message(chat_id, message_id, "–ù–∏—á–µ–≥–æ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ –ø–æ —ç—Ç–æ–π –Ω–æ—Ç–µ üò¢")
                        if success:
                            logger.info(f"[TG] Edited repeatapi not found for {chat_id}")
                        else:
                            logger.error(f"[TG] Failed to edit repeatapi not found for {chat_id}")
                    await telegram_answer_callback_query(callback_id)
                    return {"ok": True}
                else:
                    success = await telegram_send_message(chat_id, "Callback –æ–±—Ä–∞–±–æ—Ç–∞–Ω.")
                    if success:
                        logger.info(f"[TG] Sent generic callback to {chat_id}")
                    else:
                        logger.error(f"[TG] Failed to send generic callback to {chat_id}")
                    return {"ok": True}
            except Exception as e:
                logger.error(f"[TG] Exception in callback processing: {e}\n{traceback.format_exc()}")
                try:
                    await telegram_send_message(chat_id, "–ü—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ callback. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â–µ —Ä–∞–∑.")
                except:
                    logger.error("Failed to send error message to user")
                return {"ok": False, "error": str(e)}
        else:
            print('[WEBHOOK] unknown update type')
            logger.warning("[TG] Unknown update type")
            return {"ok": False}
    except Exception as e:
        print(f'[WEBHOOK] Exception: {e}')
        logger.error(f"[TG] Exception in webhook: {e}\n{traceback.format_exc()}")
        # –ù–µ –ø—ã—Ç–∞–µ–º—Å—è –æ—Ç–ø—Ä–∞–≤–ª—è—Ç—å —Å–æ–æ–±—â–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é –∑–¥–µ—Å—å, —Ç–∞–∫ –∫–∞–∫ —É –Ω–∞—Å –Ω–µ—Ç chat_id
        return {"ok": False, "error": str(e)}
print('=== [LOG] –≠–Ω–¥–ø–æ–∏–Ω—Ç webhook –æ–±—ä—è–≤–ª–µ–Ω ===')

# --- –£—Å—Ç–∞–Ω–æ–≤–∫–∞ Telegram webhook ---
async def set_telegram_webhook(base_url: str):
    url = f"https://api.telegram.org/bot{TOKEN}/setWebhook"
    webhook_url = f"{base_url}{WEBHOOK_PATH}"
    async with httpx.AsyncClient() as client:
        resp = await client.post(url, data={"url": webhook_url})
        logger.info(f"Set webhook response: {resp.text}")
        return resp.json()

# --- –≠–Ω–¥–ø–æ–∏–Ω—Ç—ã FastAPI ---
@app.on_event("startup")
async def startup_event():
    logger.info("=== STARTUP EVENT ===")
    
    # –ó–∞–ø—É—Å–∫–∞–µ–º –ø–ª–∞–Ω–∏—Ä–æ–≤—â–∏–∫ –µ–∂–µ–Ω–µ–¥–µ–ª—å–Ω—ã—Ö —Å–æ–æ–±—â–µ–Ω–∏–π
    start_weekly_scheduler()
    
    base_url = os.getenv("WEBHOOK_BASE_URL")
    if not base_url:
        logger.warning("WEBHOOK_BASE_URL –Ω–µ –∑–∞–¥–∞–Ω, webhook –Ω–µ –±—É–¥–µ—Ç —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω!")
        return
    try:
        result = await set_telegram_webhook(base_url)
        logger.info(f"Webhook set result: {result}")
    except Exception as e:
        logger.error(f"Failed to set webhook: {e}\n{traceback.format_exc()}")
    logger.info("=== STARTUP EVENT COMPLETE ===")

@app.on_event("shutdown")
async def shutdown_event():
    logger.info("=== SHUTDOWN EVENT ===")
    logger.info("Application is shutting down gracefully...")
    logger.info("=== SHUTDOWN EVENT COMPLETE ===")

@app.get("/")
async def healthcheck():
    logger.info("Healthcheck requested")
    return PlainTextResponse("OK")

@app.post("/message")
async def handle_message(msg: MessageModel):
    user_id = msg.user_id
    text = msg.text.strip()
    ai_answer = await ask_chatgpt(text)
    ai_answer = ai_answer.replace('*', '')
    return JSONResponse({"answer": ai_answer, "parse_mode": "HTML"})

@app.post("/callback")
async def handle_callback(cb: CallbackModel):
    user_id = cb.user_id
    data = cb.data
    logger.info(f"[SUPERLOG] Callback data: {data}, user_id: {user_id}")
    try:
        if data != 'ai' and user_id in user_states:
            user_states.pop(user_id, None)
        if data == 'instruction':
            set_user_state(user_id, 'awaiting_note_search')
            return JSONResponse({"text": 'üçâ –ù–∞–ø–∏—à–∏ –ª—é–±—É—é –Ω–æ—Ç—É (–Ω–∞–ø—Ä–∏–º–µ—Ä, –∞–ø–µ–ª—å—Å–∏–Ω, –∫–ª—É–±–Ω–∏–∫–∞) ‚Äî —è –Ω–∞–π–¥—É –∞—Ä–æ–º–∞—Ç—ã —Å —ç—Ç–æ–π –Ω–æ—Ç–æ–π!'} )
        elif data == 'ai':
            set_user_state(user_id, 'awaiting_ai_question')
            result = greet()
            return JSONResponse({"text": result})
        elif data.startswith('repeatapi_'):
            aroma_id = data.split('_', 1)[1]
            result = await search_by_id_api(aroma_id)
            if result.get("status") == "success":
                return JSONResponse({
                    "brand": result.get("brand"),
                    "aroma": result.get("aroma"),
                    "description": result.get("description"),
                    "url": result.get("url"),
                    "aroma_id": result.get("ID")
                })
            else:
                return JSONResponse({"error": "–ù–∏—á–µ–≥–æ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ –ø–æ —ç—Ç–æ–π –Ω–æ—Ç–µ üò¢"})
        else:
            return JSONResponse({"info": "Callback –æ–±—Ä–∞–±–æ—Ç–∞–Ω."})
    except Exception as e:
        logger.error(f"[SUPERLOG] Exception in handle_callback: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/start")
async def cmd_start(msg: MessageModel):
    logger.info(f"/start command from user {msg.user_id}")
    
    # –î–æ–±–∞–≤–ª—è–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ –±–∞–∑—É –¥–∞–Ω–Ω—ã—Ö (chat_id = user_id –¥–ª—è API endpoint)
    add_user_to_db(msg.user_id, msg.user_id)
    
    text = (
        '<b>–ó–¥—Ä–∞–≤—Å—Ç–≤—É–π—Ç–µ!\n\n'
        '–Ø ‚Äî –≤–∞—à –∞—Ä–æ–º–∞—Ç–Ω—ã–π –ø–æ–º–æ—â–Ω–∏–∫ –æ—Ç BAHUR.\n'
                                'üçì –ò—â—É –Ω–æ—Ç—ã –∏ üêÜ –æ—Ç–≤–µ—á–∞—é –Ω–∞ –≤–æ–ø—Ä–æ—Å—ã —Å –ª—é–±–æ–≤—å—é. ‚ù§</b>'
    )
    return JSONResponse({"text": text, "parse_mode": "HTML"})

@app.post("/send-weekly-messages")
async def manual_weekly_send():
    """–†—É—á–Ω–æ–π –∑–∞–ø—É—Å–∫ –µ–∂–µ–Ω–µ–¥–µ–ª—å–Ω–æ–π —Ä–∞—Å—Å—ã–ª–∫–∏ (–¥–ª—è —Ç–µ—Å—Ç–∏—Ä–æ–≤–∞–Ω–∏—è)"""
    try:
        await send_weekly_messages()
        return JSONResponse({"status": "success", "message": "Weekly messages sent"})
    except Exception as e:
        logger.error(f"Manual weekly send failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/users-stats")
async def get_users_stats():
    """–ü–æ–ª—É—á–∏—Ç—å —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π"""
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        # –û–±—â–µ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
        cursor.execute('SELECT COUNT(*) FROM users')
        total_users = cursor.fetchone()[0]
        
        # –ê–∫—Ç–∏–≤–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏
        cursor.execute('SELECT COUNT(*) FROM users WHERE is_active = 1')
        active_users = cursor.fetchone()[0]
        
        # –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏, –¥–æ–±–∞–≤–ª–µ–Ω–Ω—ã–µ –∑–∞ –ø–æ—Å–ª–µ–¥–Ω—é—é –Ω–µ–¥–µ–ª—é
        week_ago = datetime.now() - timedelta(days=7)
        cursor.execute('SELECT COUNT(*) FROM users WHERE first_start_date > ?', (week_ago,))
        new_users_week = cursor.fetchone()[0]
        
        # –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏, –∫–æ—Ç–æ—Ä—ã–º –Ω—É–∂–Ω–æ –æ—Ç–ø—Ä–∞–≤–∏—Ç—å –µ–∂–µ–Ω–µ–¥–µ–ª—å–Ω–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ
        cursor.execute('''
            SELECT COUNT(*) FROM users 
            WHERE is_active = 1 AND (
                last_weekly_message IS NULL OR 
                last_weekly_message < ?
            )
        ''', (week_ago,))
        pending_weekly = cursor.fetchone()[0]
        
        conn.close()
        
        return JSONResponse({
            "total_users": total_users,
            "active_users": active_users,
            "new_users_this_week": new_users_week,
            "pending_weekly_messages": pending_weekly
        })
        
    except Exception as e:
        logger.error(f"Failed to get users stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/users-list")
async def get_users_list():
    """–ü–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π (—Ç–æ–ª—å–∫–æ –¥–ª—è –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–≤)"""
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT user_id, chat_id, first_name, username, first_start_date, 
                   last_activity, last_weekly_message, is_active 
            FROM users 
            ORDER BY first_start_date DESC 
            LIMIT 100
        ''')
        
        users = []
        for row in cursor.fetchall():
            users.append({
                "user_id": row[0],
                "chat_id": row[1],
                "first_name": row[2],
                "username": row[3],
                "first_start_date": row[4],
                "last_activity": row[5],
                "last_weekly_message": row[6],
                "is_active": bool(row[7])
            })
        
        conn.close()
        return JSONResponse({"users": users})
        
    except Exception as e:
        logger.error(f"Failed to get users list: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# --- Endpoints –¥–ª—è —Ä–∞–±–æ—Ç—ã —Å Excel –¥–∞–Ω–Ω—ã–º–∏ ---

@app.get("/products/search")
async def search_products_api(q: str, limit: int = 10):
    """–ü–æ–∏—Å–∫ —Ç–æ–≤–∞—Ä–æ–≤ –ø–æ –Ω–∞–∑–≤–∞–Ω–∏—é –±—Ä–µ–Ω–¥–∞ –∏–ª–∏ –∞—Ä–æ–º–∞—Ç–∞"""
    try:
        products = search_products(q, limit)
        return JSONResponse({
            "query": q,
            "count": len(products),
            "products": products
        })
    except Exception as e:
        logger.error(f"Search products error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/products/calculate-price")
async def calculate_price_api(brand: str, aroma: str, volume: float):
    """–†–∞—Å—Å—á–∏—Ç–∞—Ç—å —Ü–µ–Ω—É –¥–ª—è –∫–æ–Ω–∫—Ä–µ—Ç–Ω–æ–≥–æ —Ç–æ–≤–∞—Ä–∞ –∏ –æ–±—ä–µ–º–∞"""
    try:
        # –ò—â–µ–º —Ç–æ–≤–∞—Ä
        query = f"{brand} {aroma}"
        products = search_products(query, limit=1)
        
        if not products:
            return JSONResponse({
                "error": f"–¢–æ–≤–∞—Ä '{query}' –Ω–µ –Ω–∞–π–¥–µ–Ω",
                "suggestions": search_products(brand, limit=3)
            }, status_code=404)
        
        product = products[0]
        price_info = calculate_price(product, volume)
        
        if not price_info:
            return JSONResponse({
                "error": "–ù–µ —É–¥–∞–ª–æ—Å—å —Ä–∞—Å—Å—á–∏—Ç–∞—Ç—å —Ü–µ–Ω—É –¥–ª—è —É–∫–∞–∑–∞–Ω–Ω–æ–≥–æ –æ–±—ä–µ–º–∞",
                "product": product
            }, status_code=400)
        
        # –î–æ–±–∞–≤–ª—è–µ–º –æ–ø–∏—Å–∞–Ω–∏–µ –∫ –∫–∞—á–µ—Å—Ç–≤—É
        quality_raw = product.get('–ö–∞—á–µ—Å—Ç–≤–æ', 'N/A')
        quality_descriptions = {
            'TOP': 'TOP (–≤—ã—Å—à–µ–µ –∫–∞—á–µ—Å—Ç–≤–æ)',
            'Q1': 'Q1 (–æ—Ç–ª–∏—á–Ω–æ–µ –∫–∞—á–µ—Å—Ç–≤–æ)',
            'Q2': 'Q2 (—Ö–æ—Ä–æ—à–µ–µ –∫–∞—á–µ—Å—Ç–≤–æ)'
        }
        quality_with_desc = quality_descriptions.get(quality_raw, quality_raw)
        
        return JSONResponse({
            "product": {
                "brand": product.get('–ë—Ä–µ–Ω–¥'),
                "aroma": product.get('–ê—Ä–æ–º–∞—Ç'),
                "factory": product.get('–§–∞–±—Ä–∏–∫–∞'),
                "quality": quality_with_desc
            },
            "price_calculation": price_info
        })
        
    except Exception as e:
        logger.error(f"Price calculation error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/products/top")
async def get_top_products_api(
    factory: str = None, 
    quality: int = None, 
    sort_by: str = "TOP LAST", 
    limit: int = 10
):
    """–ü–æ–ª—É—á–∏—Ç—å —Ç–æ–ø —Ç–æ–≤–∞—Ä—ã –ø–æ –ø–æ–ø—É–ª—è—Ä–Ω–æ—Å—Ç–∏"""
    try:
        products = get_top_products(factory, quality, sort_by, limit)
        
        # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ –¥–ª—è API
        formatted_products = []
        quality_descriptions = {
            'TOP': 'TOP (–≤—ã—Å—à–µ–µ –∫–∞—á–µ—Å—Ç–≤–æ)',
            'Q1': 'Q1 (–æ—Ç–ª–∏—á–Ω–æ–µ –∫–∞—á–µ—Å—Ç–≤–æ)',
            'Q2': 'Q2 (—Ö–æ—Ä–æ—à–µ–µ –∫–∞—á–µ—Å—Ç–≤–æ)'
        }
        
        for product in products:
            quality_raw = product.get('–ö–∞—á–µ—Å—Ç–≤–æ', 'N/A')
            quality_with_desc = quality_descriptions.get(quality_raw, quality_raw)
            
            formatted_products.append({
                "brand": product.get('–ë—Ä–µ–Ω–¥'),
                "aroma": product.get('–ê—Ä–æ–º–∞—Ç'),
                "factory": product.get('–§–∞–±—Ä–∏–∫–∞'),
                "quality": quality_with_desc,
                "quality_code": product.get('–ö–∞—á–µ—Å—Ç–≤–æ'),
                "prices": {
                    "30_49_ml": product.get('30 GR'),
                    "50_499_ml": product.get('50 GR'),
                    "500_999_ml": product.get('500 GR'),
                    "1000_plus_ml": product.get('1 KG')
                },
                "popularity": {
                    "last_6_months": float(product.get('TOP LAST', 0)) * 100,
                    "all_time": float(product.get('TOP ALL', 0)) * 100
                }
            })
        
        return JSONResponse({
            "filters": {
                "factory": factory,
                "quality": quality,
                "sort_by": sort_by
            },
            "count": len(formatted_products),
            "products": formatted_products
        })
        
    except Exception as e:
        logger.error(f"Get top products error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/products/stats")
async def get_products_stats():
    """–ü–æ–ª—É—á–∏—Ç—å –æ–±—â—É—é —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É –ø–æ —Ç–æ–≤–∞—Ä–∞–º"""
    try:
        global excel_data
        if excel_data is None:
            load_excel_data()
        
        if excel_data is None:
            raise HTTPException(status_code=500, detail="Excel data not available")
        
        # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ —Ñ–∞–±—Ä–∏–∫–∞–º
        factory_stats = excel_data['–§–∞–±—Ä–∏–∫–∞'].value_counts().to_dict()
        
        # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ –∫–∞—á–µ—Å—Ç–≤—É
        quality_stats = excel_data['–ö–∞—á–µ—Å—Ç–≤–æ'].value_counts().to_dict()
        
        # –¶–µ–Ω–æ–≤—ã–µ –¥–∏–∞–ø–∞–∑–æ–Ω—ã
        price_stats = {}
        for col in ['30 GR', '50 GR', '500 GR', '1 KG']:
            if col in excel_data.columns:
                prices = excel_data[col].dropna()
                if len(prices) > 0:
                    price_stats[col] = {
                        "min": float(prices.min()),
                        "max": float(prices.max()),
                        "avg": float(prices.mean()),
                        "count": len(prices)
                    }
        
        return JSONResponse({
            "total_products": len(excel_data),
            "factories": factory_stats,
            "qualities": quality_stats,
            "price_ranges": price_stats,
            "last_updated": datetime.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Get products stats error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/products/reload")
async def reload_excel_data():
    """–ü–µ—Ä–µ–∑–∞–≥—Ä—É–∑–∏—Ç—å –¥–∞–Ω–Ω—ã–µ –∏–∑ Excel —Ñ–∞–π–ª–∞"""
    try:
        df = load_excel_data()
        if df is not None:
            return JSONResponse({
                "status": "success",
                "message": "Excel data reloaded successfully",
                "products_count": len(df)
            })
        else:
            raise HTTPException(status_code=500, detail="Failed to reload Excel data")
    except Exception as e:
        logger.error(f"Reload Excel data error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# --- –î–ª—è –∑–∞–ø—É—Å–∫–∞: uvicorn 1:app --reload ---
if __name__ == "__main__":
    import signal
    
    def signal_handler(signum, frame):
        logger.info("Received shutdown signal, gracefully shutting down...")
        sys.exit(0)
    
    # –†–µ–≥–∏—Å—Ç—Ä–∏—Ä—É–µ–º –æ–±—Ä–∞–±–æ—Ç—á–∏–∫–∏ —Å–∏–≥–Ω–∞–ª–æ–≤
    signal.signal(signal.SIGINT, signal_handler)
    signal.signal(signal.SIGTERM, signal_handler)
    
    port = int(os.environ.get("PORT", 8000))
    print(f"[INFO] Starting uvicorn on 0.0.0.0:{port}")
    uvicorn.run("1:app", host="0.0.0.0", port=port)

def format_aroma_response_improved(product, include_prices=True):
    """–£–ª—É—á—à–µ–Ω–Ω–∞—è —Ñ—É–Ω–∫—Ü–∏—è —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏—è –æ—Ç–≤–µ—Ç–∞ –æ–± –∞—Ä–æ–º–∞—Ç–µ —Å –ø—Ä–∞–≤–∏–ª—å–Ω—ã–º–∏ –æ—Ç—Å—Ç—É–ø–∞–º–∏"""
    try:
        brand = product.get('–ë—Ä–µ–Ω–¥', 'N/A')
        aroma_raw = product.get('–ê—Ä–æ–º–∞—Ç', 'N/A')
        aroma = format_aroma_name(aroma_raw)
        factory = product.get('–§–∞–±—Ä–∏–∫–∞', 'N/A')
        quality = product.get('–ö–∞—á–µ—Å—Ç–≤–æ', 'N/A')
        
        # –ü–æ–ª—É—á–∞–µ–º —Å—Å—ã–ª–∫—É (–≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∞ –∏–ª–∏ –æ–±—ã—á–Ω–∞—è)
        hyperlink = product.get('–ì–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∞', '')
        link = product.get('–°—Å—ã–ª–∫–∞', '')
        aroma_url = ""
        
        if hyperlink and not pd.isna(hyperlink) and str(hyperlink).strip().startswith('http'):
            aroma_url = str(hyperlink).strip()
        elif link and not pd.isna(link) and str(link).strip().startswith('http'):
            aroma_url = str(link).strip()
        
        # –§–æ—Ä–º–∏—Ä—É–µ–º –æ—Ç–≤–µ—Ç —Å –ø—Ä–∞–≤–∏–ª—å–Ω—ã–º–∏ –æ—Ç—Å—Ç—É–ø–∞–º–∏
        response = ""
        
        # –ù–∞–∑–≤–∞–Ω–∏–µ –∞—Ä–æ–º–∞—Ç–∞
        if aroma_url:
            response += f"‚ú® <a href='{aroma_url}'>{brand} {aroma}</a>\n\n"
        else:
            response += f"‚ú® {brand} {aroma}\n\n"
        
        # –ë—Ä–µ–Ω–¥ –∏ —Å—Ç—Ä–∞–Ω–∞
        response += f"¬Æ –ë—Ä–µ–Ω–¥: {brand}\n"
        country = product.get('–°—Ç—Ä–∞–Ω–∞', '')
        country_emoji = get_country_emoji(country)
        if country and not pd.isna(country) and str(country).strip():
            response += f"{country_emoji} –°—Ç—Ä–∞–Ω–∞: {str(country).strip()}\n"
        else:
            response += f"{country_emoji} –°—Ç—Ä–∞–Ω–∞: –ù–µ —É–∫–∞–∑–∞–Ω–∞\n"
        
        response += "\n"
        
        # –ù–æ—Ç—ã (–∏–∑ –ø—Ä–∞–π—Å–∞, –Ω–µ –∏–∑ API)
        top_notes = product.get('–í–µ—Ä—Ö–Ω–∏–µ –Ω–æ—Ç—ã', '')
        middle_notes = product.get('–°—Ä–µ–¥–Ω–∏–µ –Ω–æ—Ç—ã', '')
        base_notes = product.get('–ë–∞–∑–æ–≤—ã–µ –Ω–æ—Ç—ã', '')
        
        if top_notes and not pd.isna(top_notes) and str(top_notes).strip():
            response += f"üå± –í–µ—Ä—Ö–Ω–∏–µ –Ω–æ—Ç—ã: {str(top_notes).strip()}\n"
        else:
            response += f"üå± –í–µ—Ä—Ö–Ω–∏–µ –Ω–æ—Ç—ã: –ù–µ —É–∫–∞–∑–∞–Ω—ã\n"
            
        if middle_notes and not pd.isna(middle_notes) and str(middle_notes).strip():
            response += f"üåø –°—Ä–µ–¥–Ω–∏–µ –Ω–æ—Ç—ã: {str(middle_notes).strip()}\n"
        else:
            response += f"üåø –°—Ä–µ–¥–Ω–∏–µ –Ω–æ—Ç—ã: –ù–µ —É–∫–∞–∑–∞–Ω—ã\n"
            
        if base_notes and not pd.isna(base_notes) and str(base_notes).strip():
            response += f"üçÉ –ë–∞–∑–æ–≤—ã–µ –Ω–æ—Ç—ã: {str(base_notes).strip()}\n"
        else:
            response += f"üçÉ –ë–∞–∑–æ–≤—ã–µ –Ω–æ—Ç—ã: –ù–µ —É–∫–∞–∑–∞–Ω—ã\n"
        
        response += "\n"
        
        # –ü–æ–ø—É–ª—è—Ä–Ω–æ—Å—Ç—å
        top_last = product.get('TOP LAST', 0)
        top_all = product.get('TOP ALL', 0)
        if top_last and not pd.isna(top_last):
            response += f"‚ö°Ô∏è TOP LAST: {float(top_last):.2f}% (‚Ññ{get_rank(product, get_top_products(sort_by='TOP LAST', limit=None), lambda p: p.get('TOP LAST', 0))})\n"
        if top_all and not pd.isna(top_all):
            response += f"üöÄ TOP ALL: {float(top_all):.2f}% (‚Ññ{get_rank(product, get_top_products(sort_by='TOP ALL', limit=None), lambda p: p.get('TOP ALL', 0))})\n"
        
        response += "\n"
        
        # VERSION (–µ—Å–ª–∏ –µ—Å—Ç—å –≤–∞—Ä–∏–∞–Ω—Ç—ã)
        aroma_name = product.get('–ê—Ä–æ–º–∞—Ç', '')
        if aroma_name and not pd.isna(aroma_name):
            all_versions = [p for p in excel_data if p.get('–ê—Ä–æ–º–∞—Ç', '').strip().lower() == aroma_name.strip().lower()]
            if len(all_versions) > 1:
                total_popularity = sum(p.get('TOP LAST', 0) for p in all_versions)
                if total_popularity > 0:
                    factory_stats = {}
                    for version in all_versions:
                        factory = version.get('–§–∞–±—Ä–∏–∫–∞', '')
                        quality = version.get('–ö–∞—á–µ—Å—Ç–≤–æ', '')
                        popularity = version.get('TOP LAST', 0)
                        key = f"{factory} {quality}"
                        if key not in factory_stats:
                            factory_stats[key] = 0
                        factory_stats[key] += popularity
                    
                    version_percents = []
                    for factory_key, popularity in factory_stats.items():
                        percent = (popularity / total_popularity) * 100
                        version_percents.append(f"{factory_key}: {percent:.2f}%")
                    
                    if version_percents:
                        response += f"‚ôæÔ∏è VERSION: {' | '.join(version_percents)}\n\n"
        
        # –°—Ç–æ–∏–º–æ—Å—Ç—å
        if include_prices:
            response += f"üíµ –°—Ç–æ–∏–º–æ—Å—Ç—å:\n"
            price_ranges = [
                ('30 GR', 30, '30 –≥—Ä–∞–º–º'),
                ('50 GR', 50, '50 –≥—Ä–∞–º–º'),
                ('500 GR', 500, '500 –≥—Ä–∞–º–º'),
                ('1 KG', 1000, '1000 –≥—Ä–∞–º–º')
            ]
            for col, volume, volume_text in price_ranges:
                price_per_g = product.get(col)
                if price_per_g and not pd.isna(price_per_g):
                    total_price = int(price_per_g * volume)
                    response += f"üíß{volume_text} = {total_price}‚ÇΩ ({price_per_g}‚ÇΩ - –∑–∞ 1 –≥—Ä–∞–º–º)\n"
        
        return response.strip()
        
    except Exception as e:
        logger.error(f"Error formatting aroma response: {e}")
        return f"‚ùå –û—à–∏–±–∫–∞ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏—è –¥–∞–Ω–Ω—ã—Ö –æ –ø—Ä–æ–¥—É–∫—Ç–µ"