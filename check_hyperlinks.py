import pandas as pd
import requests
import io
import urllib3
from openpyxl import load_workbook

# Отключаем SSL предупреждения
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def check_hyperlinks():
    try:
        print("🔍 Проверяем гиперссылки в Excel...")
        
        # URL Google Sheets
        url = "https://docs.google.com/spreadsheets/d/1J70LlZwh6g7JOryDG2br-weQrYfv6zTc/export?format=xlsx"
        
        # Загружаем данные
        session = requests.Session()
        session.verify = False
        response = session.get(url, timeout=30)
        
        if response.status_code != 200:
            print(f"❌ Ошибка доступа: {response.status_code}")
            return
        
        # Сохраняем во временный файл для openpyxl
        with open('temp.xlsx', 'wb') as f:
            f.write(response.content)
        
        # Загружаем с помощью openpyxl для проверки гиперссылок
        wb = load_workbook('temp.xlsx')
        ws = wb.active
        
        print(f"✅ Excel загружен: {ws.max_row} строк, {ws.max_column} столбцов")
        
        # Проверяем гиперссылки в столбце B (индекс 2)
        print(f"\n🔗 Проверяем гиперссылки в столбце B:")
        hyperlinks_found = 0
        
        for row in range(4, min(15, ws.max_row + 1)):  # Начинаем с 4-й строки (после заголовков)
            cell = ws.cell(row=row, column=2)  # Столбец B
            
            if cell.hyperlink:
                print(f"✅ Строка {row}: {cell.hyperlink.target}")
                hyperlinks_found += 1
            elif cell.value:
                print(f"⚠️  Строка {row}: {cell.value} (текст без гиперссылки)")
            else:
                print(f"❌ Строка {row}: пустая ячейка")
        
        print(f"\n📊 Найдено гиперссылок: {hyperlinks_found}")
        
        # Также проверяем обычные данные через pandas
        df = pd.read_excel(io.BytesIO(response.content), header=2, skiprows=[3])
        print(f"\n📋 Pandas данные: {len(df)} строк, {len(df.columns)} столбцов")
        
        # Удаляем временный файл
        import os
        os.remove('temp.xlsx')
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")

if __name__ == "__main__":
    check_hyperlinks() 